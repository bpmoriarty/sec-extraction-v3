"""
build_spreadsheet.py — assemble the extracted per-filing JSONs into one analysis workbook.

Reads every JSON in data/extracted/ and writes data/dataset/semiliquid_bdc_dataset.xlsx with
four tabs:

  Data         — one row per filing (fund-period grain), ~65 fields grouped by section.
                 Flag-and-keep values are marked: a `status`/`flags` column, an amber tint on
                 any review row, AND a cell-level highlight on the specific fields tied to each
                 failing rule (C4 -> the fair-value cells, C5 -> NII, etc.).
  ShareClasses — one row per filing x share class (NAV / shares / net assets / yield).
  Review       — every filing with a validation failure + the rules + messages.
  Check (Gold) — a hand-verification view for ~15 representative filings: each filing's key
                 fields stacked tall, with provenance + blank verdict columns. Fill the
                 ✓ (Y/N) column against the actual SEC filing; the accuracy % computes itself.

Run:  uv run python src/output/build_spreadsheet.py
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
EXTRACTED = ROOT / "data" / "extracted"
HOLDINGS = ROOT / "data" / "holdings"
OUT = ROOT / "data" / "dataset" / "semiliquid_bdc_dataset.xlsx"

# Columns in the per-filing holdings CSVs (written by run_extraction._write_holdings).
HOLDING_CSV_COLS = ["issuer", "affiliation", "fair_value", "cost", "principal", "rate",
                    "spread", "pik_rate", "floor", "shares", "commitment", "pct_na"]

# ── Column spec for the Data tab ────────────────────────────────────────────────
# (section, column_name, section_obj, field) — section_obj is the top-level JSON key
# whose Fact lives at [section_obj][field]["value"]. Meta columns are handled separately.
DATA_FIELDS: list[tuple[str, str, str, str]] = [
    # Balance sheet (§2)
    ("Balance", "total_assets", "balance_sheet", "total_assets"),
    ("Balance", "total_liabilities", "balance_sheet", "total_liabilities"),
    ("Balance", "total_net_assets", "balance_sheet", "total_net_assets"),
    ("Balance", "investments_at_fair_value", "balance_sheet", "investments_at_fair_value"),
    ("Balance", "cash_and_equivalents", "balance_sheet", "cash_and_equivalents"),
    ("Balance", "total_debt", "balance_sheet", "total_debt"),
    # Balance-sheet detail (§2, Theme 4)
    ("Balance", "interest_receivable", "balance_sheet", "interest_receivable"),
    ("Balance", "receivable_for_investments", "balance_sheet", "receivable_for_investments"),
    ("Balance", "other_assets", "balance_sheet", "other_assets"),
    ("Balance", "payable_for_investments", "balance_sheet", "payable_for_investments"),
    ("Balance", "interest_payable", "balance_sheet", "interest_payable"),
    ("Balance", "management_fee_payable", "balance_sheet", "management_fee_payable"),
    ("Balance", "distribution_payable", "balance_sheet", "distribution_payable"),
    ("Balance", "additional_paid_in_capital", "balance_sheet", "additional_paid_in_capital"),
    ("Balance", "accumulated_deficit", "balance_sheet", "accumulated_deficit"),
    ("Balance", "liabilities_and_equity", "balance_sheet", "liabilities_and_equity"),
    # Income statement (§4)
    ("Income", "interest_income", "income_statement", "interest_income"),
    ("Income", "pik_interest_income", "income_statement", "pik_interest_income"),
    ("Income", "dividend_income", "income_statement", "dividend_income"),
    ("Income", "other_investment_income", "income_statement", "other_investment_income"),
    ("Income", "total_investment_income", "income_statement", "total_investment_income"),
    ("Income", "total_expenses", "income_statement", "total_expenses"),
    ("Income", "income_tax_expense", "income_statement", "income_tax_expense"),
    ("Income", "net_investment_income", "income_statement", "net_investment_income"),
    ("Income", "net_realized_gain_loss", "income_statement", "net_realized_gain_loss"),
    ("Income", "net_change_unrealized", "income_statement", "net_change_unrealized"),
    ("Income", "net_increase_in_net_assets_ops", "income_statement", "net_increase_in_net_assets_ops"),
    # Fair-value hierarchy (§6)
    ("FairValue", "fv_level_1", "fair_value", "fv_level_1"),
    ("FairValue", "fv_level_2", "fair_value", "fv_level_2"),
    ("FairValue", "fv_level_3", "fair_value", "fv_level_3"),
    ("FairValue", "fv_nav_practical_expedient", "fair_value", "fv_nav_practical_expedient"),
    ("FairValue", "fv_total", "fair_value", "fv_total"),
    # Fair-value % of total investments (computed = level / fv_total). The $ values stay above;
    # "__calc__" marks a derived column whose `field` names the level it divides by fv_total.
    ("FairValue", "pct_fv_level_1", "__calc__", "fv_level_1"),
    ("FairValue", "pct_fv_level_2", "__calc__", "fv_level_2"),
    ("FairValue", "pct_fv_level_3", "__calc__", "fv_level_3"),
    ("FairValue", "pct_fv_nav_practical_expedient", "__calc__", "fv_nav_practical_expedient"),
    # Statement of changes (§5)
    ("Changes", "beginning_net_assets", "statement_of_changes", "beginning_net_assets"),
    ("Changes", "capital_raised", "statement_of_changes", "capital_raised"),
    ("Changes", "repurchases", "statement_of_changes", "repurchases"),
    ("Changes", "distributions_declared", "statement_of_changes", "distributions_declared"),
    ("Changes", "ending_net_assets", "statement_of_changes", "ending_net_assets"),
    # Capital share activity (§5 detail) — summed across share classes
    ("Changes", "shares_issued_new", "statement_of_changes", "shares_issued_new"),
    ("Changes", "proceeds_new_issues", "statement_of_changes", "proceeds_new_issues"),
    ("Changes", "shares_issued_drip", "statement_of_changes", "shares_issued_drip"),
    ("Changes", "value_drip", "statement_of_changes", "value_drip"),
    ("Changes", "shares_repurchased", "statement_of_changes", "shares_repurchased"),
    # Cash-flow statement (§5b)
    ("CashFlow", "net_cash_operating", "cash_flow", "net_cash_operating"),
    ("CashFlow", "net_cash_investing", "cash_flow", "net_cash_investing"),
    ("CashFlow", "net_cash_financing", "cash_flow", "net_cash_financing"),
    ("CashFlow", "effect_of_fx", "cash_flow", "effect_of_fx"),
    ("CashFlow", "net_change_in_cash", "cash_flow", "net_change_in_cash"),
    ("CashFlow", "interest_paid", "cash_flow", "interest_paid"),
    ("CashFlow", "investment_purchases", "cash_flow", "investment_purchases"),
    ("CashFlow", "investment_sales", "cash_flow", "investment_sales"),
    # Fees (§11)
    ("Fees", "management_fee", "fees", "management_fee"),
    ("Fees", "incentive_fee", "fees", "incentive_fee"),
    ("Fees", "expense_support_net", "fees", "expense_support_net"),
    # Expense breakdown (§11, Theme 5)
    ("Fees", "interest_expense", "fees", "interest_expense"),
    ("Fees", "administrative_fees", "fees", "administrative_fees"),
    ("Fees", "professional_fees", "fees", "professional_fees"),
    ("Fees", "other_g_and_a", "fees", "other_g_and_a"),
    ("Fees", "director_trustee_fees", "fees", "director_trustee_fees"),
    ("Fees", "amortization_of_financing_costs", "fees", "amortization_of_financing_costs"),
    # Financial highlights (§7)
    ("Highlights", "expense_ratio", "financial_highlights", "expense_ratio"),
    ("Highlights", "gross_expense_ratio", "financial_highlights", "gross_expense_ratio"),
    ("Highlights", "net_investment_income_ratio", "financial_highlights", "net_investment_income_ratio"),
    ("Highlights", "total_return", "financial_highlights", "total_return"),
    ("Highlights", "portfolio_turnover", "financial_highlights", "portfolio_turnover"),
    # Distributions & leverage (§8)
    ("DistLev", "distributions_per_share", "distributions_leverage", "distributions_per_share"),
    ("DistLev", "return_of_capital_distribution", "distributions_leverage", "return_of_capital_distribution"),
    ("DistLev", "return_of_capital_pct", "distributions_leverage", "return_of_capital_pct"),
    ("DistLev", "asset_coverage_ratio", "distributions_leverage", "asset_coverage_ratio"),
    ("DistLev", "weighted_avg_interest_rate", "distributions_leverage", "weighted_avg_interest_rate"),
    # Tax basis (§13, Theme 6)
    ("TaxBasis", "tax_cost_of_investments", "tax_basis", "tax_cost_of_investments"),
    ("TaxBasis", "tax_unrealized_appreciation", "tax_basis", "tax_unrealized_appreciation"),
    ("TaxBasis", "tax_unrealized_depreciation", "tax_basis", "tax_unrealized_depreciation"),
    ("TaxBasis", "tax_unrealized_net", "tax_basis", "tax_unrealized_net"),
    ("TaxBasis", "undistributed_ordinary_income", "tax_basis", "undistributed_ordinary_income"),
    ("TaxBasis", "undistributed_lt_capital_gains", "tax_basis", "undistributed_lt_capital_gains"),
    # Portfolio summary (§9)
    ("Portfolio", "num_holdings", "portfolio_summary", "num_holdings"),
    ("Portfolio", "investments_at_cost", "portfolio_summary", "investments_at_cost"),
    ("Portfolio", "non_accrual_fair_value", "portfolio_summary", "non_accrual_fair_value"),
    ("Portfolio", "non_accrual_at_cost", "portfolio_summary", "non_accrual_at_cost"),
    ("Portfolio", "weighted_avg_portfolio_yield", "portfolio_summary", "weighted_avg_portfolio_yield"),
    ("Portfolio", "weighted_avg_spread", "portfolio_summary", "weighted_avg_spread"),
    ("Portfolio", "pct_floating_rate", "portfolio_summary", "pct_floating_rate"),
    ("Portfolio", "pct_holdings_with_pik", "portfolio_summary", "pct_holdings_with_pik"),
    ("Portfolio", "pct_holdings_with_pik_fv_weighted", "portfolio_summary", "pct_holdings_with_pik_fv_weighted"),
    ("Portfolio", "pct_affiliated", "portfolio_summary", "pct_affiliated"),
    ("Portfolio", "capitalized_pik_balance", "portfolio_summary", "capitalized_pik_balance"),
    ("Portfolio", "top_10_concentration", "portfolio_summary", "top_10_concentration"),
    # Liquidity & obligations (§12)
    ("Liquidity", "unfunded_commitments", "liquidity", "unfunded_commitments"),
    ("Liquidity", "undrawn_debt_capacity", "liquidity", "undrawn_debt_capacity"),
    ("Liquidity", "weighted_avg_debt_maturity", "liquidity", "weighted_avg_debt_maturity"),
    ("Liquidity", "repurchase_offered", "liquidity", "repurchase_offered"),
    ("Liquidity", "repurchase_requested", "liquidity", "repurchase_requested"),
    ("Liquidity", "repurchase_repurchased", "liquidity", "repurchase_repurchased"),
    ("Liquidity", "repurchase_proration_pct", "liquidity", "repurchase_proration_pct"),
    # Derived metrics (§10)
    ("Derived", "leverage_ratio", "derived", "leverage_ratio"),
    ("Derived", "asset_coverage_pct", "derived", "asset_coverage_pct"),
    ("Derived", "net_debt", "derived", "net_debt"),
    ("Derived", "pik_income_ratio", "derived", "pik_income_ratio"),
    ("Derived", "non_accrual_pct_fv", "derived", "non_accrual_pct_fv"),
    ("Derived", "non_accrual_pct_cost", "derived", "non_accrual_pct_cost"),
    ("Derived", "distribution_coverage_ratio", "derived", "distribution_coverage_ratio"),
    ("Derived", "portfolio_mark", "derived", "portfolio_mark"),
    # net_lending_spread computed in-workbook = weighted_avg_portfolio_yield - weighted_avg_interest_rate
    ("Derived", "net_lending_spread", "__netspread__", ""),
    ("Derived", "liquidity_coverage", "derived", "liquidity_coverage"),
    # Holdings reconciliation diagnostic (computed in-workbook): sum of schedule-of-investments
    # fair value / balance-sheet investments_at_fair_value. ~1.0 = the SOI ties to the balance
    # sheet; far from 1.0 = a data-quality concern (flagged amber). NOT a validation rule.
    ("Derived", "holdings_fv_recon", "__recon__", ""),
]

META_COLS = ["cik", "fund_name", "form_type", "reporting_date", "period_months",
             "vehicle_type", "status", "flags"]

# Which Data columns each validation rule implicates (for cell-level highlighting). Per-class
# rules (C2/A2) highlight in the ShareClasses tab instead, so they're not here.
RULE_FIELDS: dict[str, list[str]] = {
    "C1": ["total_assets", "total_liabilities", "total_net_assets"],
    "C1b": ["total_assets", "liabilities_and_equity"],
    "C3": ["total_net_assets"],
    "C4": ["fv_level_1", "fv_level_2", "fv_level_3", "fv_nav_practical_expedient", "fv_total",
           "pct_fv_level_1", "pct_fv_level_2", "pct_fv_level_3", "pct_fv_nav_practical_expedient"],
    "C5": ["total_investment_income", "total_expenses", "net_investment_income"],
    "C7": ["interest_income", "pik_interest_income", "dividend_income",
           "other_investment_income", "total_investment_income"],
    "C9": ["net_cash_operating", "net_cash_investing", "net_cash_financing", "effect_of_fx",
           "net_change_in_cash"],
    "C10": ["interest_expense", "administrative_fees", "professional_fees", "other_g_and_a",
            "director_trustee_fees", "amortization_of_financing_costs"],
    "C11": ["tax_unrealized_appreciation", "tax_unrealized_depreciation", "tax_unrealized_net"],
    "A1": ["total_net_assets"],
    "I1": ["asset_coverage_ratio"],
    "I2": ["leverage_ratio"],
}

# Gold sample (~15): clean + messy filers, single + multi-class, LLC + corp. Each entry is a
# (fund_name prefix, preferred form) — we take that fund's latest filing of the preferred form.
GOLD_SELECTION = [
    ("Apollo Debt Solutions", "10-K"),       # clean, annual (verify full-year flows)
    ("Blackstone Private Credit", "10-Q"),   # clean, large, NAV-PE bucket
    ("HPS Corporate Lending", "10-Q"),       # clean, 4-class, fair-value sum2
    ("AB Private Lending", "10-Q"),          # excise-tax / C5 waiver case
    ("Blue Owl Credit Income", "10-Q"),      # PIK-dividend (C7) case
    ("First Eagle Private Credit", "10-Q"),  # messy: C2/C3/C4 — verify flags are correct
    ("Antares Private Credit", "10-Q"),      # C5 flag — verify NII value is right anyway
    ("PGIM Private Credit", "10-Q"),         # C4 residual
    ("Golub Capital Private Credit", "10-K"),# clean, 4-class, annual
    ("John Hancock Comvest", "10-Q"),        # most share classes (5)
    ("T. Rowe Price OHA Select", "10-Q"),    # newer sponsor entrant — broaden gold coverage
    ("Fidelity Private Credit", "10-Q"),     # LLC, single-class
    ("Oaktree Strategic Credit", "10-Q"),    # C5 split-tax case
    ("Bain Capital Private Credit", "10-Q"), # clean, 4-class
    ("Crescent Private Credit", "10-Q"),     # C4 cleared via NAV-PE — verify
]

# Key fields shown in the gold check view (curated subset, in reading order).
GOLD_FIELDS = [
    ("balance_sheet", "total_assets"), ("balance_sheet", "total_liabilities"),
    ("balance_sheet", "total_net_assets"), ("balance_sheet", "investments_at_fair_value"),
    ("balance_sheet", "cash_and_equivalents"), ("balance_sheet", "total_debt"),
    ("income_statement", "total_investment_income"), ("income_statement", "total_expenses"),
    ("income_statement", "net_investment_income"), ("income_statement", "pik_interest_income"),
    ("fair_value", "fv_level_1"), ("fair_value", "fv_level_2"), ("fair_value", "fv_level_3"),
    ("fair_value", "fv_nav_practical_expedient"), ("fair_value", "fv_total"),
    ("portfolio_summary", "num_holdings"), ("portfolio_summary", "investments_at_cost"),
]

# ── Definitions content ──────────────────────────────────────────────────────────
# Calculated / derived data points. Each: (column_name, formula, plain-language meaning).
# Formulas use the EXACT Data-tab column names so any value can be traced and hand-checked.
DERIVED_DEFS: list[tuple[str, str, str]] = [
    ("leverage_ratio", "total_debt / total_net_assets",
     "Borrowings as a multiple of net assets — how levered the fund is."),
    ("net_debt", "total_debt - cash_and_equivalents",
     "Borrowings net of cash on hand."),
    ("asset_coverage_pct", "(total_assets - (total_liabilities - total_debt)) / total_debt",
     "Assets available per dollar of debt (assets minus non-debt liabilities, divided by debt). "
     "The 1940-Act leverage-analysis ratio; ~2.0 = 200% coverage. Distinct from the as-tagged "
     "asset_coverage_ratio column, which is the filer's reported regulatory figure."),
    ("portfolio_mark", "investments_at_fair_value / investments_at_cost",
     "Fair value of the portfolio vs. its cost. Below 1.0 = marked below cost (net unrealized loss)."),
    ("pik_income_ratio", "pik_interest_income / total_investment_income",
     "Share of income that is paid-in-kind (accrued, not received in cash) — a credit-stress signal."),
    ("distribution_coverage_ratio", "net_investment_income / distributions_declared",
     "How well net investment income covers declared distributions. Below 1.0 = distributions "
     "exceed income earned (likely partial return of capital)."),
    ("non_accrual_pct_fv", "non_accrual_fair_value / investments_at_fair_value",
     "Portfolio (by fair value) on non-accrual / not paying. NOT YET POPULATED — non-accrual "
     "amounts aren't in clean XBRL (schedule-of-investments footnotes); pending the LLM/HTML phase."),
    ("non_accrual_pct_cost", "non_accrual_at_cost / investments_at_cost",
     "Portfolio (by cost) on non-accrual; usually higher than the fair-value version. "
     "NOT YET POPULATED — pending the LLM/HTML phase."),
    ("net_lending_spread", "weighted_avg_portfolio_yield - weighted_avg_interest_rate",
     "Gross spread between what the portfolio earns and the fund's cost of debt. Computed "
     "in-workbook where BOTH legs exist; yield is holdings-derived and the cost-of-debt leg "
     "(weighted_avg_interest_rate) is tagged for only ~86 filings, so coverage is limited."),
    ("liquidity_coverage", "(cash_and_equivalents + undrawn_debt_capacity) / unfunded_commitments",
     "Available liquidity vs. commitments the fund may have to fund. Computed only when all "
     "three inputs are present; undrawn_debt_capacity is tagged by only some filers (see below), "
     "so coverage is partial."),
]

# Percent columns computed inside this workbook (not stored in the JSON).
WORKBOOK_CALC_DEFS: list[tuple[str, str, str]] = [
    ("pct_fv_level_1", "fv_level_1 / fv_total",
     "Level 1 (quoted prices) as a % of total investments at fair value."),
    ("pct_fv_level_2", "fv_level_2 / fv_total",
     "Level 2 (observable inputs) as a % of total investments at fair value."),
    ("pct_fv_level_3", "fv_level_3 / fv_total",
     "Level 3 (unobservable / model inputs) as a % of total — most direct-lending loans sit here."),
    ("pct_fv_nav_practical_expedient", "fv_nav_practical_expedient / fv_total",
     "NAV-measured holdings (money-market / alternative funds) as a % of total."),
]

# Holdings-derived §9 metrics — computed from the schedule of investments (the separate
# per-filing holdings CSV in data/holdings/). Formulas use the holdings CSV column names.
HOLDINGS_DERIVED_DEFS: list[tuple[str, str, str]] = [
    ("num_holdings", "count of holdings (current period)",
     "Number of portfolio positions (schedule-of-investments line items) at the reporting date."),
    ("top_10_concentration", "sum(10 largest fair_value) / sum(fair_value)",
     "Fair value of the 10 largest holdings as a share of the whole portfolio — concentration."),
    ("pct_floating_rate", "sum(fair_value where spread tagged) / sum(fair_value)",
     "Share of the portfolio (by fair value) on a floating rate (carries a spread over a base rate)."),
    ("weighted_avg_portfolio_yield", "sum(fair_value * rate) / sum(fair_value), over holdings with a rate",
     "Fair-value-weighted all-in interest rate. NULL when the rate concept is tagged for <60% of "
     "fair value (e.g. Apollo, First Eagle don't tag it) — reported only when reliable."),
    ("weighted_avg_spread", "sum(fair_value * spread) / sum(fair_value), over holdings with a spread",
     "Fair-value-weighted spread over the base rate. More robust than the all-in yield (spread is "
     "tagged across filers); mis-scaled outliers (>=1) are excluded from the average."),
    ("pct_holdings_with_pik", "count(holdings with a pik_rate) / num_holdings",
     "Share of positions (count basis) carrying a payment-in-kind rate — a credit-stress signal."),
    ("pct_holdings_with_pik_fv_weighted", "sum(fair_value where pik_rate tagged) / sum(fair_value)",
     "Share of the portfolio (by fair value) in PIK-bearing holdings — sizes PIK DOLLAR exposure, "
     "vs pct_holdings_with_pik which counts positions. A few large PIK loans can be a small count "
     "but a large fair-value share (or vice versa)."),
    ("pct_affiliated", "sum(fair_value where affiliated) / sum(fair_value)",
     "Share (by fair value) in affiliated/controlled issuers. NULL when issuer affiliation isn't "
     "parseable from the holding label (some filers don't use the 'Issuer | Affiliation' convention)."),
    ("unfunded_commitments", "sum(commitment) over holdings",
     "Total unfunded commitments the fund may still have to fund (feeds liquidity_coverage). "
     "Stored under the Liquidity section. NULL for filers that don't tag per-holding commitments."),
    ("holdings_fv_recon", "sum(holdings fair_value) / investments_at_fair_value",
     "DATA-QUALITY DIAGNOSTIC (not a validation rule): does the schedule of investments tie to "
     "the balance-sheet investments line? ~1.0 = reconciles; far from 1.0 (flagged amber) means "
     "the holdings or the balance-sheet line is structurally off for that filer (e.g. fund-of-"
     "funds look-through) — a follow-up signal, not a confirmed error."),
]

# Extracted (not derived) fields whose extraction METHODOLOGY is non-obvious enough to document.
EXTRACTED_METHOD_DEFS: list[tuple[str, str, str]] = [
    ("undrawn_debt_capacity", "us-gaap:LineOfCreditFacilityRemainingBorrowingCapacity",
     "Undrawn revolver / credit-facility capacity. Prefer the UNDIMENSIONED total where tagged "
     "(clean fund-level figure). Otherwise the rows are tagged only per-facility and cross-tabbed "
     "(the same total appears under a coarse axis and a finer breakdown), so we group by "
     "axis-signature and take the LARGEST single group's sum — never summing across signatures "
     "(which would double-count). We do NOT derive it from maximum-capacity − drawn (the maximum "
     "facts double-count across facility/SPV axes). NULL for filers that tag capacity only via "
     "maximum (not remaining) capacity or not at all — coverage is partial by design. Feeds "
     "liquidity_coverage."),
]

# Tax basis (§13, Theme 6) — the portfolio's tax position + distribution character.
TAX_BASIS_DEFS: list[tuple[str, str, str]] = [
    ("tax_cost_of_investments", "us-gaap:TaxBasisOfInvestmentsCostForIncomeTaxPurposes",
     "Aggregate cost of the portfolio for federal income-tax purposes. Differs from book cost "
     "(investments_at_cost) due to tax adjustments (wash sales, OID, etc.)."),
    ("tax_unrealized_appreciation", "us-gaap:TaxBasisOfInvestmentsGrossUnrealizedAppreciation",
     "Gross built-in GAIN on a tax basis (sum of positions above tax cost)."),
    ("tax_unrealized_depreciation", "us-gaap:TaxBasisOfInvestmentsGrossUnrealizedDepreciation",
     "Gross built-in LOSS on a tax basis. C11 checks apprec/deprec reconcile to the net."),
    ("tax_unrealized_net", "us-gaap:TaxBasisOfInvestmentsUnrealizedAppreciationDepreciationNet",
     "Net built-in gain/loss on a tax basis = appreciation − depreciation."),
    ("undistributed_ordinary_income", "InvestmentCompanyDistributableEarnings…OrdinaryIncomeLoss",
     "Accumulated ordinary income not yet distributed (a net-asset tax bucket) — capacity for "
     "future income distributions."),
    ("undistributed_lt_capital_gains", "InvestmentCompanyDistributableEarnings…LongTermCapitalGainLoss",
     "Accumulated long-term capital gains not yet distributed."),
    ("return_of_capital_distribution", "us-gaap:InvestmentCompanyTaxReturnOfCapitalDistribution",
     "$ of distributions characterized as RETURN OF CAPITAL (not income/gains) for tax — a return "
     "of the investor's own capital. Tagged by ~half of filers."),
    ("return_of_capital_pct", "return_of_capital_distribution / distributions_declared",
     "Share of distributions that is return of capital. High = the fund is paying out more than it "
     "earns (a distribution-quality concern). Computed only when the ROC $ is tagged."),
]

# Expense breakdown (§11, Theme 5) — gross expense lines decomposing total_expenses.
EXPENSE_DETAIL_DEFS: list[tuple[str, str, str]] = [
    ("interest_expense", "us-gaap:InterestExpenseDebt (or Borrowings/Operating/…)",
     "Dollar interest cost of the fund's borrowings — usually the largest non-fee expense. "
     "Tagged under several concepts across filers; we take the total."),
    ("administrative_fees", "us-gaap:AdministrativeFeesExpense", "Administration / fund-services fees."),
    ("professional_fees", "us-gaap:ProfessionalFees", "Legal, audit, and other professional fees."),
    ("other_g_and_a", "us-gaap:OtherGeneralAndAdministrativeExpense", "Other general & administrative expense."),
    ("director_trustee_fees", "us-gaap:TrusteeFees / NoninterestExpenseDirectorsFees",
     "Board of directors / trustees fees."),
    ("amortization_of_financing_costs", "us-gaap:AmortizationOfFinancingCosts",
     "Non-cash amortization of deferred debt-issuance costs (added back in operating cash flow)."),
    ("(note)", "GROSS vs NET",
     "These are GROSS expense lines; their sum can exceed the NET total_expenses (which is after "
     "fee waivers / expense support). C10 only flags an implausibly large sum (>2x total)."),
]

# Balance-sheet detail (§2, Theme 4) — direct instant extractions; a few have non-obvious meaning.
BALANCE_DETAIL_DEFS: list[tuple[str, str, str]] = [
    ("receivable_for_investments", "us-gaap:ReceivableInvestmentSale",
     "Unsettled trades — investments sold but not yet settled/collected (a receivable)."),
    ("payable_for_investments", "us-gaap:PayableInvestmentPurchase",
     "Unsettled trades — investments bought but not yet paid for (a payable)."),
    ("accumulated_deficit", "us-gaap:RetainedEarningsAccumulatedDeficit",
     "Accumulated distributed earnings (losses) — cumulative earnings net of distributions. "
     "Negative = the fund has distributed more than it has earned to date."),
    ("liabilities_and_equity", "us-gaap:LiabilitiesAndStockholdersEquity",
     "The tagged 'total liabilities and net assets' line. MUST equal total_assets — used as the "
     "C1b cross-check. (interest_receivable, other_assets, interest_payable, management_fee_payable, "
     "distribution_payable, additional_paid_in_capital are direct, self-explanatory line items.)"),
]

# Capital share activity (§5 detail) — extracted per share class and summed.
SHARE_ACTIVITY_DEFS: list[tuple[str, str, str]] = [
    ("shares_issued_new", "us-gaap:StockIssuedDuringPeriodSharesNewIssues (sum over classes)",
     "Shares sold to new/existing investors during the period (subscriptions), summed across share "
     "classes. Detail behind capital_raised."),
    ("proceeds_new_issues", "us-gaap:StockIssuedDuringPeriodValueNewIssues (sum over classes)",
     "Dollars raised from new subscriptions, summed across classes. May differ from capital_raised "
     "(which is the cash-flow ProceedsFromIssuanceOfCommonStock)."),
    ("shares_issued_drip", "us-gaap:StockIssuedDuringPeriodSharesDividendReinvestmentPlan (sum)",
     "Shares issued via dividend reinvestment (DRIP), summed across classes."),
    ("value_drip", "us-gaap:StockIssuedDuringPeriodValueDividendReinvestmentPlan (sum)",
     "Dollar value of reinvested distributions. A wash in cash terms (a distribution that comes "
     "back as shares) — increases net assets in the roll-forward."),
    ("shares_repurchased", "us-gaap:StockRepurchasedDuringPeriodShares (sum over classes)",
     "Shares bought back via the periodic tender/repurchase offer, summed across classes. Detail "
     "behind repurchases."),
]

# Cash-flow statement fields (§5b) — extracted as-tagged, but a couple of conventions are
# non-obvious enough to document.
CASH_FLOW_DEFS: list[tuple[str, str, str]] = [
    ("net_cash_operating", "us-gaap:NetCashProvidedByUsedInOperatingActivities",
     "Cash from operations. For an investment company this INCLUDES buying/selling investments "
     "(there is usually no separate investing section), so it is typically a large negative number "
     "in a growing fund (deploying capital)."),
    ("net_cash_financing", "us-gaap:NetCashProvidedByUsedInFinancingActivities",
     "Cash from financing — share issuance/repurchases and net borrowings. Usually the positive "
     "offset to operating in a growing fund."),
    ("net_cash_investing", "us-gaap:NetCashProvidedByUsedInInvestingActivities",
     "Investing-activities cash. Usually ABSENT for BDCs (investments are an operating activity); "
     "treated as 0 in the C9 footing check when not tagged."),
    ("net_change_in_cash", "PeriodIncreaseDecrease...IncludingExchangeRateEffect",
     "Bottom-line change in cash for the period, INCLUDING the FX effect. C9 checks that "
     "operating + investing + financing + effect_of_fx equals this."),
    ("interest_paid", "us-gaap:InterestPaidNet",
     "Cash interest actually paid in the period (vs. the accrued interest expense on the income "
     "statement)."),
    ("investment_purchases", "us-gaap:PaymentsForPurchaseOfInvestmentOperatingActivity",
     "Cash used to buy investments during the period — a gross deployment / turnover signal."),
    ("investment_sales", "us-gaap:ProceedsFromDispositionOfInvestmentOperatingActivity",
     "Cash from selling / repaying investments — the other half of the turnover signal."),
]

# Validation/review codes for the Review-tab key: (code, short name, type, what it verifies).
# Identity = an accounting equation that MUST hold (a fail means the extraction is likely wrong).
# Reasonableness = the value is KEPT and flagged for a human to eyeball (flag-and-keep policy).
REVIEW_CODES: list[tuple[str, str, str, str]] = [
    ("C1", "Balance sheet equation", "Identity",
     "total_assets = total_liabilities + total_net_assets."),
    ("C1b", "Liab+equity total = assets", "Reasonableness",
     "The tagged liabilities_and_equity (total liabilities and net assets) equals total_assets. "
     "A redundant cross-check; fires only when both are present."),
    ("C2", "NAV per share", "Identity",
     "Per ShareClasses row: nav_per_share = net_assets / shares."),
    ("C3", "Class net assets sum", "Identity",
     "The ShareClasses net_assets for a filing add up to total_net_assets."),
    ("C4", "Fair-value hierarchy sum", "Identity",
     "fv_level_1 + fv_level_2 + fv_level_3 + fv_nav_practical_expedient = fv_total."),
    ("C5", "NII reconciles", "Identity",
     "net_investment_income = total_investment_income - total_expenses (less income_tax_expense "
     "where tax sits above NII), or equals the filer's tagged income subtotal."),
    ("C7", "Income components sum", "Identity",
     "interest_income + pik_interest_income + dividend_income + other_investment_income reconcile "
     "to total_investment_income (a paid-in-kind band is allowed)."),
    ("A1", "Net assets positive", "Reasonableness",
     "total_net_assets > 0 (a basic sanity check)."),
    ("A2", "NAV in plausible range", "Reasonableness",
     "A share class's nav_per_share is within $1-$100 (catches ~1000x unit errors); empty/dormant "
     "classes are skipped."),
    ("I1", "Asset coverage >= 150%", "Reasonableness",
     "asset_coverage_pct >= 1.5 (150%, the 1940-Act regulatory minimum)."),
    ("I2", "Leverage in range", "Reasonableness",
     "leverage_ratio sits within 0-2."),
    ("C8", "Undrawn debt capacity plausible", "Reasonableness",
     "undrawn_debt_capacity is non-negative and does not exceed total_assets (a larger value "
     "would signal a double-count / mis-tag). Value is kept either way."),
    ("C9", "Cash flow statement foots", "Identity",
     "net_cash_operating + net_cash_investing + net_cash_financing + effect_of_fx = "
     "net_change_in_cash (investing/fx default to 0 when a filer omits them)."),
    ("C10", "Expense breakdown sane", "Reasonableness",
     "The captured expense components (management/incentive/interest/admin/professional/G&A/"
     "trustee/amortization) sum to <= 2x total_expenses. They are GROSS lines, so a loose bound "
     "(not an exact footing) avoids false flags when waivers reduce the net total."),
    ("C11", "Tax unrealized nets out", "Reasonableness",
     "tax_unrealized_appreciation and tax_unrealized_depreciation reconcile to tax_unrealized_net "
     "(accepting either sign convention for depreciation)."),
]

# ── Styling ─────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="2E5496")
REVIEW_ROW_FILL = PatternFill("solid", fgColor="FFF2CC")   # light amber row tint
FLAG_CELL_FILL = PatternFill("solid", fgColor="F4B183")    # stronger amber for flagged cells
PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = "#,##0"      # default: dollar amounts
PCT_FMT = "0.0%"         # decimals we compute as fractions (fair-value %, PIK ratio, …)
RATIO_FMT = "0.000"      # multiples / coverage ratios we compute (portfolio_mark, leverage, …)
DEC_FMT = "0.0000"       # as-tagged rates/ratios whose scale we don't force into %

# Per-column number format (column NAME -> format). Anything not listed uses MONEY_FMT.
NUMBER_FORMATS = {
    "pct_fv_level_1": PCT_FMT, "pct_fv_level_2": PCT_FMT, "pct_fv_level_3": PCT_FMT,
    "pct_fv_nav_practical_expedient": PCT_FMT,
    "pik_income_ratio": PCT_FMT, "non_accrual_pct_fv": PCT_FMT, "non_accrual_pct_cost": PCT_FMT,
    # Holdings-derived §9 metrics (computed fractions -> show as %)
    "top_10_concentration": PCT_FMT, "pct_holdings_with_pik": PCT_FMT,
    "pct_holdings_with_pik_fv_weighted": PCT_FMT, "pct_affiliated": PCT_FMT,
    "weighted_avg_spread": PCT_FMT,
    "portfolio_mark": RATIO_FMT, "leverage_ratio": RATIO_FMT, "asset_coverage_ratio": RATIO_FMT,
    "asset_coverage_pct": RATIO_FMT, "distribution_coverage_ratio": RATIO_FMT,
    "net_lending_spread": PCT_FMT, "liquidity_coverage": RATIO_FMT,
    "holdings_fv_recon": RATIO_FMT,
    "distributions_per_share": DEC_FMT, "num_holdings": "#,##0",
    # capital share activity — share counts (values use the default money format)
    "shares_issued_new": "#,##0", "shares_issued_drip": "#,##0", "shares_repurchased": "#,##0",
    # as-tagged §7/§8 rates & ratios (scale varies by filer) -> plain decimal, not money
    "expense_ratio": DEC_FMT, "gross_expense_ratio": DEC_FMT, "net_investment_income_ratio": DEC_FMT,
    "total_return": DEC_FMT, "portfolio_turnover": DEC_FMT, "return_of_capital_pct": PCT_FMT,
    "weighted_avg_interest_rate": DEC_FMT, "weighted_avg_portfolio_yield": PCT_FMT,
    "pct_floating_rate": PCT_FMT, "repurchase_proration_pct": DEC_FMT,
    "weighted_avg_debt_maturity": "0.0",
}


def fact_value(j: dict, section: str, field: str):
    """Return the numeric .value for a Fact field, or None."""
    sec = j.get(section) or {}
    f = sec.get(field) or {}
    return f.get("value")


def fact_source(j: dict, section: str, field: str) -> str:
    sec = j.get(section) or {}
    f = sec.get(field) or {}
    return f.get("raw_text") or (f.get("source") or "")


def failing_rule_bases(j: dict) -> list[str]:
    return sorted({c["rule"].split("[")[0] for c in j.get("validation_checks", [])
                   if c["status"] == "fail"})


def load_filings() -> list[dict]:
    out = []
    for p in sorted(EXTRACTED.glob("*.json")):
        out.append(json.loads(p.read_text(encoding="utf-8")))
    return out


def filing_stem(j: dict) -> str:
    """The shared filename stem cik_form_reportingdate (JSON, holdings CSV)."""
    return f"{j.get('cik')}_{j.get('form_type')}_{j.get('reporting_date')}"


def read_holdings(stem: str) -> list[dict]:
    """Holding rows for one filing from its CSV (numeric fields coerced to float)."""
    path = HOLDINGS / f"{stem}.csv"
    if not path.exists():
        return []
    rows = []
    with path.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            for c in HOLDING_CSV_COLS:
                if c not in ("issuer", "affiliation"):
                    r[c] = float(r[c]) if r.get(c) not in (None, "") else None
            rows.append(r)
    return rows


def holdings_fv_sum(stem: str) -> float | None:
    """Sum of current-period holding fair values for one filing (for the reconciliation)."""
    rows = read_holdings(stem)
    s = sum(r["fair_value"] for r in rows if r.get("fair_value") is not None)
    return s or None


def sec_url(j: dict) -> str:
    acc = j.get("accession_no") or ""
    cik = j.get("cik", "").lstrip("0")
    if not acc or not cik:
        return ""
    return f"https://www.sec.gov/Archives/edgar/data/{cik}/{acc.replace('-', '')}/{acc}-index.htm"


def style_header(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_data_tab(wb, filings: list[dict]):
    ws = wb.create_sheet("Data")
    headers = META_COLS + [name for _, name, _, _ in DATA_FIELDS]
    # Row 1: section bands; Row 2: column names
    ws.append(["META"] * len(META_COLS) + [sec for sec, _, _, _ in DATA_FIELDS])
    ws.append(headers)
    style_header(ws, 2, len(headers))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center")

    field_col = {name: META_COLS.__len__() + i + 1 for i, (_, name, _, _) in enumerate(DATA_FIELDS)}
    for j in filings:
        fails = failing_rule_bases(j)
        row = [
            j.get("cik"), j.get("fund_name"), j.get("form_type"), j.get("reporting_date"),
            j.get("period_months"), j.get("vehicle_type"), j.get("validation_status"),
            ",".join(fails),
        ]
        stem = filing_stem(j)
        for _, _, sec, fld in DATA_FIELDS:
            if sec == "__calc__":      # fair-value % = level / fv_total
                num = fact_value(j, "fair_value", fld)
                den = fact_value(j, "fair_value", "fv_total")
                row.append((num / den) if (num is not None and den) else None)
            elif sec == "__netspread__":   # yield - cost of debt
                y = fact_value(j, "portfolio_summary", "weighted_avg_portfolio_yield")
                cod = fact_value(j, "distributions_leverage", "weighted_avg_interest_rate")
                row.append((y - cod) if (y is not None and cod is not None) else None)
            elif sec == "__recon__":       # sum(holdings FV) / investments_at_fair_value
                sfv = holdings_fv_sum(stem)
                ifv = fact_value(j, "balance_sheet", "investments_at_fair_value")
                row.append((sfv / ifv) if (sfv is not None and ifv) else None)
            else:
                row.append(fact_value(j, sec, fld))
        ws.append(row)
        r = ws.max_row
        is_review = j.get("validation_status") == "review"
        # row tint + per-column number formats
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if is_review:
                cell.fill = REVIEW_ROW_FILL
            if c > len(META_COLS) and isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMATS.get(headers[c - 1], MONEY_FMT)
        # cell-level highlight on fields implicated by each failing rule
        flagged_fields: set[str] = set()
        for rule in fails:
            flagged_fields.update(RULE_FIELDS.get(rule, []))
        for fld in flagged_fields:
            col = field_col.get(fld)
            if col:
                ws.cell(row=r, column=col).fill = FLAG_CELL_FILL
        # Data-quality flag: holdings don't reconcile to the balance sheet (>5% off).
        recon_col = field_col.get("holdings_fv_recon")
        recon_val = ws.cell(row=r, column=recon_col).value if recon_col else None
        if isinstance(recon_val, (int, float)) and abs(recon_val - 1.0) > 0.05:
            ws.cell(row=r, column=recon_col).fill = FLAG_CELL_FILL
    ws.freeze_panes = "C3"
    _autosize(ws, headers, start_row=2)
    return ws


def build_shareclasses_tab(wb, filings: list[dict]):
    ws = wb.create_sheet("ShareClasses")
    headers = ["cik", "fund_name", "reporting_date", "class", "net_assets",
               "shares", "nav_per_share", "distribution_yield", "flags"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for j in filings:
        fails = failing_rule_bases(j)
        # per-class fails (C2/A2) carry the class label in brackets
        class_fail = {}
        for c in j.get("validation_checks", []):
            if c["status"] == "fail" and "[" in c["rule"]:
                lbl = c["rule"].split("[")[1].rstrip("]")
                class_fail.setdefault(lbl, set()).add(c["rule"].split("[")[0])
        for sc in j.get("share_classes_nav", []):
            lbl = sc.get("class_label")
            ws.append([
                j.get("cik"), j.get("fund_name"), j.get("reporting_date"), lbl,
                (sc.get("class_net_assets") or {}).get("value"),
                (sc.get("class_shares_outstanding") or {}).get("value"),
                (sc.get("class_nav_per_share") or {}).get("value"),
                (sc.get("distribution_yield") or {}).get("value"),
                ",".join(sorted(class_fail.get(lbl, set()))),
            ])
            r = ws.max_row
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = BORDER
            for col in (5, 6, 7):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = MONEY_FMT
            if class_fail.get(lbl):
                ws.cell(row=r, column=7).fill = FLAG_CELL_FILL  # NAV cell
    ws.freeze_panes = "A2"
    _autosize(ws, headers, start_row=1)
    return ws


def build_review_tab(wb, filings: list[dict]):
    ws = wb.create_sheet("Review")

    # ── Code key (legend) — what each failing rule means ──────────────────────
    ws.append(["VALIDATION CODE KEY — what each failing rule checks"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(["Identity = accounting equation that must hold (a fail means the extraction is "
               "likely wrong). Reasonableness = value is kept and flagged for a human to eyeball."])
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    key_hdr = ws.max_row + 1
    ws.append(["code", "check", "type", "what it verifies"])
    style_header(ws, key_hdr, 4)
    for code, short, typ, meaning in REVIEW_CODES:
        ws.append([code, short, typ, meaning])
        r = ws.max_row
        ws.cell(row=r, column=1).font = Font(bold=True)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top")  # meaning overflows right
    ws.append([])

    # ── The review queue itself (one row per flagged filing) ──────────────────
    headers = ["cik", "fund_name", "reporting_date", "status", "failing_rules", "messages"]
    hdr_row = ws.max_row + 1
    ws.append(headers)
    style_header(ws, hdr_row, len(headers))
    for j in filings:
        if j.get("validation_status") != "review":
            continue
        fails = [c for c in j.get("validation_checks", []) if c["status"] == "fail"]
        ws.append([
            j.get("cik"), j.get("fund_name"), j.get("reporting_date"), "review",
            ",".join(sorted({c["rule"] for c in fails})),
            " | ".join(c["message"] for c in fails if c.get("message")),
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = REVIEW_ROW_FILL
            ws.cell(row=r, column=c).border = BORDER
    ws.freeze_panes = "A" + str(hdr_row + 1)
    _autosize(ws, headers, start_row=hdr_row, max_w=70)
    return ws


def pick_gold(filings: list[dict]) -> list[dict]:
    chosen = []
    for prefix, form in GOLD_SELECTION:
        cands = [j for j in filings if j.get("fund_name", "").startswith(prefix)
                 and j.get("form_type") == form]
        if not cands:  # fall back to any form
            cands = [j for j in filings if j.get("fund_name", "").startswith(prefix)]
        if cands:
            chosen.append(max(cands, key=lambda x: x.get("reporting_date", "")))
    return chosen


def build_gold_tab(wb, filings: list[dict]):
    ws = wb.create_sheet("Check (Gold)")
    gold = pick_gold(filings)
    # Accuracy summary block (formulas fill in as you complete the verdict column F).
    ws.append(["GOLD ACCURACY CHECK — fill column F (Y/N) against each filing on SEC.gov"])
    ws["A1"].font = Font(bold=True, size=12)
    # Verdict cells live in column F from row 5 down. Reference that range explicitly (NOT
    # the whole F:F column) so the Accuracy cell in F2 doesn't count itself -> no circular ref.
    ws.append(["Checked:", "=COUNTIF(F5:F10000,\"Y\")+COUNTIF(F5:F10000,\"N\")",
               "Correct:", "=COUNTIF(F5:F10000,\"Y\")",
               "Accuracy:", "=IFERROR(COUNTIF(F5:F10000,\"Y\")/(COUNTIF(F5:F10000,\"Y\")+COUNTIF(F5:F10000,\"N\")),\"-\")"])
    ws["F2"].number_format = "0.0%"
    for c in ("A2", "C2", "E2"):
        ws[c].font = Font(bold=True)
    ws.append([])
    headers = ["field", "extracted", "source (xbrl concept)", "confidence", "filing / link", "✓ (Y/N)", "true value", "note"]
    hdr_row = ws.max_row + 1
    ws.append(headers)
    style_header(ws, hdr_row, len(headers))

    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv)

    for j in gold:
        # filing header band
        label = f"{j.get('fund_name')}  |  {j.get('form_type')}  {j.get('reporting_date')}  (classes: {', '.join(j.get('share_classes') or ['single'])})"
        ws.append([label])
        br = ws.max_row
        ws.cell(row=br, column=1).font = Font(bold=True, color="FFFFFF")
        for c in range(1, len(headers) + 1):
            ws.cell(row=br, column=c).fill = SECTION_FILL
        ws.cell(row=br, column=5).value = sec_url(j)
        ws.cell(row=br, column=5).font = Font(color="FFFFFF", underline="single")
        # fund-level key fields
        for sec, fld in GOLD_FIELDS:
            val = fact_value(j, sec, fld)
            ws.append([fld, val, fact_source(j, sec, fld),
                       (j.get(sec, {}).get(fld, {}) or {}).get("confidence"), "", "", "", ""])
            r = ws.max_row
            if isinstance(val, (int, float)):
                ws.cell(row=r, column=2).number_format = MONEY_FMT
            dv.add(ws.cell(row=r, column=6))
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = BORDER
        # per-class NAV rows
        for sc in j.get("share_classes_nav", []):
            nav = (sc.get("class_nav_per_share") or {}).get("value")
            ws.append([f"NAV (class {sc.get('class_label')})", nav,
                       (sc.get("class_nav_per_share") or {}).get("raw_text", ""),
                       (sc.get("class_nav_per_share") or {}).get("confidence"), "", "", "", ""])
            r = ws.max_row
            dv.add(ws.cell(row=r, column=6))
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = BORDER
    ws.freeze_panes = "A" + str(hdr_row + 1)
    _autosize(ws, headers, start_row=hdr_row, max_w=55)
    return ws, len(gold)


def build_holdings_tab(wb, filings: list[dict]):
    """Holding-level schedule of investments for the GOLD sample only (the full ~130k rows
    stay in data/holdings/ CSVs — putting them all here would bloat the workbook). One row
    per holding, sorted by fund then fair value descending."""
    ws = wb.create_sheet("Holdings (Gold)")
    headers = ["fund_name", "reporting_date", "issuer", "affiliation", "fair_value", "cost",
               "principal", "rate", "spread", "pik_rate", "floor", "shares", "commitment", "pct_na"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    money_cols = {"fair_value", "cost", "principal", "commitment"}
    pct_cols = {"rate", "spread", "pik_rate", "floor", "pct_na"}
    n_funds = 0
    for j in pick_gold(filings):
        rows = read_holdings(filing_stem(j))
        if not rows:
            continue
        n_funds += 1
        rows.sort(key=lambda h: (h.get("fair_value") or 0), reverse=True)
        for h in rows:
            ws.append([j.get("fund_name"), j.get("reporting_date"), h.get("issuer"),
                       h.get("affiliation")] + [h.get(c) for c in headers[4:]])
            r = ws.max_row
            for ci, name in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=ci)
                cell.border = BORDER
                if isinstance(cell.value, (int, float)):
                    cell.number_format = (MONEY_FMT if name in money_cols
                                          else PCT_FMT if name in pct_cols else "#,##0")
    ws.freeze_panes = "A2"
    _autosize(ws, headers, start_row=1, max_w=45)
    return ws, n_funds


def build_definitions_tab(wb):
    """A glossary tab: every calculated/derived data point with its formula (in exact Data-tab
    column names) and a plain-language explanation. No numbers — methodology only."""
    ws = wb.create_sheet("Definitions")
    ws.append(["DEFINITIONS — calculated & derived data points"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(["Formulas use the EXACT column names from the Data tab so each value can be traced "
               "and hand-checked. '/' = divide, '-' = subtract."])
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    ws.append([])

    def section(title, rows):
        ws.append([title])
        tr = ws.max_row
        ws.cell(row=tr, column=1).font = Font(bold=True, color="FFFFFF")
        for c in range(1, 4):
            ws.cell(row=tr, column=c).fill = SECTION_FILL
        ws.append(["data point", "formula / methodology", "explanation"])
        style_header(ws, ws.max_row, 3)
        for name, formula, expl in rows:
            ws.append([name, formula, expl])
            r = ws.max_row
            ws.cell(row=r, column=1).font = Font(bold=True)
            for c in range(1, 4):
                ws.cell(row=r, column=c).border = BORDER
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        ws.append([])

    section("Derived metrics — computed during extraction (Data tab section: Derived, §10)",
            DERIVED_DEFS)
    section("Fair-value % columns — computed in this workbook (Data tab section: FairValue)",
            WORKBOOK_CALC_DEFS)
    section("Portfolio metrics — computed from the schedule of investments / holdings (§9)",
            HOLDINGS_DERIVED_DEFS)
    section("Expense breakdown (§11) — gross lines decomposing total_expenses",
            EXPENSE_DETAIL_DEFS)
    section("Tax basis & distribution character (§13) — portfolio tax position + return of capital",
            TAX_BASIS_DEFS)
    section("Balance-sheet detail (§2) — direct extractions; non-obvious ones noted",
            BALANCE_DETAIL_DEFS)
    section("Capital share activity (§5) — extracted per share class and summed",
            SHARE_ACTIVITY_DEFS)
    section("Cash-flow statement (§5b) — extracted; conventions to note",
            CASH_FLOW_DEFS)
    section("Extracted fields — non-obvious extraction methodology",
            EXTRACTED_METHOD_DEFS)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 85
    return ws


def _autosize(ws, headers, start_row: int, max_w: int = 22):
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), 10)
        for r in range(start_row, min(ws.max_row, start_row + 60) + 1):
            v = ws.cell(row=r, column=i).value
            if v is not None:
                width = max(width, min(len(str(v)), max_w))
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, max_w)


def main():
    filings = load_filings()
    if not filings:
        raise SystemExit(f"No JSONs found in {EXTRACTED}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default sheet
    build_data_tab(wb, filings)
    build_shareclasses_tab(wb, filings)
    build_review_tab(wb, filings)
    _, n_gold = build_gold_tab(wb, filings)
    _, n_hold = build_holdings_tab(wb, filings)
    build_definitions_tab(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    n_review = sum(1 for j in filings if j.get("validation_status") == "review")
    print(f"Wrote {OUT}")
    print(f"  Data: {len(filings)} filings | Review: {n_review} | Gold: {n_gold} | "
          f"Holdings tab: {n_hold} gold funds")


if __name__ == "__main__":
    main()
