"""
holdings_compare.py — Cross-BDC holdings & mark comparison (docs/HOLDINGS_COMPARISON_PLAN.md).

Independent of the extractor. Reads the per-filing schedule-of-investments CSVs in data/holdings/
(one row per holding) and works toward matching the SAME underlying credit across BDCs so we can
compare each holder's mark (fair value as a % of par).

This file currently implements PHASE 1 (consolidate + clean + parse) and a light issuer
normalization + diagnostic, so we can measure parse quality and confirm the named anchors
(Anaplan, Flexera, ...) surface before building issuer clustering / issue matching / mark
comparison (phases 2-5).

Run the diagnostic:
    uv run python src/analysis/holdings_compare.py --diagnose
Write the cleaned consolidated table to data/dataset/:
    uv run python src/analysis/holdings_compare.py --build

Design notes (why the parsing is shaped this way):
  - The raw `issuer` field is the XBRL InvestmentIdentifierAxis member verbatim. Its format varies
    by FILER, not by row: comma-delimited ("Name, Sector, Instrument"), em-dash ("Name - Debt"),
    or a denormalized category PATH with the real issuer buried at the end ("Non-Control...
    Application Software Airship Group"). No single split works; we detect the format per row.
  - We DEGRADE HONESTLY: a row we can't confidently parse gets parse_ok=False and is reported in
    the coverage stats rather than force-matched. The plan's promise is best-effort discovery with
    confidence, not exhaustive reconciliation.
"""

from __future__ import annotations

import argparse
import glob
import json
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz, process

PROJECT_ROOT = Path(__file__).resolve().parents[2]
HOLDINGS_DIR = PROJECT_ROOT / "data" / "holdings"
OUT_DIR = PROJECT_ROOT / "data" / "dataset"

# ---------------------------------------------------------------------------
# Cleaning vocab
# ---------------------------------------------------------------------------

# Category / hierarchy / sector phrases that denormalized filers prepend to the issuer
# ("Investments United States Debt Investments Health Care Equipment & Services ACI Group ...").
# Stripped from the FRONT token-by-token, repeatedly, stopping at the first non-boilerplate token.
# This is a generic vocabulary (affiliation buckets + geographies + GICS sectors/industries + asset
# classes), NOT per-filer special-casing. '&' is normalized to 'and' for matching.
_CATEGORY_PHRASES = [
    # affiliation / control buckets
    "non-controlled/non-affiliated investments", "non-control/non-affiliate investments",
    "non-controlled non-affiliated investments", "non-controlled/non-affiliated",
    "non-control/non-affiliate", "controlled affiliate investments",
    "non-controlled affiliate investments", "controlled investments", "control investments",
    "affiliate investments", "affiliated investments", "affiliate", "affiliated",
    "non-controlled", "non-affiliated",
    # SOI section headers
    "portfolio company debt securities", "portfolio company warrant investments",
    "portfolio company equity investments", "debt securities portfolio", "debt securities",
    "debt investments", "equity investments", "warrant investments",
    "us corporate debt", "u s corporate debt", "corporate debt", "us debt", "u s debt",
    "issuer name", "issuer", "investments", "investment",
    # sector phrases seen as headers on specific denormalized filers
    "drug discovery and development", "drug discovery", "cannabis",
    "transportation services", "utilities services", "business services",
    "consumer products and services", "consumer products and services",
    "consumer products", "consumer services", "special retail", "specialty retail",
    # asset classes / seniority used as a bucket header
    "senior secured first lien debt", "first and second lien debt", "first lien debt",
    "second lien debt", "secured debt", "subordinated debt", "unsecured debt",
    "first lien senior secured", "1st lien senior secured debt", "1st lien last out unitranche",
    "last out unitranche", "1st lien", "2nd lien", "3rd lien", "senior secured",
    "first lien", "second lien", "unitranche", "broadly syndicated", "one stop",
    "preferred equity", "common equity", "common stock", "preferred stock", "common shares",
    "preferred shares", "structured products", "joint ventures", "short term investments",
    # sector / asset-class leads on specific denormalized filers (leading-strip only, so safe)
    "services", "industrial conglomerates", "retail and consumer products",
    "consumer products and services", "diversified",
    # geographies
    "united states", "united kingdom", "australia", "canada", "netherlands", "luxembourg",
    "germany", "france", "ireland", "switzerland", "new zealand", "europe", "north america",
    "cayman islands", "bermuda", "jersey", "spain", "italy", "sweden", "norway", "denmark",
    # GICS sectors
    "energy", "materials", "industrials", "consumer discretionary", "consumer staples",
    "health care", "healthcare", "financials", "information technology", "communication services",
    "utilities", "real estate",
    # GICS industry groups / industries common in BDC SOIs
    "software", "it services", "high tech industries", "technology hardware", "semiconductors",
    "communications equipment", "electronic equipment", "internet",
    "health care equipment and services", "health care equipment", "health care technology",
    "health care providers and services", "health care providers", "pharmaceuticals",
    "biotechnology", "life sciences tools and services", "life sciences tools",
    "commercial services and supplies", "commercial and professional services",
    "commercial services", "professional services", "equity securities", "debt securities",
    "capital goods", "aerospace and defense", "machinery", "building products",
    "construction and engineering", "electrical equipment", "trading companies and distributors",
    "media", "entertainment", "media and entertainment", "interactive media and services",
    "hotels restaurants and leisure", "hotels, restaurants and leisure", "consumer services",
    "diversified consumer services", "specialty retail", "multiline retail", "distributors",
    "food products", "food and staples retailing", "beverages", "household durables",
    "household products", "personal products", "leisure products",
    "textiles apparel and luxury goods", "textiles, apparel and luxury goods",
    "chemicals", "containers and packaging", "metals and mining", "paper and forest products",
    "insurance", "banks", "capital markets", "diversified financial services",
    "consumer finance", "financial services", "thrifts and mortgage finance",
    "road and rail", "transportation", "air freight and logistics", "airlines",
    "energy equipment and services", "oil gas and consumable fuels", "oil, gas and consumable fuels",
    "automobiles", "auto components", "automotive",
]

# Structured-attribute markers used by denormalized filers ("... Issuer Investment Type First Lien
# ... Interest Rate 8.79% Maturity Date 11/2030"). The issuer name sits BEFORE the first marker, so
# we cut the member there — discarding attribute noise we already capture as clean XBRL facts.
_STRUCT_MARKER = re.compile(
    r"\b(investment type|type of investment|interest term|interest rate|reference rate(?: and spread)?|"
    r"maturity\s*/?\s*dissolution|maturity|commitment type|commitment expiration|"
    r"investment date|acquisition date|acquisition\s+\d|\bindustry\b|\bsector\b|asset type|"
    r"facility type|all[- ]?in rate|benchmark|variable index|variable interest rate|"
    r"original purchase|purchase date|initial purchase)\b", re.I)

# A stray subtotal percentage or pure-punctuation token between path levels ("Debt Investments
# 233.2% United States - 220.5% 1st Lien ...") — treated as strippable boilerplate.
_PCT_TOKEN_RE = re.compile(r"^[-–—(]*\d+(\.\d+)?\s*%[)]*$")
_PUNCT_TOKEN_RE = re.compile(r"^[-–—:;,.]+$")

# Legal-entity suffixes — when the token right after the first comma is one of these, the comma is
# INSIDE the legal name ("Belnick, LLC"), so the issuer name extends past it.
_LEGAL_SUFFIXES = {
    "llc", "l.l.c.", "inc", "inc.", "incorporated", "corp", "corp.", "corporation",
    "co", "co.", "company", "lp", "l.p.", "llp", "l.l.p.", "ltd", "ltd.", "limited",
    "plc", "n.v.", "nv", "s.a.", "sa", "gmbh", "ag", "ab", "sas", "s.a.s.", "bv", "b.v.",
    "holdings", "holding", "group", "partners", "lp.",
}

# Instrument / seniority keywords (searched in the full description, lowercased).
_SENIORITY = [
    (r"\bfirst lien\b|\b1st lien\b", "First Lien"),
    (r"\bsecond lien\b|\b2nd lien\b", "Second Lien"),
    (r"\bthird lien\b|\b3rd lien\b", "Third Lien"),
    (r"\bsenior secured\b", "Senior Secured"),
    (r"\bsubordinat", "Subordinated"),
    (r"\bjunior\b|\bmezzanine\b", "Subordinated"),
    (r"\bunsecured\b", "Unsecured"),
    (r"\bsenior\b", "Senior"),
]
_INSTRUMENT = [
    (r"\bdelayed draw\b|\bddtl\b", "Delayed Draw Term Loan"),
    (r"\brevolv|\brevolver\b|\bline of credit\b", "Revolver"),
    (r"\bterm loan\b|\bterm debt\b", "Term Loan"),
    (r"\bnotes?\b", "Notes"),
    (r"\bpreferred\b", "Preferred"),
    (r"\bwarrant", "Warrant"),
    (r"\bcommon\b", "Common Equity"),
    (r"\bequity\b|\bmember(ship)? units?\b|\bllc units?\b|\bshares?\b", "Equity"),
    (r"\bbond", "Notes"),
    (r"\bloan\b", "Loan"),
]

# Subtotal / aggregate rows to drop entirely (not real holdings).
_SUBTOTAL_RE = re.compile(r"\btotal\b|\bsubtotal\b", re.I)
_PCT_ONLY_RE = re.compile(r"^[^a-zA-Z]*\(\s*\d+\.?\d*\s*%\s*\)\s*$")  # e.g. "(2.02%)"
_JUNK_EXACT = {"investment", "investments", "", "nan", "none"}

# Named validation anchors (broadly-held credits found in the recon) — used by --diagnose.
_ANCHORS = ["anaplan", "flexera", "finastra", "avalara", "zendesk", "petvet", "icefall"]


def _fix_mojibake(s: str) -> str:
    """The em-dash separator some filers use comes through as a replacement char (�) or cp1252
    garble. Normalize any of those to ' - ' so the dash-format parser can split on it."""
    for ch in ("�", "", "", "–", "—", "–", "—"):
        s = s.replace(ch, " - ")
    return re.sub(r"\s+", " ", s).strip()


def _norm_tok(t: str) -> str:
    return t.lower().replace("&", "and").strip(" ,.:;-")


# category phrases as token tuples, longest-first (so "health care equipment" beats "health care")
_CATEGORY_TOKENS = sorted(
    ([_norm_tok(t) for t in p.replace("&", "and").split()] for p in _CATEGORY_PHRASES),
    key=len, reverse=True)


def _strip_category_prefix(s: str) -> str:
    """Strip leading category/geo/sector phrases token-by-token, REPEATEDLY — denormalized members
    stack several ('Investments United States Health Care Equipment & Services <issuer>'). Stops at
    the first token that isn't boilerplate, so a real issuer ('ACI Group Holdings, Inc.') survives.
    Operates on the original tokens (preserving case/punctuation) guided by a normalized view."""
    toks = s.split()
    changed = True
    while changed and toks:
        changed = False
        # strip a leading stray percentage / bare number / pure-punctuation token first
        # (denormalized filers embed subtotal percentages, sometimes written "2 2" without a %)
        if toks and (_PCT_TOKEN_RE.match(toks[0]) or _PUNCT_TOKEN_RE.match(toks[0])
                     or re.fullmatch(r"\d+(\.\d+)?", toks[0])):
            toks = toks[1:]
            changed = True
            continue
        norm = [_norm_tok(t) for t in toks]
        for phrase in _CATEGORY_TOKENS:
            n = len(phrase)
            if n <= len(toks) and norm[:n] == phrase:
                toks = toks[n:]
                changed = True
                break
    return " ".join(toks).lstrip(" ,-:").strip()


def parse_issuer(raw: str) -> dict:
    """Parse a raw SOI member string into {issuer_name, instrument_text, parse_ok}.

    Format detection (per row):
      1. comma-delimited  "Name[, Suffix], Sector, Instrument"  -> issuer = leading name parts
      2. em-dash          "Name - Instrument"                   -> issuer = part before dash
      3. fallback         denormalized path / freeform          -> strip category prefix; best guess

    parse_ok=False when the result is empty or still looks like a category bucket (so it can be
    excluded from matching and counted in the coverage report rather than silently mismatched)."""
    if raw is None:
        return {"issuer_name": None, "instrument_text": None, "parse_ok": False}
    s = _fix_mojibake(str(raw))
    low = s.lower().strip()
    if low in _JUNK_EXACT or _PCT_ONLY_RE.match(s) or _SUBTOTAL_RE.search(s):
        return {"issuer_name": None, "instrument_text": None, "parse_ok": False}

    # Cut at the first structured-attribute marker (denormalized filers). Everything from the
    # marker on is attribute noise (type/rate/maturity) we already have as XBRL facts; the issuer
    # is in the head. Keep the tail as instrument_text so seniority/type derivation still sees it.
    struct_tail = None
    m = _STRUCT_MARKER.search(s)
    if m and m.start() > 0:
        struct_tail = s[m.start():]
        s = s[:m.start()].strip(" ,-:")

    s = _strip_category_prefix(s)
    # drop a trailing enumerator ("... Term Loan 2", "... Secured Debt 1") — same instrument, copy n
    s_noenum = re.sub(r"\s+\d{1,2}$", "", s).strip()

    instrument_text = None
    issuer_name = None

    if "," in s_noenum:
        # comma format
        parts = [p.strip() for p in s_noenum.split(",") if p.strip()]
        if parts:
            name_parts = [parts[0]]
            i = 1
            # absorb legal-suffix / parenthetical continuations into the name
            while i < len(parts):
                tok = parts[i].lower().strip().rstrip(".")
                first_word = tok.split()[0] if tok else ""
                if tok in _LEGAL_SUFFIXES or first_word in _LEGAL_SUFFIXES \
                        or parts[i].startswith("(") or "d/b/a" in tok or "f/k/a" in tok \
                        or "fka" in tok or "dba" in tok:
                    # a legal-suffix part may itself carry a dash-delimited instrument tail
                    # ("Inc. - Delayed Draw Loan") — keep only the pre-dash piece in the name.
                    seg, dash, tail = parts[i].partition(" - ")
                    name_parts.append(seg)
                    if dash:
                        instrument_text = tail.strip()
                        i += 1
                        break
                    i += 1
                else:
                    break
            issuer_name = ", ".join(name_parts)
            if instrument_text is None:
                instrument_text = ", ".join(parts[i:]) if i < len(parts) else None
    elif " - " in s_noenum:
        head, _, tail = s_noenum.partition(" - ")
        issuer_name = head.strip()
        instrument_text = tail.strip() or None
    else:
        # freeform / denormalized path — keep as the issuer guess (lossy)
        issuer_name = s_noenum or None
        instrument_text = None

    # the struct tail (type/seniority words) feeds instrument-type/seniority derivation
    if struct_tail:
        instrument_text = (instrument_text + " " + struct_tail) if instrument_text else struct_tail
    ok = bool(issuer_name) and issuer_name.lower() not in _JUNK_EXACT \
        and not _SUBTOTAL_RE.search(issuer_name) and len(issuer_name) > 1
    # Reject names whose core_key (distinctive tokens only) is empty — these are all-boilerplate
    # strings like "in Securities" (core="") or "unaffiliated issuer" (core="") that parse_ok=True
    # but would otherwise form a mega-cluster under a single preposition or boilerplate token.
    # This keeps the "degrade honestly" design: they're counted as unparsed, not force-matched.
    if ok:
        ok = bool(core_key(normalize_issuer(issuer_name) or ""))
    return {"issuer_name": issuer_name, "instrument_text": instrument_text, "parse_ok": ok}


def derive_seniority(text: str) -> str | None:
    if not text:
        return None
    low = text.lower()
    for pat, label in _SENIORITY:
        if re.search(pat, low):
            return label
    return None


def derive_instrument_type(text: str) -> str | None:
    if not text:
        return None
    low = text.lower()
    for pat, label in _INSTRUMENT:
        if re.search(pat, low):
            return label
    return None


_SUFFIX_TOKENS = re.compile(
    r"\b(llc|l\.l\.c\.|inc|incorporated|corp|corporation|co|company|lp|l\.p\.|llp|ltd|limited|"
    r"plc|nv|n\.v\.|sa|s\.a\.|gmbh|ag|holdings?|group|partners|the)\b", re.I)
_PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_issuer(name: str) -> str | None:
    """Aggressive normalization for clustering/anchor tests: drop parentheticals (d/b/a, fka),
    legal suffixes, punctuation; lowercase; collapse whitespace. 'Belnick, LLC (d/b/a ...)' ->
    'belnick'."""
    if not name:
        return None
    s = _PAREN_RE.sub(" ", str(name).lower())
    s = re.sub(r"[^a-z0-9&\s]", " ", s)
    s = _SUFFIX_TOKENS.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


# ---------------------------------------------------------------------------
# Phase 2 — issuer clustering (fuzzy)
# ---------------------------------------------------------------------------

# Generic tokens dropped when building an issuer's CORE KEY (the distinctive-token fingerprint used
# for clustering). Three families: (a) sponsor-backed borrower scaffolding ("X Buyer", "Y Bidco",
# "Z Holdings"); (b) SOI boilerplate that survived Phase-1 parsing on denormalized filers
# (controlled/affiliated/lien/secured/debt/...); (c) single-word GICS sector remnants. Dropping
# these means a denormalized "non controlled affiliated software avalara avalara first lien" and a
# clean "avalara" collapse to the SAME core ('avalara') and merge safely — while boilerplate can no
# longer bridge two unrelated issuers (the bug that produced a 14k-name mega-cluster).
_GENERIC_TOKENS = {
    # sponsor-backed scaffolding / legal
    "holdings", "holding", "holdco", "topco", "midco", "bidco", "buyer", "parent", "intermediate",
    "acquisition", "acquisitions", "group", "capital", "partners", "investors", "investor",
    "investments", "investment", "corp", "company", "inc", "llc", "ltd", "plc", "purchaser",
    "services", "service", "solutions", "systems", "technologies", "technology", "brands",
    "usa", "global", "international", "the", "borrower", "finance", "financial", "newco",
    "national", "american", "ventures", "enterprises", "industries", "and",
    # SOI boilerplate that can survive denormalized parsing
    "non", "controlled", "noncontrolled", "affiliated", "affiliate", "nonaffiliated", "control",
    "lien", "1st", "2nd", "3rd", "first", "second", "third", "senior", "secured", "subordinated",
    "unsecured", "debt", "loan", "loans", "notes", "note", "revolver", "revolving", "term",
    "delayed", "draw", "ddtl", "unitranche", "equity", "common", "preferred", "warrant",
    "warrants", "units", "unit", "interest", "interests", "stock", "shares", "share", "credit",
    "facility", "due", "sofr", "libor", "prime", "euribor", "spread", "fund", "fixed", "last",
    "out", "line", "bank", "related", "party", "undrawn", "commitment", "pssl", "pslf", "rate",
    # structured-template remnants (denormalized filers)
    "issuer", "name", "maturity", "original", "purchase", "date", "variable", "index", "benchmark",
    "floor", "initial", "type", "cash", "all", "corporate", "pik", "index", "reference", "loc",
    "securities", "uk", "warrants", "warrant",
    # prepositions and connectives — too generic to distinguish any issuer; a name whose only
    # remaining tokens are these (e.g. "in securities" after "securities" is already generic)
    # would otherwise create a mega-cluster on the single token "in".
    "in", "of", "to", "at", "on", "for", "by", "as", "or",
    # affiliation-bucket words that escape stripping in some member formats
    "unaffiliated", "nonaffiliate",
    # single-word GICS sector remnants
    "software", "healthcare", "industrials", "materials", "energy", "utilities", "financials",
    "media", "entertainment", "insurance", "banks", "chemicals", "machinery", "transportation",
    "pharmaceuticals", "biotechnology", "automobiles", "automotive", "retail", "distributors",
    "beverages", "internet", "semiconductors", "airlines", "products", "manufacturing",
}


class _UnionFind:
    def __init__(self, items):
        self.p = {x: x for x in items}

    def find(self, x):
        while self.p[x] != x:
            self.p[x] = self.p[self.p[x]]
            x = self.p[x]
        return x

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.p[ra] = rb


def _sig_tokens(norm: str) -> list[str]:
    """Distinctive tokens of a normalized name: length>=2, not generic, not all-digits. Keeping
    2-char tokens matters — short prefixes ('gs'/'pl' in 'gs acquisitionco', 'g1' in 'g1
    therapeutics') are the only thing distinguishing otherwise-identical sponsor names."""
    return [t for t in norm.split()
            if len(t) >= 2 and t not in _GENERIC_TOKENS and not t.isdigit()]


def core_key(norm: str) -> str:
    """The distinctive-token fingerprint of an issuer name: significant tokens, de-duplicated and
    order-independent. 'finastra usa' -> 'finastra'; 'anaplan anaplan' -> 'anaplan'; a denormalized
    'non controlled software avalara avalara first lien' -> 'avalara'. Empty when the name is all
    boilerplate (such rows are left unclustered — honest, not force-matched)."""
    return " ".join(sorted(set(_sig_tokens(norm))))


def cluster_issuers(norms_freq: dict[str, int], threshold: int = 92,
                    max_block: int = 1500) -> dict[str, str]:
    """Cluster normalized issuer names into one canonical name each, in two safe steps:

      1. EXACT CORE — names with the same core_key merge (handles 'finastra'/'finastra usa',
         'anaplan'/'anaplan anaplan', and denormalized variants that collapse to a clean core).
      2. FUZZY CORE — distinct cores that are typo/spacing variants ('lendingpoint'/'lending point',
         'u s renal care'/'us renal care') merge when token_sort_ratio >= threshold. Blocked by a
         shared significant token so it stays tractable; cdist within each block.

    Because clustering keys are DISTINCTIVE tokens only, boilerplate ('controlled', 'lien', 'debt')
    can't bridge unrelated issuers. Canonical = the original norm with the highest row frequency in
    the cluster (ties -> shortest). Returns {issuer_norm -> canonical_norm}. Empty-core norms map to
    themselves (singletons)."""
    uniq = list(norms_freq)
    uf = _UnionFind(uniq)

    core_of = {u: core_key(u) for u in uniq}
    by_core: dict[str, list[str]] = defaultdict(list)
    for u in uniq:
        by_core[core_of[u]].append(u)

    # step 1: exact-core merges (skip the empty core — unclusterable boilerplate-only names)
    core_rep: dict[str, str] = {}   # core_key -> a representative norm
    for c, members in by_core.items():
        if not c:
            continue
        for m in members[1:]:
            uf.union(members[0], m)
        core_rep[c] = members[0]

    # step 2: fuzzy-merge distinct cores, blocked by significant token
    blocks: dict[str, list[str]] = defaultdict(list)
    for c in core_rep:
        for t in set(c.split()):
            blocks[t].append(c)
    seen_pairs: set = set()
    for tok, cores in blocks.items():
        if len(cores) < 2 or len(cores) > max_block:
            continue
        scores = process.cdist(cores, cores, scorer=fuzz.token_sort_ratio, workers=-1)
        for i in range(len(cores)):
            row = scores[i]
            for j in range(i + 1, len(cores)):
                if row[j] >= threshold:
                    uf.union(core_rep[cores[i]], core_rep[cores[j]])

    groups: dict[str, list[str]] = defaultdict(list)
    for u in uniq:
        groups[uf.find(u)].append(u)
    canon: dict[str, str] = {}
    for members in groups.values():
        # prefer the CLEANEST display name: fewest words, then most frequent, then alphabetical.
        # (a denormalized 'common stock 1 0 medeanalytics' loses to the plain 'medeanalytics')
        best = min(members, key=lambda x: (len(x.split()), -norms_freq[x], x))
        for u in members:
            canon[u] = best
    return canon


# ---------------------------------------------------------------------------
# Consolidation
# ---------------------------------------------------------------------------

def load_consolidated() -> pd.DataFrame:
    """Read every per-filing holdings CSV into one DataFrame, parse the issuer field, derive
    seniority / instrument type / price (FV/par). Adds: issuer_name, issuer_norm, instrument_text,
    seniority, instrument_type, price, price_basis, parse_ok."""
    files = sorted(glob.glob(str(HOLDINGS_DIR / "*.csv")))
    if not files:
        raise SystemExit(f"No holdings CSVs in {HOLDINGS_DIR}")
    frames = []
    for f in files:
        try:
            frames.append(pd.read_csv(f, dtype=str))
        except Exception:
            continue
    df = pd.concat(frames, ignore_index=True)

    # numeric coercions
    for c in ("fair_value", "cost", "principal", "spread", "rate", "pik_rate"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    parsed = df["issuer"].apply(parse_issuer).apply(pd.Series)
    df = pd.concat([df, parsed], axis=1)
    df["issuer_norm"] = df["issuer_name"].apply(normalize_issuer)
    # seniority/instrument derived from the full raw member (instrument_text can be empty)
    src = df["issuer"].fillna("") + " " + df["instrument_text"].fillna("")
    df["seniority"] = src.apply(derive_seniority)
    df["instrument_type"] = src.apply(derive_instrument_type)

    # price = fair value as a fraction of par (cents on the dollar), comparable across holders.
    # Two candidate bases: FV/principal (true %-of-par) and FV/cost (mark vs carrying cost).
    #   - Normally use FV/principal.
    #   - PARTIAL-FUNDING FIX: some filers put the full COMMITMENT in `principal` while only part is
    #     drawn (FV ~= cost << principal), which makes FV/principal read like a deep markdown when
    #     the funded slice is actually near par. When FV/principal is implausibly low (<0.5) but
    #     FV/cost is normal (0.6-1.15), the position is partially funded -> use FV/cost. Genuine
    #     deep marks (FV << cost too) are preserved as par-distressed.
    #   - When principal is missing entirely, fall back to FV/cost.
    fv, par, cost = df["fair_value"], df["principal"], df["cost"]
    p_par = (fv / par).where((par > 0) & fv.notna())
    p_cost = (fv / cost).where((cost > 0) & fv.notna())
    # COMMITMENT OVERHANG: principal MUCH larger than amortized cost (>=2.5x) => `principal` is the
    # full commitment with only part drawn; FV/principal would read as a deep (false) markdown, so
    # use FV/cost. The 2.5x bar is deliberately high: a distressed loan bought at a DISCOUNT also
    # has cost < par (e.g. 1.2-1.6x), and must NOT be rerouted — its true mark IS FV/par. Only a
    # genuine undrawn commitment pushes par/cost past ~2.5x.
    overhang = (par > 0) & (cost > 0) & (par > cost * 2.5)
    price = pd.Series(float("nan"), index=df.index, dtype="float64")
    basis = pd.Series(pd.NA, index=df.index, dtype="object")
    # 1) clean par: plausible and not a commitment overhang
    m = p_par.between(0.5, 1.15) & ~overhang
    price[m] = p_par[m]; basis[m] = "par"
    # 2) cost: overhang, or par missing/implausible (covers partial funding + genuine distress)
    m = price.isna() & p_cost.between(0.4, 1.2)
    price[m] = p_cost[m]; basis[m] = "cost"
    # 3) leftover plausible par with no cost to confirm
    m = price.isna() & p_par.between(0.5, 1.15)
    price[m] = p_par[m]; basis[m] = "par"
    df["price"] = price
    df["price_basis"] = basis
    df.loc[(df["price"] <= 0) | (df["price"] > 1.6), "price"] = pd.NA
    return df


def add_clusters(df: pd.DataFrame, threshold: int = 90) -> pd.DataFrame:
    """Add issuer_cluster (canonical normalized name) via Phase-2 fuzzy clustering. Rows that
    didn't parse keep a null cluster."""
    freq = df.loc[df["parse_ok"], "issuer_norm"].dropna().value_counts().to_dict()
    canon = cluster_issuers(freq, threshold=threshold)
    df["issuer_cluster"] = df["issuer_norm"].map(canon)
    return df


# ---------------------------------------------------------------------------
# Phase 3 — issue (tranche) matching within an issuer cluster
# ---------------------------------------------------------------------------

_EQUITY_TYPES = {"Equity", "Preferred", "Common Equity", "Warrant"}
_UNFUNDED_TYPES = {"Revolver", "Delayed Draw Term Loan"}
_YEAR_RE = re.compile(r"(19|20)\d{2}")


def _mat_year(m) -> int | None:
    if m is None or (isinstance(m, float) and pd.isna(m)):
        return None
    hit = _YEAR_RE.search(str(m))
    return int(hit.group()) if hit else None


def _spread_bps(s) -> int | None:
    """Spread in basis points, rounded to the nearest 5bps to absorb tagging noise while keeping
    genuinely different tranches apart (S+500 vs S+575)."""
    if s is None or pd.isna(s):
        return None
    return int(round(float(s) * 10000 / 5.0)) * 5


def _classify(it, price, fair_value) -> str:
    """debt | equity | unfunded — only 'debt' rows with a usable price feed the mark comparison."""
    if it in _EQUITY_TYPES:
        return "equity"
    if (price is None or pd.isna(price)) and it in _UNFUNDED_TYPES:
        return "unfunded"
    return "debt"


def match_issues(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Within each (issuer_cluster, reporting_date), group debt holdings into specific issues
    (tranches) and score match confidence. Returns (holdings_with_issue_cols, issues_summary).

    Grain: the strong, contractual keys are SENIORITY + SPREAD (survive rate resets). Rows with a
    spread are grouped by (seniority, spread_bps). Debt rows WITHOUT a spread are attached to the
    sole spread-issue of their seniority when there's exactly one (co-occurrence inference); else
    they form a seniority-only group. MATURITY is a corroborator: when holders disagree on maturity
    year within an issue, we flag it (mat_conflict) and cap confidence — a likely mis-merge of two
    tranches. Equity/preferred/warrant and unfunded revolver/DDTL rows are tagged and excluded from
    the priced comparison (kept for issuer-level context).

    Confidence (only High/Medium are apples-to-apples price comparisons):
      Single — one holder (well-identified but nothing to compare against).
      High   — spread + seniority known, no maturity conflict, 2-15 holders (club-deal band).
      Medium — spread known but (single-tranche ambiguity: >15 holders, maturity conflict, or
               unknown seniority), or no-spread rows inferred into a spread-issue.
      Low    — matched on seniority only (no spread) / unmatched.
    """
    d = df[df["parse_ok"] & df["issuer_cluster"].notna()].copy()
    d["hold_class"] = [_classify(it, pr, fv) for it, pr, fv in
                       zip(d["instrument_type"], d["price"], d["fair_value"])]
    d["spread_bps"] = d["spread"].apply(_spread_bps)
    d["mat_year"] = d["maturity"].apply(_mat_year)
    d["sen"] = d["seniority"].where(d["seniority"].notna(), None)
    # price that may enter a comparison: funded debt, in the plausible loan band (0.3-1.10; a mark
    # above ~110 usually means accrued interest landed in fair value, not a real premium. Equity is
    # already excluded via hold_class; partial-funding artifacts fixed in load_consolidated.)
    d["cmp_price"] = d["price"].where((d["hold_class"] == "debt")
                                      & d["price"].between(0.3, 1.10))
    d["issue_id"] = None

    issues: list[dict] = []
    for (cl, dt), g in d.groupby(["issuer_cluster", "reporting_date"], sort=False):
        debt = g[g["hold_class"] == "debt"]
        if debt.empty:
            continue
        groups: dict[tuple, list] = {}
        for idx, row in debt[debt["spread_bps"].notna()].iterrows():
            groups.setdefault((row["sen"], int(row["spread_bps"])), []).append(idx)
        sen_keys: dict = defaultdict(list)
        for key in groups:
            sen_keys[key[0]].append(key)
        for idx, row in debt[debt["spread_bps"].isna()].iterrows():
            keys = sen_keys.get(row["sen"])
            if keys and len(keys) == 1:
                groups[keys[0]].append(idx)            # co-occurrence attach
            else:
                groups.setdefault((row["sen"], "nospread"), []).append(idx)

        for (sen, bps), idxs in groups.items():
            sub = d.loc[idxs]
            iid = f"{cl}|{sen}|{bps}|{dt}"
            d.loc[idxs, "issue_id"] = iid
            holders = sub["cik"].nunique()
            # ONE mark per holder: collapse a fund's multiple lots (funded TL + partially-funded
            # piece, etc.) to that fund's median price, so within-fund lot spread is NOT mistaken
            # for cross-holder disagreement. Dispersion is then measured across HOLDERS.
            prices = sub.dropna(subset=["cmp_price"]).groupby("cik")["cmp_price"].median()
            # Robust dispersion: trim holder marks >25pts from the median as artifacts (a lone
            # holder 68pts below the pack on a 1st-lien loan is partial-funding/unit noise, not a
            # real mark). Genuine disagreement keeps multiple holders within the band, so it stays.
            med_all = prices.median() if len(prices) else None
            clean = prices[(prices - med_all).abs() <= 0.25] if med_all is not None else prices
            n_out = int(len(prices) - len(clean))
            myrs = sorted(sub["mat_year"].dropna().unique().tolist())
            mat_conflict = len(myrs) > 1
            spread_known = bps != "nospread"
            if holders < 2:
                conf = "Single"
            elif spread_known and sen and not mat_conflict and 2 <= holders <= 15:
                conf = "High"
            elif spread_known:
                conf = "Medium"
            else:
                conf = "Low"
            issues.append({
                "issuer_cluster": cl, "reporting_date": dt, "issue_id": iid,
                "seniority": sen, "spread_bps": (bps if spread_known else None),
                "maturity_years": ";".join(map(str, myrs)) if myrs else None,
                "mat_conflict": mat_conflict, "n_holders": holders,
                "n_prices": int(len(prices)), "n_clean": int(len(clean)), "n_outliers": n_out,
                "price_median": (round(med_all, 4) if med_all is not None else None),
                "price_min": (round(clean.min(), 4) if len(clean) else None),
                "price_max": (round(clean.max(), 4) if len(clean) else None),
                "price_range_pts": (round((clean.max() - clean.min()) * 100, 2)
                                    if len(clean) >= 2 else None),
                "price_stdev_pts": (round(clean.std() * 100, 2) if len(clean) >= 2 else None),
                "raw_min": (round(prices.min(), 4) if len(prices) else None),
                "raw_max": (round(prices.max(), 4) if len(prices) else None),
                "confidence": conf,
            })
    issues_df = pd.DataFrame(issues)
    return d, issues_df


# ---------------------------------------------------------------------------
# Diagnostic
# ---------------------------------------------------------------------------

def diagnose() -> None:
    df = load_consolidated()
    n = len(df)
    print(f"consolidated rows: {n:,} from {df['cik'].nunique()} CIKs, "
          f"{df['reporting_date'].nunique()} reporting dates\n")

    ok = df["parse_ok"].sum()
    print("PARSE COVERAGE")
    print(f"  issuer parsed (parse_ok): {ok:,} ({ok/n:.1%})")
    print(f"  seniority derived:        {df['seniority'].notna().sum():,} "
          f"({df['seniority'].notna().mean():.1%})")
    print(f"  instrument_type derived:  {df['instrument_type'].notna().sum():,} "
          f"({df['instrument_type'].notna().mean():.1%})")
    print(f"  price computable:         {df['price'].notna().sum():,} "
          f"({df['price'].notna().mean():.1%})  "
          f"[par={ (df['price_basis']=='par').sum():,}, cost={(df['price_basis']=='cost').sum():,}]")
    print(f"  spread present:           {df['spread'].notna().sum():,} "
          f"({df['spread'].notna().mean():.1%})")
    print(f"  maturity present:         {df['maturity'].notna().sum():,} "
          f"({df['maturity'].notna().mean():.1%})")

    # cross-fund overlap on normalized issuer within a reporting date
    g = (df[df["parse_ok"]]
         .groupby(["issuer_norm", "reporting_date"])["cik"].nunique())
    multi = g[g >= 2]
    print("\nCROSS-FUND OVERLAP (parsed issuers)")
    print(f"  (issuer_norm, date) pairs held by >=2 funds: {len(multi):,}")
    print(f"  ... by >=3 funds: {(g>=3).sum():,}   >=5: {(g>=5).sum():,}   >=10: {(g>=10).sum():,}")

    print("\nNAMED-ANCHOR CHECK (max funds holding, any date)")
    for a in _ANCHORS:
        hit = df[df["issuer_norm"].fillna("").str.contains(a, na=False) & df["parse_ok"]]
        if hit.empty:
            print(f"  {a:12} —  not found")
            continue
        top = hit.groupby("reporting_date")["cik"].nunique().max()
        norms = hit["issuer_norm"].value_counts().head(3).index.tolist()
        print(f"  {a:12} max {top} funds/date   norms={norms}")

    print("\nSAMPLE PARSES (random)")
    samp = df[df["parse_ok"]].sample(min(8, int(ok)), random_state=3)
    for _, r in samp.iterrows():
        print(f"  raw={str(r['issuer'])[:60]!r}")
        print(f"      -> issuer={r['issuer_name']!r}  norm={r['issuer_norm']!r}  "
              f"sen={r['seniority']} type={r['instrument_type']} price={r['price']}")

    bad = df[~df["parse_ok"]]
    print(f"\nUNPARSED / DROPPED: {len(bad):,} rows. Sample raw members:")
    for v in bad["issuer"].dropna().drop_duplicates().head(8):
        print(f"  {str(v)[:80]!r}")


def diagnose_clusters(threshold: int = 90) -> None:
    df = add_clusters(load_consolidated(), threshold=threshold)
    parsed = df[df["parse_ok"]]
    n_norm = parsed["issuer_norm"].nunique()
    n_clust = parsed["issuer_cluster"].nunique()
    print(f"clustering @ WRatio>={threshold}")
    print(f"  distinct issuer_norm: {n_norm:,}  ->  clusters: {n_clust:,} "
          f"(merged {n_norm - n_clust:,})\n")

    # cross-fund overlap AFTER clustering (compare to the pre-cluster numbers)
    g = parsed.groupby(["issuer_cluster", "reporting_date"])["cik"].nunique()
    print("CROSS-FUND OVERLAP (clustered)")
    print(f"  (cluster, date) pairs held by >=2 funds: {(g>=2).sum():,}")
    print(f"  ... >=3: {(g>=3).sum():,}   >=5: {(g>=5).sum():,}   >=10: {(g>=10).sum():,}\n")

    print("NAMED-ANCHOR CHECK (norms folded into each anchor's cluster)")
    for a in _ANCHORS:
        hit = parsed[parsed["issuer_norm"].fillna("").str.contains(a, na=False)]
        if hit.empty:
            print(f"  {a:12} not found"); continue
        # the cluster that most of the anchor's rows map to
        cl = hit["issuer_cluster"].mode().iat[0]
        members = sorted(parsed.loc[parsed["issuer_cluster"] == cl, "issuer_norm"].unique(),
                         key=len)
        funds = parsed.loc[parsed["issuer_cluster"] == cl]\
            .groupby("reporting_date")["cik"].nunique().max()
        print(f"  {a:12} -> '{cl}'  ({len(members)} norms, max {funds} funds)")
        print(f"               folds: {members[:6]}{' ...' if len(members) > 6 else ''}")

    print("\nLARGEST CLUSTERS (by #distinct norms folded — eyeball for over-merge)")
    sizes = parsed.groupby("issuer_cluster")["issuer_norm"].nunique().sort_values(ascending=False)
    for cl, k in sizes.head(12).items():
        sample = sorted(parsed.loc[parsed["issuer_cluster"] == cl, "issuer_norm"].unique(),
                        key=len)[:4]
        print(f"  [{k:3} norms] {cl[:34]:34}  e.g. {sample}")

    print("\nSAMPLE MULTI-NORM MERGES (random clusters with 2-5 norms — eyeball for correctness)")
    multi = sizes[(sizes >= 2) & (sizes <= 5)].index
    import pandas as _pd  # local alias to avoid confusion
    for cl in _pd.Series(list(multi)).sample(min(10, len(multi)), random_state=5):
        members = sorted(parsed.loc[parsed["issuer_cluster"] == cl, "issuer_norm"].unique(), key=len)
        print(f"  {cl[:30]:30} <- {members}")


_ANCHOR_CLUSTERS = ["anaplan", "flexera", "avalara", "petvet care centers", "integrity marketing"]


def diagnose_issues(threshold: int = 90) -> None:
    df = add_clusters(load_consolidated(), threshold=threshold)
    holdings, issues = match_issues(df)

    print(f"issues identified: {len(issues):,}")
    by_conf = issues["confidence"].value_counts()
    print("\nBY CONFIDENCE")
    for c in ("High", "Medium", "Low", "Single"):
        sub = issues[issues["confidence"] == c]
        comparable = sub[sub["n_prices"] >= 2]
        print(f"  {c:7} {len(sub):6,}   (with >=2 prices to compare: {len(comparable):,})")

    hi = issues[(issues["confidence"].isin(["High", "Medium"]))
                & (issues["n_holders"] >= 3) & (issues["n_clean"] >= 3)]
    print(f"\nMARK-DISPERSION PAYOFF — High/Medium issues, >=3 CLEAN holders, widest CLEAN spread:")
    print("(outlier prices >25pts from median trimmed as partial-funding/unit artifacts -> n_out)")
    cols = ["issuer_cluster", "reporting_date", "seniority", "spread_bps", "n_holders",
            "n_outliers", "price_median", "price_min", "price_max", "price_range_pts", "confidence"]
    top = hi.sort_values("price_range_pts", ascending=False).head(15)
    with pd.option_context("display.width", 230, "display.max_colwidth", 26):
        print(top[cols].to_string(index=False))

    print("\nTIGHTEST broad club deals (>=6 holders, smallest dispersion — healthy/agreed):")
    tight = hi[hi["n_holders"] >= 6].sort_values("price_range_pts").head(8)
    with pd.option_context("display.width", 230, "display.max_colwidth", 26):
        print(tight[cols].to_string(index=False))

    print("\nANCHOR ISSUE BREAKDOWN (how each anchor splits into tranches, latest common date):")
    for a in _ANCHOR_CLUSTERS:
        sub = issues[issues["issuer_cluster"] == a]
        if sub.empty:
            continue
        dt = sub.loc[sub["n_holders"].idxmax(), "reporting_date"]
        sub = sub[sub["reporting_date"] == dt]
        print(f"\n  {a}  ({dt}):")
        for _, r in sub.sort_values("n_holders", ascending=False).iterrows():
            sp = f"S+{r['spread_bps']}" if pd.notna(r["spread_bps"]) else "no-spread"
            med = f"{r['price_median']*100:.1f}" if pd.notna(r["price_median"]) else "  -"
            rng = f"{r['price_range_pts']:.1f}pt" if pd.notna(r["price_range_pts"]) else "  -"
            print(f"     {str(r['seniority']):14} {sp:9} {r['n_holders']:2}h  "
                  f"med={med:>6}  range={rng:>7}  [{r['confidence']}]")


# ---------------------------------------------------------------------------
# Phase 6 — trend ownership: WHO owns the biggest-moving credits
# ---------------------------------------------------------------------------
# The Trend tab answers "which credits moved most" but drops holder identity on the way to a
# median. This section puts it back, and adds each holder's position size.
#
# TWO GUARDS, both arrived at by measuring rather than assuming:
#
# 1. RECENCY. Trend's net change is (last observed - first observed) over a series that only
#    exists on dates where >=3 funds reported a mark. When an issuer drops below that
#    threshold the series simply STOPS — so a headline "-44pt decline" can be a 2023->2024
#    move on a credit no fund has reported since. Ranking that beside a live 12-quarter slide
#    conflates credit deterioration with holder-coverage CHURN, and asking "who owns it" of
#    such an issuer has no honest answer. Issuers still observed at the corpus's latest date
#    are ranked; the rest are quarantined into their own table WITH the date last seen —
#    recorded rather than silently ranked, and not discarded either.
#
# 2. THE PORTFOLIO DENOMINATOR IS THE TAGGED TOTAL, NOT THE SOI SUM. Summing the schedule of
#    investments looks like the obvious way to size a position against its portfolio, and it
#    is wrong: filers tag INDUSTRY-LEVEL AGGREGATE rows on the same InvestmentIdentifierAxis
#    ("Trading companies & distributors", fair value $311m), so the sum double-counts. MEASURED
#    on the 2026-03 corpus: only 44% of fund-dates fell within +/-10% of the XBRL-tagged
#    `investments_at_fair_value`, 95th percentile 3.18x. So we use the tagged total — the
#    filer's own portfolio figure, and the source of truth the rest of the pipeline already
#    validates — and emit NaN plus a flag where it is absent, never a confident wrong percent.

EXTRACTED_DIR = PROJECT_ROOT / "data" / "extracted"

TREND_MIN_FUNDS = 3        # a date counts only if >=3 funds marked the credit (matches Trend)
TREND_MIN_QUARTERS = 3     # need a few quarters before a "trend" is meaningful (matches Trend)
TREND_MIN_RANGE_PTS = 3.0  # ignore credits that barely moved (matches Trend)
TREND_MIN_NET_PTS = -5.0   # a "decliner" must have fallen at least this far, first->last
TREND_TOP_N = 150          # backstop on issuers given holder detail. Set ABOVE the observed
                           # population (95 live decliners at 2026-03) so it drops nothing in
                           # practice; whatever it does drop is reported on the Coverage tab
                           # rather than silently truncated.
# A single BDC name position above this share of the portfolio is implausible and indicates an
# aggregate row swept into the numerator — flagged, not silently published.
CONCENTRATION_FLAG_PCT = 25.0
# Gap between the raw last-quarter move and the same move on a CONSTANT holder set, beyond
# which the move is substantially a composition change rather than a repricing.
COMPOSITION_FLAG_PTS = 5.0
# Spread across current holders' own marks on the same credit, beyond which the "issue" is more
# likely two different tranches merged than a genuine valuation disagreement.
DISAGREEMENT_FLAG_PTS = 20.0


def tagged_portfolio_fv() -> pd.DataFrame:
    """Portfolio fair value per (cik, reporting_date) from the XBRL-tagged balance sheet in
    data/extracted/*.json. This is the denominator for 'how big is this loan in their book'.
    Returns an EMPTY frame (not an error) when the extracted dir is absent, so the workbook
    still builds — the percentage columns then come out flagged rather than wrong."""
    if not EXTRACTED_DIR.exists():
        return pd.DataFrame(columns=["cik", "reporting_date", "portfolio_fv"])
    rows = []
    for path in EXTRACTED_DIR.glob("*.json"):
        try:
            d = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        fact = (d.get("balance_sheet") or {}).get("investments_at_fair_value") or {}
        val = fact.get("value")
        if val:
            rows.append({"cik": str(d.get("cik")).zfill(10),
                         "reporting_date": str(d.get("reporting_date"))[:10],
                         "portfolio_fv": float(val)})
    if not rows:
        return pd.DataFrame(columns=["cik", "reporting_date", "portfolio_fv"])
    return pd.DataFrame(rows).drop_duplicates(["cik", "reporting_date"])


def issuer_mark_history(fm: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """(marks, fund_counts) pivots — issuer x reporting_date. Built from the same fund_marks
    frame the Trend tab uses, so the two tabs cannot disagree about a credit's trajectory."""
    ok = fm.dropna(subset=["cmp_price"])
    g = (ok.groupby(["issuer_cluster", "reporting_date"])
         .agg(mark=("cmp_price", "median"), funds=("cik", "nunique")).reset_index())
    g = g[g["funds"] >= TREND_MIN_FUNDS]
    marks = g.pivot(index="issuer_cluster", columns="reporting_date", values="mark").mul(100)
    counts = g.pivot(index="issuer_cluster", columns="reporting_date", values="funds")
    cols = sorted(marks.columns)
    marks, counts = marks[cols], counts[cols]
    keep = marks.notna().sum(axis=1) >= TREND_MIN_QUARTERS
    return marks[keep], counts[keep]


def trend_summary(marks: pd.DataFrame, counts: pd.DataFrame) -> pd.DataFrame:
    """One row per trending issuer: net change, swing, the window actually observed, and the
    fund count at each end (so a move measured across a shrinking holder set is visible)."""
    out = []
    for iss, row in marks.iterrows():
        s = row.dropna()
        if len(s) < 2:
            continue
        first, last = s.index[0], s.index[-1]
        out.append({
            "issuer_cluster": iss,
            "net_change_pts": round(s.iloc[-1] - s.iloc[0], 1),
            "range_pts": round(s.max() - s.min(), 1),
            "first_obs": first, "last_obs": last, "n_quarters": int(len(s)),
            "first_mark_pts": round(s.iloc[0], 1), "last_mark_pts": round(s.iloc[-1], 1),
            "peak_mark_pts": round(s.max(), 1), "trough_mark_pts": round(s.min(), 1),
            "funds_at_first": int(counts.at[iss, first]),
            "funds_at_last": int(counts.at[iss, last]),
        })
    df = pd.DataFrame(out)
    return df[df["range_pts"] >= TREND_MIN_RANGE_PTS] if not df.empty else df


# ---------------------------------------------------------------------------
# Phase 7 — per-holder history of the moving credits (the LoanHistory tab)
# ---------------------------------------------------------------------------
# TrendOwners is a snapshot: it says who holds a decliner NOW. This says what each holder's
# mark, position and portfolio weight were at a series of as-of dates, so the arc of a holding
# is one row rather than something to reconstruct.
#
# GRAIN IS ISSUER, NOT TRANCHE — and that was measured, not assumed. The obvious key for "the
# same loan over time" is the issue_id minus its date, i.e. (issuer, seniority, spread). It does
# not survive: of 1,059 such tranches on the Dispersion tab across these six dates, 581 (55%)
# appear on only ONE date and exactly ONE appears on all six, because repricings, amendments and
# tagging drift move the spread. At issuer grain 674 of 1,160 appear on all six. So a tranche-keyed
# history would be mostly blanks. The cost of issuer grain is that a borrower's several tranches
# collapse into one median mark and a summed exposure.
#
# TWO COLUMNS THAT LOOK COMPARABLE ACROSS DATES AND ARE NOT:
#   - TotalExp$mm is summed over the funds that REPORTED that date, and that count moves (53 at
#     2023-12 down to 43 at 2026-06, because funds with later fiscal year-ends have not filed).
#     Medallia's $1,501m -> $950m is partly Monroe Capital dropping out. Hence a Funds column
#     beside it at every date: the denominator of the sum, made visible.
#   - %Port moves on its DENOMINATOR too. Apollo's Medallia weight fell 0.24% -> 0.07% partly
#     because Apollo's own book went $14.6bn -> $25.4bn. Hence FundPort$mm at every date.

HISTORY_DATES = ["2023-12-31", "2024-06-30", "2024-12-31",
                 "2025-06-30", "2025-12-31", "2026-06-30"]
# The as-of date whose mark and total exposure are also bolted onto TrendOwners for a
# side-by-side against "now". A named constant so next year is a one-line edit.
COMPARISON_DATE = "2025-12-31"


def _history_long(holdings: pd.DataFrame, universe: set[str],
                  dates: list[str]) -> pd.DataFrame:
    """Long per-(issuer, fund, date) frame with every metric the history needs. Shared by the
    wide LoanHistory tab and by TrendOwners' as-of columns, so the two cannot disagree."""
    d = holdings[holdings["issuer_cluster"].isin(universe)
                 & (holdings["hold_class"] == "debt")
                 & holdings["cmp_price"].notna()
                 & holdings["reporting_date"].isin(dates)].copy()
    if d.empty:
        return pd.DataFrame()
    d["cik"] = d["cik"].astype(str).str.zfill(10)
    g = (d.groupby(["issuer_cluster", "cik", "fund_name", "reporting_date"])
         .agg(mark=("cmp_price", "median"), pos_fv=("fair_value", "sum")).reset_index())
    ctx = (g.groupby(["issuer_cluster", "reporting_date"])
           .agg(total_exp_fv=("pos_fv", "sum"), funds_reporting=("cik", "nunique"))
           .reset_index())
    g = g.merge(ctx, on=["issuer_cluster", "reporting_date"], how="left")
    g = g.merge(tagged_portfolio_fv(), on=["cik", "reporting_date"], how="left")
    g["mark_pts"] = (g["mark"] * 100).round(1)
    g["pos_mm"] = (g["pos_fv"] / 1e6).round(2)
    g["pct_port"] = (100 * g["pos_fv"] / g["portfolio_fv"]).round(3)
    g["total_exp_mm"] = (g["total_exp_fv"] / 1e6).round(1)
    g["port_mm"] = (g["portfolio_fv"] / 1e6).round(0)
    return g


# (metric column, label prefix) in the order the groups appear on the tab
_HISTORY_METRICS = [
    ("mark_pts", "Mark"), ("pos_mm", "Pos$mm"), ("pct_port", "%Port"),
    ("port_mm", "FundPort$mm"), ("total_exp_mm", "TotalExp$mm"), ("funds_reporting", "Funds"),
]


def loan_history(holdings: pd.DataFrame, universe: set[str],
                 dates: list[str] | None = None) -> pd.DataFrame:
    """Wide history: one row per (issuer, fund), a column group per metric per as-of date.

    A blank is meaningful and deliberately left blank — the fund either did not hold the credit
    on that date or had not filed for it. Never zero-filled, which would read as a real position
    of zero. `Dates present` counts the dates a pair actually has, so a one-date row cannot be
    mistaken for a history."""
    dates = dates or HISTORY_DATES
    g = _history_long(holdings, universe, dates)
    if g.empty:
        return pd.DataFrame()
    idx = ["issuer_cluster", "fund_name"]
    out = None
    for col, prefix in _HISTORY_METRICS:
        piv = g.pivot_table(index=idx, columns="reporting_date", values=col, aggfunc="first")
        piv = piv.reindex(columns=dates)          # every date present, in order, even if empty
        piv.columns = [f"{prefix} {c[:7]}" for c in piv.columns]
        out = piv if out is None else out.join(piv)
    out = out.reset_index()

    present = g.groupby(idx)["reporting_date"].nunique().rename("dates_present")
    out = out.merge(present.reset_index(), on=idx, how="left")

    # Net mark move across whatever dates the pair actually has — the natural sort key, and it
    # must ignore blanks rather than treating a gap as flat.
    mark_cols = [f"Mark {d[:7]}" for d in dates]
    def _net(row):
        s = row[mark_cols].dropna()
        return round(s.iloc[-1] - s.iloc[0], 1) if len(s) >= 2 else None
    out["mark_chg_pts"] = out.apply(_net, axis=1)

    last_pos = [f"Pos$mm {d[:7]}" for d in dates][::-1]
    out["_sort_pos"] = out[last_pos].bfill(axis=1).iloc[:, 0]
    out = out.sort_values(["issuer_cluster", "_sort_pos"], ascending=[True, False]) \
             .drop(columns="_sort_pos")
    # metric-major order: all Marks across time, then all Pos$mm, and so on
    cols = idx + ["dates_present", "mark_chg_pts"]
    for _, prefix in _HISTORY_METRICS:
        cols += [f"{prefix} {d[:7]}" for d in dates]
    return out[cols]


def last_quarter_change(fm: pd.DataFrame, marks: pd.DataFrame) -> pd.DataFrame:
    """Change across an issuer's last two OBSERVED dates, computed two ways: over every fund
    reporting on each date (raw), and over only the funds present on BOTH dates (stable set).

    GUARD 3. A fund that has not filed yet leaves the sample silently, and if its marks sat
    above the pack the median falls without any loan being repriced. MEASURED on this corpus:
    'central parent' shows -38.7pts Mar->Jun, but BOTH Prospect funds — which marked it 86.1
    and 100.0, the two highest — have a June fiscal year-end and had not filed at extraction
    time. Part of that move is composition, not credit. The recency guard cannot catch this,
    because the issuer IS still observed at the latest date; only holding the holder set
    constant separates the two. The stable-set number is the defensible one, and a wide gap
    between raw and stable is flagged rather than left for the reader to discover.
    """
    ok = fm.dropna(subset=["cmp_price"])
    per = ok.groupby(["issuer_cluster", "reporting_date", "cik"])["cmp_price"].median()
    rows = []
    for iss, row in marks.iterrows():
        s = row.dropna()
        if len(s) < 2:
            continue
        d0, d1 = s.index[-2], s.index[-1]
        try:
            a, b = per.loc[(iss, d0)], per.loc[(iss, d1)]
        except KeyError:
            continue
        common = a.index.intersection(b.index)
        raw = s.iloc[-1] - s.iloc[-2]
        stable = ((b[common].median() - a[common].median()) * 100
                  if len(common) else float("nan"))
        rows.append({
            "issuer_cluster": iss, "prev_obs": d0,
            "last_q_chg_pts": round(raw, 1),
            "last_q_chg_stable_pts": (round(stable, 1) if pd.notna(stable) else None),
            "n_stable_holders": int(len(common)),
            "n_left": int(len(a.index.difference(b.index))),
            "n_joined": int(len(b.index.difference(a.index))),
            "composition_gap_pts": (round(raw - stable, 1) if pd.notna(stable) else None),
        })
    return pd.DataFrame(rows)


def trend_ownership(holdings: pd.DataFrame, fm: pd.DataFrame) -> tuple[
        pd.DataFrame, pd.DataFrame, dict]:
    """WHO owns the biggest decliners, and how big the position is in each owner's portfolio.

    Returns (owners, ended, stats):
      owners — one row per (trending issuer, holding fund) at the latest reporting date, with
               the fund's own mark, its deviation from the cross-holder consensus, the position
               in $mm, and the position as a % of that fund's tagged portfolio fair value.
               Also carries when the fund FIRST and LAST held the credit, so holders who sat
               through the whole decline are distinguishable from ones who bought into it.
      ended  — trending issuers whose series stopped before the latest date (guard 1).
      stats  — counts for the Coverage tab, including what the TREND_TOP_N cap dropped.
    """
    marks, counts = issuer_mark_history(fm)
    summ = trend_summary(marks, counts)
    if summ.empty:
        return pd.DataFrame(), pd.DataFrame(), {}

    latest = max(holdings["reporting_date"].dropna().astype(str))
    live = summ[summ["last_obs"] == latest]
    ended = (summ[summ["last_obs"] != latest]
             .sort_values("net_change_pts")
             .reset_index(drop=True))

    decliners = live[live["net_change_pts"] <= TREND_MIN_NET_PTS].sort_values("net_change_pts")
    n_decliners = len(decliners)
    ranked = decliners.head(TREND_TOP_N).copy()
    ranked["rank"] = range(1, len(ranked) + 1)
    targets = set(ranked["issuer_cluster"])

    # Position detail. Debt rows only (equity/unfunded are not what the mark trend measures),
    # restricted to rows that actually matched into an issue so the numerator is real holdings
    # rather than a category aggregate.
    d = holdings[holdings["issuer_cluster"].isin(targets)
                 & (holdings["hold_class"] == "debt")
                 & holdings["issue_id"].notna()].copy()
    d["cik"] = d["cik"].astype(str).str.zfill(10)

    # tenure across ALL dates — who has been holding through the slide
    tenure = (d.groupby(["issuer_cluster", "cik"])
              .agg(first_held=("reporting_date", "min"), last_held=("reporting_date", "max"),
                   quarters_held=("reporting_date", "nunique")).reset_index())

    cur = d[d["reporting_date"] == latest]
    own = (cur.groupby(["issuer_cluster", "cik", "fund_name"])
           .agg(position_fv=("fair_value", "sum"), fund_mark=("cmp_price", "median"),
                n_positions=("fair_value", "size"),
                seniority=("seniority", lambda s: "; ".join(sorted(set(s.dropna())))or None))
           .reset_index())
    own = own.merge(tenure, on=["issuer_cluster", "cik"], how="left")

    port = tagged_portfolio_fv()
    port = port[port["reporting_date"] == latest][["cik", "portfolio_fv"]]
    own = own.merge(port, on="cik", how="left")

    own["pct_of_portfolio"] = (100 * own["position_fv"] / own["portfolio_fv"]).round(3)
    # Guard: no denominator -> no percentage (never a fabricated one); implausible share ->
    # keep the number but say so, since it signals an aggregate row in the numerator.
    flags = []
    for _, r in own.iterrows():
        f = []
        if pd.isna(r["portfolio_fv"]):
            f.append("no_portfolio_total")
        elif r["pct_of_portfolio"] > CONCENTRATION_FLAG_PCT:
            f.append("implausible_concentration")
        if pd.isna(r["fund_mark"]):
            f.append("no_usable_mark")
        flags.append("; ".join(f))
    own["flags"] = flags
    own.loc[own["portfolio_fv"].isna(), "pct_of_portfolio"] = pd.NA

    # Issuer-level context: total exposure across every BDC we can see, holder count, and how
    # far apart the holders are TODAY. A wide current spread is the tell for a tranche mismatch:
    # MEASURED, 'truist insurance' holders agreed within 1.5pts in March and 50.8pts in June,
    # which is not how five sophisticated managers disagree about one loan. Flagged so a
    # mis-merge is not read as a repricing.
    roll = (own.groupby("issuer_cluster")
            .agg(n_holders_now=("cik", "nunique"),
                 total_exposure_fv=("position_fv", "sum"),
                 holder_spread_pts=("fund_mark", lambda s: round(
                     (s.max() - s.min()) * 100, 1) if s.notna().sum() >= 2 else None),
                 ).reset_index())

    # As-of columns: the same holder's mark and the same credit's total exposure at
    # COMPARISON_DATE, for a side-by-side against "now". Sourced from _history_long so these
    # cannot drift from the LoanHistory tab.
    asof = _history_long(holdings, targets, [COMPARISON_DATE])
    if not asof.empty:
        asof = asof[["issuer_cluster", "cik", "mark_pts", "total_exp_mm"]].rename(columns={
            "mark_pts": "mark_at_comparison", "total_exp_mm": "exposure_at_comparison"})
        own = own.merge(asof, on=["issuer_cluster", "cik"], how="left")
    else:
        own["mark_at_comparison"] = pd.NA
        own["exposure_at_comparison"] = pd.NA

    lq = last_quarter_change(fm, marks)
    own = own.merge(ranked, on="issuer_cluster", how="inner").merge(
        roll, on="issuer_cluster", how="left").merge(lq, on="issuer_cluster", how="left")
    # Guard 3: say so when the last-quarter move is substantially a holder-set change.
    own.loc[own["composition_gap_pts"].abs() >= COMPOSITION_FLAG_PTS, "flags"] = (
        own["flags"].fillna("") + "; composition_shift").str.strip("; ")
    own.loc[own["holder_spread_pts"] >= DISAGREEMENT_FLAG_PTS, "flags"] = (
        own["flags"].fillna("") + "; holders_disagree_check_tranche").str.strip("; ")
    own["fund_mark_pts"] = (own["fund_mark"] * 100).round(1)
    own["dev_vs_consensus_pts"] = (own["fund_mark_pts"] - own["last_mark_pts"]).round(1)
    own["position_mm"] = (own["position_fv"] / 1e6).round(2)
    own["total_exposure_mm"] = (own["total_exposure_fv"] / 1e6).round(1)
    own["portfolio_mm"] = (own["portfolio_fv"] / 1e6).round(0)
    own["held_since_peak"] = (own["first_held"] <= own["first_obs"]).map(
        {True: "yes", False: "no"})
    own = own.sort_values(["rank", "position_fv"], ascending=[True, False])

    stats = {
        "latest reporting date": latest,
        "trending issuers (range>=%.0fpts)" % TREND_MIN_RANGE_PTS: int(len(summ)),
        "  still observed at latest date": int(len(live)),
        "  series ENDED earlier (quarantined)": int(len(ended)),
        "decliners (net<=%.0fpts) among live" % TREND_MIN_NET_PTS: n_decliners,
        "  given holder detail (TREND_TOP_N cap)": int(len(ranked)),
        "  DROPPED by the cap": int(max(0, n_decliners - TREND_TOP_N)),
        "owner rows": int(len(own)),
        "  missing a portfolio denominator": int(own["portfolio_fv"].isna().sum()),
        "  flagged implausible concentration":
            int((own["flags"].str.contains("implausible", na=False)).sum()),
        "issuers flagged composition_shift (guard 3)":
            int(own.loc[own["flags"].str.contains("composition_shift", na=False),
                        "issuer_cluster"].nunique()),
        "issuers flagged holders_disagree (tranche check)":
            int(own.loc[own["flags"].str.contains("holders_disagree", na=False),
                        "issuer_cluster"].nunique()),
    }
    return own, ended, stats


# ---------------------------------------------------------------------------
# Phase 4 — cross-holder mark-comparison workbook
# ---------------------------------------------------------------------------

WORKBOOK = OUT_DIR / "holdings_marks_comparison.xlsx"


def _spread_label(bps) -> str:
    return f"S+{int(bps)}" if pd.notna(bps) else "—"


def fund_marks(holdings: pd.DataFrame, issues: pd.DataFrame) -> pd.DataFrame:
    """Per-fund mark for each matched issue (a fund's lots collapsed to their median price), with
    deviation from the issue's consensus median. The actionable 'who marks where' table."""
    h = holdings[(holdings["hold_class"] == "debt") & holdings["cmp_price"].notna()
                 & holdings["issue_id"].notna()]
    fm = (h.groupby(["issue_id", "cik", "fund_name"])["cmp_price"].median().reset_index())
    meta = issues[["issue_id", "issuer_cluster", "reporting_date", "seniority", "spread_bps",
                   "price_median", "n_holders", "price_range_pts", "confidence"]]
    fm = fm.merge(meta, on="issue_id", how="left")
    fm["dev_pts"] = (fm["cmp_price"] - fm["price_median"]) * 100
    fm["stance"] = pd.cut(fm["dev_pts"], [-999, -1, 1, 999],
                          labels=["Cheap (marks low)", "In line", "Rich (marks high)"])
    return fm


def _style_sheet(ws, n_cols: int, pct_cols: list[int] | None = None,
                 widths: dict[int, int] | None = None) -> None:
    """Bold/filled header, freeze header row, autofilter, 1-dp number format on mark columns."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    for col, w in (widths or {}).items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for col in (pct_cols or []):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "0.0"


def load_matched_cache() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Read back the (holdings, issues) that `--build` already wrote, instead of recomputing
    them. `--build` writes exactly `match_issues(...)`'s two outputs, so this is the same data
    by construction — the point of the CLAUDE.md rule about deriving from the source of truth
    rather than recomputing it. Clustering + issue matching is ~70 minutes on the full corpus,
    so this turns a workbook-only change from an hour into seconds. Raises if the cache is
    absent, so it can never silently fall back to stale or partial input."""
    hp, ip = OUT_DIR / "holdings_matched.csv", OUT_DIR / "issues.csv"
    for p in (hp, ip):
        if not p.exists():
            raise SystemExit(f"--from-cache needs {p.name}; run --build first")
    return (pd.read_csv(hp, low_memory=False), pd.read_csv(ip, low_memory=False))


def build_workbook(threshold: int = 92, from_cache: bool = False) -> None:
    """Phase-4 deliverable: a cross-holder mark-comparison workbook (data/dataset/
    holdings_marks_comparison.xlsx). Marks are expressed in POINTS of par (price x 100)."""
    if from_cache:
        holdings, issues = load_matched_cache()
    else:
        df = add_clusters(load_consolidated(), threshold=threshold)
        holdings, issues = match_issues(df)
    fm = fund_marks(holdings, issues)

    # points-of-par views
    iss = issues.copy()
    for col in ("price_median", "price_min", "price_max"):
        iss[col + "_pts"] = (iss[col] * 100).round(1)
    iss["spread"] = iss["spread_bps"].apply(_spread_label)
    base_cols = {
        "issuer_cluster": "Issuer", "reporting_date": "Date", "seniority": "Seniority",
        "spread": "Spread", "maturity_years": "Maturity", "n_holders": "Holders",
        "n_outliers": "Outliers", "price_median_pts": "Median", "price_min_pts": "Min",
        "price_max_pts": "Max", "price_range_pts": "Range(pts)", "price_stdev_pts": "Stdev(pts)",
        "confidence": "Confidence", "issue_id": "issue_id",
    }
    hi = iss[iss["confidence"].isin(["High", "Medium"])]

    # --- tab data ---
    dispersion = (hi[(hi["n_holders"] >= 3) & (hi["n_clean"] >= 3)]
                  .sort_values("price_range_pts", ascending=False))[list(base_cols)]
    consensus = (hi[(hi["n_holders"] >= 5) & (hi["price_range_pts"] <= 2)]
                 .sort_values("n_holders", ascending=False))[list(base_cols)]

    # holder detail for the meaningfully-dispersed issues (range >= 3 pts)
    disp_ids = set(hi[(hi["n_holders"] >= 3) & (hi["price_range_pts"] >= 3)]["issue_id"])
    hd = fm[fm["issue_id"].isin(disp_ids)].copy()
    hd["spread"] = hd["spread_bps"].apply(_spread_label)
    hd["fund_mark_pts"] = (hd["cmp_price"] * 100).round(1)
    hd["issue_median_pts"] = (hd["price_median"] * 100).round(1)
    hd["dev_pts"] = hd["dev_pts"].round(1)
    hd = hd.sort_values(["issuer_cluster", "reporting_date", "issue_id", "dev_pts"])
    hd_cols = {"issuer_cluster": "Issuer", "reporting_date": "Date", "seniority": "Seniority",
               "spread": "Spread", "fund_name": "Fund", "fund_mark_pts": "Fund mark",
               "issue_median_pts": "Issue median", "dev_pts": "Deviation(pts)", "stance": "Stance",
               "confidence": "Confidence", "issue_id": "issue_id"}

    # issuer-grain rollup (both grains, per the plan)
    pr = issues[issues["price_median"].notna()].copy()
    isum = (pr.groupby(["issuer_cluster", "reporting_date"])
            .agg(max_holders=("n_holders", "max"), priced_tranches=("issue_id", "nunique"),
                 issuer_median=("price_median", "median"),
                 widest_range=("price_range_pts", "max")).reset_index())
    isum = isum[isum["max_holders"] >= 2].sort_values("max_holders", ascending=False)
    isum["issuer_median_pts"] = (isum["issuer_median"] * 100).round(1)
    isum["widest_range_pts"] = isum["widest_range"].round(1)
    isum["dispersed?"] = (isum["widest_range"] >= 5).map({True: "yes", False: ""})
    isum_cols = {"issuer_cluster": "Issuer", "reporting_date": "Date",
                 "max_holders": "Max holders", "priced_tranches": "Priced tranches",
                 "issuer_median_pts": "Issuer median", "widest_range_pts": "Widest tranche range",
                 "dispersed?": "Dispersed(>=5pts)?"}

    # anchors validation
    anchor_rows = []
    for a in _ANCHOR_CLUSTERS + ["finastra", "icefall parent", "zendesk"]:
        sub = iss[iss["issuer_cluster"] == a]
        if sub.empty:
            continue
        dt = sub.loc[sub["n_holders"].idxmax(), "reporting_date"]
        for _, r in sub[sub["reporting_date"] == dt].sort_values("n_holders", ascending=False).iterrows():
            anchor_rows.append({"Issuer": a, "Date": dt, "Seniority": r["seniority"],
                                "Spread": r["spread"], "Holders": r["n_holders"],
                                "Median": r["price_median_pts"], "Range(pts)": r["price_range_pts"],
                                "Confidence": r["confidence"]})
    anchors = pd.DataFrame(anchor_rows)

    # summary stats for the overview
    debt = holdings[holdings["hold_class"] == "debt"]
    matched_hi = holdings[holdings["issue_id"].isin(set(hi["issue_id"]))]
    stats = {
        "Funds (CIKs)": int(holdings["cik"].nunique()),
        "Reporting dates": int(holdings["reporting_date"].nunique()),
        "Issuer clusters": int(holdings["issuer_cluster"].nunique()),
        "Debt holdings": int(len(debt)),
        "Matched issues (all)": int(len(issues)),
        "  High confidence": int((issues["confidence"] == "High").sum()),
        "  Medium": int((issues["confidence"] == "Medium").sum()),
        "  Low": int((issues["confidence"] == "Low").sum()),
        "  Single holder": int((issues["confidence"] == "Single").sum()),
        "Comparable issues (>=2 holders, High/Med)": int((hi["n_holders"] >= 2).sum()),
        "Dispersed issues (>=3 holders, range>=5pts)":
            int(len(hi[(hi["n_holders"] >= 3) & (hi["price_range_pts"] >= 5)])),
    }

    # ---- Phase 5: coverage stats, stratified review sample, period-over-period trend ----
    debt_priced = debt[debt["cmp_price"].notna()]
    in_hi = debt_priced["issue_id"].isin(set(hi[hi["n_holders"] >= 2]["issue_id"]))
    g_cd = debt.groupby(["issuer_cluster", "reporting_date"])["cik"].nunique()
    rng = hi[hi["n_holders"] >= 3]["price_range_pts"].dropna()
    coverage = {
        "MATCH RATE": "",
        "Debt holdings (rows)": int(len(debt)),
        "  with a comparable price": int(len(debt_priced)),
        "  matched into a >=2-holder High/Med issue": int(in_hi.sum()),
        "  match rate (% of priced debt)":
            (round(100 * in_hi.sum() / len(debt_priced), 1) if len(debt_priced) else 0),
        "ISSUERS / ISSUES": "",
        "Issuer clusters (parsed)": int(holdings["issuer_cluster"].nunique()),
        "  held by >=2 funds on a date": int(g_cd[g_cd >= 2].index.get_level_values(0).nunique()),
        "(cluster, date) pairs >=2 funds": int((g_cd >= 2).sum()),
        "High/Med comparable issues (>=2 holders)": int((hi["n_holders"] >= 2).sum()),
        "DISPERSION BANDS (High/Med, >=3 holders)": "",
        "  tight (<2 pts)": int((rng < 2).sum()),
        "  moderate (2-5 pts)": int(((rng >= 2) & (rng < 5)).sum()),
        "  wide (5-10 pts)": int(((rng >= 5) & (rng < 10)).sum()),
        "  very wide (10-25 pts)": int(((rng >= 10) & (rng < 25)).sum()),
        "  extreme (>=25 pts — review)": int((rng >= 25).sum()),
        "NOTE": "Matching is EXACT reporting-date; funds with different fiscal quarter-ends "
                "do not co-match (a coverage limit, not an error).",
    }

    # stratified review sample (confidence x dispersion band), with each issue's holder marks inline
    fm_ok = fm.dropna(subset=["cmp_price"]).copy()
    fm_ok["lab"] = (fm_ok["fund_name"].astype(str).str[:16] + ":"
                    + (fm_ok["cmp_price"] * 100).round(0).astype(int).astype(str))
    marks_by_issue = (fm_ok.sort_values("cmp_price").groupby("issue_id")["lab"]
                      .apply(lambda s: "  ".join(s.head(10))))
    band = pd.cut(iss["price_range_pts"], [-1, 2, 10, 9999],
                  labels=["tight", "moderate", "wide"])
    samp_src = iss.assign(band=band)
    samp_rows = []
    for conf in ("High", "Medium", "Low"):
        for bnd in ("tight", "moderate", "wide"):
            pool = samp_src[(samp_src["confidence"] == conf) & (samp_src["band"] == bnd)
                            & (samp_src["n_holders"] >= 2)]
            take = pool.sample(min(6, len(pool)), random_state=11) if len(pool) else pool
            for _, r in take.iterrows():
                samp_rows.append({
                    "Confidence": conf, "Band": bnd, "Issuer": r["issuer_cluster"],
                    "Date": r["reporting_date"], "Seniority": r["seniority"],
                    "Spread": r["spread"], "Holders": r["n_holders"],
                    "Median": r["price_median_pts"], "Range(pts)": r["price_range_pts"],
                    "Holder marks (fund:pts)": marks_by_issue.get(r["issue_id"], ""),
                    "Verdict (Y/N)": "", "Notes": "",
                })
    review_sample = pd.DataFrame(samp_rows)

    # period-over-period MEDIAN mark per issuer (>=3 funds that date), to spot drift
    idate = (fm_ok.groupby(["issuer_cluster", "reporting_date"])
             .agg(mark=("cmp_price", "median"), funds=("cik", "nunique")).reset_index())
    idate = idate[idate["funds"] >= 3]
    piv = idate.pivot(index="issuer_cluster", columns="reporting_date", values="mark").mul(100).round(1)
    piv = piv[sorted(piv.columns)]
    piv = piv[piv.notna().sum(axis=1) >= 3]          # need a few quarters of history

    def _net(row):
        s = row.dropna()
        return round(s.iloc[-1] - s.iloc[0], 1) if len(s) >= 2 else None
    net = piv.apply(_net, axis=1)
    rng_t = (piv.max(axis=1) - piv.min(axis=1)).round(1)
    trend = piv.copy()
    trend.insert(0, "Range over time", rng_t)
    trend.insert(0, "Net change (first->last)", net)
    trend = trend[trend["Range over time"] >= 3].sort_values("Net change (first->last)")
    trend = trend.reset_index().rename(columns={"issuer_cluster": "Issuer"})

    # ---- Phase 6: who owns the biggest-moving credits (see the section header for the
    # two guards this applies — recency, and the tagged-total denominator) ----
    owners, ended, tstats = trend_ownership(holdings, fm)
    own_cols = {
        "rank": "Rank", "issuer_cluster": "Issuer", "net_change_pts": "Net change(pts)",
        "first_mark_pts": "Mark at start", "last_mark_pts": "Mark now",
        "n_quarters": "Quarters", "n_holders_now": "Holders now",
        "last_q_chg_pts": "Last qtr(pts)",
        "last_q_chg_stable_pts": "Last qtr, same holders(pts)",
        "composition_gap_pts": "Composition gap(pts)",
        "holder_spread_pts": "Holder spread(pts)",
        "total_exposure_mm": "Total BDC exposure($mm)",
        "exposure_at_comparison": f"Total BDC exposure @{COMPARISON_DATE}($mm)",
        "fund_name": "Fund",
        "fund_mark_pts": "Fund mark", "mark_at_comparison": f"Fund mark @{COMPARISON_DATE}",
        "dev_vs_consensus_pts": "Dev vs consensus(pts)",
        "position_mm": "Position($mm)", "pct_of_portfolio": "% of fund portfolio",
        "portfolio_mm": "Fund portfolio($mm)", "seniority": "Seniority",
        "n_positions": "Lots", "first_held": "First held", "quarters_held": "Quarters held",
        "held_since_peak": "Held since start?", "flags": "Flags",
    }
    ended_cols = {
        "issuer_cluster": "Issuer", "net_change_pts": "Net change(pts)",
        "range_pts": "Range(pts)", "first_obs": "First observed", "last_obs": "LAST OBSERVED",
        "n_quarters": "Quarters", "first_mark_pts": "Mark at start",
        "last_mark_pts": "Mark when last seen", "funds_at_first": "Funds at start",
        "funds_at_last": "Funds at end",
    }
    if tstats:
        coverage["TREND OWNERSHIP (Phase 6)"] = ""
        coverage.update(tstats)

    # ---- Phase 7: per-holder history of those same decliners, wide by as-of date ----
    hist = (loan_history(holdings, set(owners["issuer_cluster"]))
            if not owners.empty else pd.DataFrame())
    if not hist.empty:
        mark_cols = [f"Mark {d[:7]}" for d in HISTORY_DATES]
        blanks = hist[mark_cols].isna().sum().sum()
        coverage["LOAN HISTORY (Phase 7)"] = ""
        coverage.update({
            "as-of dates": ", ".join(HISTORY_DATES),
            "(issuer, fund) rows": int(len(hist)),
            "  with >=4 of the as-of dates": int((hist["dates_present"] >= 4).sum()),
            "  with only ONE date (not a history)": int((hist["dates_present"] == 1).sum()),
            "blank mark cells (fund did not hold, or had not filed)":
                f"{int(blanks)} of {len(hist)*len(mark_cols)} "
                f"({blanks/(len(hist)*len(mark_cols)):.0%})",
            "NOTE (history)": "TotalExp$mm is summed over the funds that REPORTED each date - "
                              "read the Funds column beside it. %Port moves on its denominator "
                              "too - read FundPort$mm beside it.",
        })

    _write_workbook(dispersion, base_cols, consensus, hd, hd_cols, isum, isum_cols,
                    anchors, stats, coverage, review_sample, trend,
                    owners, own_cols, ended, ended_cols, hist)
    print(f"wrote {WORKBOOK}")
    if tstats:
        print(f"  TrendOwners: {len(owners)} owner rows over "
              f"{owners['issuer_cluster'].nunique() if not owners.empty else 0} issuers; "
              f"TrendEnded: {len(ended)} quarantined")


def _write_workbook(dispersion, base_cols, consensus, hd, hd_cols, isum, isum_cols,
                    anchors, stats, coverage, review_sample, trend,
                    owners=None, own_cols=None, ended=None, ended_cols=None,
                    hist=None) -> None:
    import openpyxl
    from openpyxl.styles import Font

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CAVEATS = [
        "Cross-BDC mark comparison — how differently BDCs value the SAME credit at the SAME date.",
        "Marks are in POINTS OF PAR (100 = par). price = fair value / par, with a cost-based",
        "fallback when only part of a commitment is drawn (commitment-overhang fix).",
        "",
        "Read the tabs:",
        "  Dispersion    — matched issues where holders DISAGREE most (the analytical payoff).",
        "  Consensus     — broadly-held issues with tight agreement (the 'market mark').",
        "  HolderDetail  — per-fund mark vs the issue median for dispersed issues (who's rich/cheap).",
        "  IssuerSummary — issuer-grain rollup across all tranches.",
        "  Anchors       — known broadly-held credits, for validation.",
        "  Coverage      — match-rate, confidence + dispersion-band stats.",
        "  ReviewSample  — stratified hand-check sample (fill the Verdict column).",
        "  Trend         — median mark (>=3 holders) per issuer over time (spot deterioration).",
        "  TrendOwners   — WHO owns the biggest decliners, each holder's mark, position $mm and",
        "                  share of their portfolio. Ranked only for credits still reported at the",
        "                  latest date; 'Held since start?' = held it before the slide began.",
        "  TrendEnded    — trending credits whose series STOPPED before the latest date. Their",
        "                  net change is measured over an old window and a shrinking holder set,",
        "                  so they are quarantined here rather than ranked. Check LAST OBSERVED.",
        "  LoanHistory   — the SAME decliners, one row per (issuer, fund), read left to right",
        "                  through six as-of dates: each holder's mark, position $mm, % of its",
        "                  portfolio, its portfolio size, the credit's total BDC exposure, and how",
        "                  many funds that exposure covers. A BLANK means the fund did not hold",
        "                  the credit then or had not filed — never zero-filled. 'Dates present'",
        "                  counts what a row actually has, so a one-date row is not a history.",
        "                  Grain is ISSUER, not tranche: measured, 55% of (issuer,seniority,spread)",
        "                  tranches survive only one of these dates, so a tranche-keyed history",
        "                  would be mostly blank. Several tranches of one borrower are therefore",
        "                  collapsed into a median mark and a summed exposure.",
        "",
        "Confidence: High = seniority+spread matched, no maturity conflict, 2-15 holders.",
        "            Medium = spread matched but ambiguous; Low = seniority only; Single = 1 holder.",
        "Caveats: marks are manager estimates on illiquid Level-3 loans; different fiscal quarter-",
        "ends are aligned only within the same reporting date; a lone holder >25pts off the median",
        "is trimmed as a unit/partial-funding artifact (counted in 'Outliers'). Best-effort match,",
        "not exact reconciliation — investigate wide spreads before relying on them.",
        "",
        "On TrendOwners, read THREE columns before trusting a decline:",
        "  'Last qtr, same holders' — the move recomputed over only the funds reporting on BOTH",
        "    dates. 'Composition gap' is how much of the raw move was a holder-set change instead",
        "    (a fund with a later fiscal year-end drops out and takes its mark with it).",
        "  'Holder spread' — how far apart current holders are. Above ~20pts, suspect two",
        "    different tranches merged into one issue rather than a real disagreement.",
        "",
        "'% of fund portfolio' divides the position by the fund's XBRL-TAGGED investments at fair",
        "value, NOT by the sum of its schedule of investments. Measured reason: filers tag",
        "industry-level AGGREGATE rows on the same axis, so the SOI sum double-counts (only 44% of",
        "fund-dates came within +/-10% of the tagged total). Blank = no tagged total for that",
        "fund-date; see the Flags column.",
    ]
    with pd.ExcelWriter(WORKBOOK, engine="openpyxl") as xl:
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Overview", index=False)
        dispersion.rename(columns=base_cols).to_excel(xl, sheet_name="Dispersion", index=False)
        consensus.rename(columns=base_cols).to_excel(xl, sheet_name="Consensus", index=False)
        hd[list(hd_cols)].rename(columns=hd_cols).to_excel(xl, sheet_name="HolderDetail", index=False)
        isum[list(isum_cols)].rename(columns=isum_cols).to_excel(
            xl, sheet_name="IssuerSummary", index=False)
        if not anchors.empty:
            anchors.to_excel(xl, sheet_name="Anchors", index=False)
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Coverage", index=False)
        if not review_sample.empty:
            review_sample.to_excel(xl, sheet_name="ReviewSample", index=False)
        if not trend.empty:
            trend.to_excel(xl, sheet_name="Trend", index=False)
        if owners is not None and not owners.empty:
            owners[list(own_cols)].rename(columns=own_cols).to_excel(
                xl, sheet_name="TrendOwners", index=False)
        if ended is not None and not ended.empty:
            ended[list(ended_cols)].rename(columns=ended_cols).to_excel(
                xl, sheet_name="TrendEnded", index=False)
        if hist is not None and not hist.empty:
            hist.rename(columns={"issuer_cluster": "Issuer", "fund_name": "Fund",
                                 "dates_present": "Dates present",
                                 "mark_chg_pts": "Mark chg(pts)"}).to_excel(
                xl, sheet_name="LoanHistory", index=False)

        wb = xl.book
        ov = wb["Overview"]
        ov["A1"] = "BDC Holdings — Cross-Holder Mark Comparison"
        ov["A1"].font = Font(bold=True, size=14)
        r = 3
        for line in CAVEATS:
            ov.cell(row=r, column=1, value=line); r += 1
        r += 1
        ov.cell(row=r, column=1, value="SUMMARY").font = Font(bold=True); r += 1
        for k, v in stats.items():
            ov.cell(row=r, column=1, value=k)
            ov.cell(row=r, column=2, value=v); r += 1
        ov.column_dimensions["A"].width = 70
        ov.column_dimensions["B"].width = 14

        # mark columns (1-dp) per sheet
        _style_sheet(wb["Dispersion"], len(base_cols), pct_cols=[8, 9, 10, 11, 12],
                     widths={1: 30, 2: 12, 3: 12, 4: 9, 5: 12, 14: 34})
        _style_sheet(wb["Consensus"], len(base_cols), pct_cols=[8, 9, 10, 11, 12],
                     widths={1: 30, 2: 12, 3: 12, 4: 9, 5: 12, 14: 34})
        _style_sheet(wb["HolderDetail"], len(hd_cols), pct_cols=[6, 7, 8],
                     widths={1: 28, 2: 12, 3: 12, 4: 9, 5: 30, 9: 18, 11: 34})
        _style_sheet(wb["IssuerSummary"], len(isum_cols), pct_cols=[5, 6],
                     widths={1: 30, 2: 12, 3: 12, 4: 14, 5: 14, 6: 18, 7: 16})
        if not anchors.empty:
            _style_sheet(wb["Anchors"], anchors.shape[1], pct_cols=[6, 7],
                         widths={1: 22, 2: 12, 3: 12, 4: 9})

        # Coverage tab (key/value, section headers bold)
        cv = wb["Coverage"]
        cv["A1"] = "Coverage & match-rate"
        cv["A1"].font = Font(bold=True, size=13)
        rr = 3
        for k, v in coverage.items():
            cell = cv.cell(row=rr, column=1, value=k)
            if v == "":           # a section header
                cell.font = Font(bold=True)
            else:
                cv.cell(row=rr, column=2, value=v)
            rr += 1
        cv.column_dimensions["A"].width = 48
        cv.column_dimensions["B"].width = 60

        if not review_sample.empty:
            _style_sheet(wb["ReviewSample"], review_sample.shape[1], pct_cols=[8, 9],
                         widths={1: 9, 2: 9, 3: 26, 4: 12, 5: 12, 6: 8, 10: 60, 11: 11, 12: 26})
        if not trend.empty:
            # cols: Issuer, Net change, Range, then one col per date — all numeric except Issuer
            _style_sheet(wb["Trend"], trend.shape[1],
                         pct_cols=list(range(2, trend.shape[1] + 1)), widths={1: 30})
        if owners is not None and not owners.empty:
            # Column positions are derived from own_cols BY NAME, never hard-coded — adding a
            # column upstream would otherwise silently number-format the wrong ones.
            keys = list(own_cols)
            def _at(name: str) -> int:
                return keys.index(name) + 1
            one_dp = [_at(k) for k in (
                "net_change_pts", "first_mark_pts", "last_mark_pts", "last_q_chg_pts",
                "last_q_chg_stable_pts", "composition_gap_pts", "holder_spread_pts",
                "total_exposure_mm", "fund_mark_pts", "dev_vs_consensus_pts", "position_mm")
                if k in keys]
            _style_sheet(wb["TrendOwners"], len(own_cols), pct_cols=one_dp,
                         widths={_at("issuer_cluster"): 26, _at("fund_name"): 32,
                                 _at("last_q_chg_stable_pts"): 16,
                                 _at("composition_gap_pts"): 14,
                                 _at("pct_of_portfolio"): 15, _at("flags"): 42})
            for row in range(2, wb["TrendOwners"].max_row + 1):
                wb["TrendOwners"].cell(
                    row=row, column=_at("pct_of_portfolio")).number_format = "0.00"
        if ended is not None and not ended.empty:
            _style_sheet(wb["TrendEnded"], len(ended_cols), pct_cols=[2, 3, 7, 8],
                         widths={1: 30, 4: 14, 5: 15, 7: 14, 8: 20})
        if hist is not None and not hist.empty:
            ws = wb["LoanHistory"]
            hcols = list(hist.columns)
            # 1-dp everywhere except %Port (3-dp, the values are small) and Funds (integer)
            one_dp = [i + 1 for i, c in enumerate(hcols)
                      if c.startswith(("Mark ", "Pos$mm ", "FundPort$mm ", "TotalExp$mm "))
                      or c == "mark_chg_pts"]
            _style_sheet(ws, len(hcols), pct_cols=one_dp,
                         widths={hcols.index("issuer_cluster") + 1: 26,
                                 hcols.index("fund_name") + 1: 32,
                                 hcols.index("dates_present") + 1: 9,
                                 hcols.index("mark_chg_pts") + 1: 11})
            for i, c in enumerate(hcols):
                if c.startswith("%Port "):
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=i + 1).number_format = "0.000"
            # Freeze BOTH header row and the two label columns — with 38 columns you lose track
            # of which issuer/fund a row belongs to as soon as you scroll right.
            ws.freeze_panes = "C2"


def build(threshold: int = 90) -> None:
    df = add_clusters(load_consolidated(), threshold=threshold)
    holdings, issues = match_issues(df)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT_DIR / "holdings_consolidated.csv", index=False, encoding="utf-8")
    holdings.to_csv(OUT_DIR / "holdings_matched.csv", index=False, encoding="utf-8")
    issues.to_csv(OUT_DIR / "issues.csv", index=False, encoding="utf-8")
    print(f"wrote {len(df):,} holdings ({df['issuer_cluster'].nunique():,} clusters), "
          f"{len(issues):,} issues -> {OUT_DIR}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--diagnose", action="store_true", help="report parse coverage + anchors")
    ap.add_argument("--cluster", action="store_true", help="report Phase-2 issuer clustering")
    ap.add_argument("--issues", action="store_true", help="report Phase-3 issue matching")
    ap.add_argument("--build", action="store_true", help="write consolidated + matched + issues CSVs")
    ap.add_argument("--workbook", action="store_true", help="write Phase-4 mark-comparison .xlsx")
    ap.add_argument("--threshold", type=int, default=92, help="fuzzy merge threshold (Phase 2)")
    ap.add_argument("--from-cache", action="store_true",
                    help="with --workbook: reuse the holdings_matched/issues CSVs written by "
                         "--build instead of re-clustering (~70min -> seconds). Only valid when "
                         "those CSVs are current; re-run --build after any parsing change.")
    args = ap.parse_args()
    if args.workbook:
        build_workbook(threshold=args.threshold, from_cache=args.from_cache)
    elif args.build:
        build(threshold=args.threshold)
    elif args.cluster:
        diagnose_clusters(threshold=args.threshold)
    elif args.issues:
        diagnose_issues(threshold=args.threshold)
    else:
        diagnose()
