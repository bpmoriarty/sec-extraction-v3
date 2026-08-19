"""
credit_migration.py — Credit migration & fund attribution (docs/CREDIT_MIGRATION_PLAN.md).

Answers a seven-question review list as THREE engines over one window:

  A. Credit migration  — assets and issuer counts crossing price thresholds  (items 2, 3, 4)
  B. Issuer importance — exposure-weighted price change, and breadth        (items 1, 5a)
  C. Fund attribution  — what the deteriorating slice cost each fund        (items 5b, 6)

Writes its OWN workbook (data/dataset/credit_migration.xlsx). Nothing mixes into
holdings_marks_comparison.xlsx. No API cost; every input is already on disk.

Run the reconciliation gate on its own (checkpoint 1 — do this before trusting anything else):
    uv run python src/analysis/credit_migration.py --gate
Full workbook:
    uv run python src/analysis/credit_migration.py --build

FOUR THINGS THAT MAKE THE NUMBERS DEFENSIBLE, each measured rather than assumed. The plan doc
carries the evidence; the short version:

1. A DEBT HOLDING MUST CARRY A PAR AMOUNT. Filers tag industry-level AGGREGATE rows on the same
   InvestmentIdentifierAxis, so summing holdings does not reconcile to the tagged portfolio
   (median 1.391 at 2024-12, p90 6.15). Pattern-matching those aggregate rows recovers almost
   nothing (51.5% -> 51.7% of assets). Requiring a par amount recovers nearly everything
   (-> 88.9%), because aggregates are fair-value-only sector totals while real debt holdings have
   par. It is also conceptually required: percent of par is undefined without par.

2. THE BUCKET SHARES SELF-VALIDATE. Every share is (priced debt in bucket) / (XBRL-tagged
   portfolio total), so across buckets they must sum to <= ~1.0. A fund that sums above
   RECON_MAX has double-counted rows and is flagged out automatically — the test IS the metric,
   so contamination cannot pass silently.

3. WEIGHTS ARE FIXED AT THE START OF THE WINDOW. Weighting by ENDING fair value is circular: a
   marked-down loan carries a smaller ending weight and so understates its own contribution.

4. A WEIGHTED MARK CHANGE IS NOT A RETURN. It omits interest income — which dominates BDC total
   return — plus realised gains, leverage and fees. It is labelled a valuation drag everywhere.
   Reported total_return is carried separately, and is fiscal YEAR-TO-DATE (period_months does
   NOT describe it: Golub reports 2.77/4.72/6.96 then resets to 1.65), so a window-aligned return
   is chained best-effort and left blank where the chain breaks.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "src" / "analysis"))

OUT_DIR = PROJECT_ROOT / "data" / "dataset"
EXTRACTED_DIR = PROJECT_ROOT / "data" / "extracted"
CONSOLIDATED = OUT_DIR / "holdings_consolidated.csv"
WORKBOOK = OUT_DIR / "credit_migration.xlsx"

# ── Window (docs/CREDIT_MIGRATION_PLAN.md §3 — Dec 2024 beats Dec 2023 on every axis) ────────
START = "2024-12-31"
END = "2026-06-30"

# ── Price buckets, in points of par ──────────────────────────────────────────────────────────
# >110 is held out rather than treated as premium: it is usually accrued interest landing in
# fair value, not a loan trading above par.
BUCKET_EDGES = [0, 85, 90, 95, 98, 100, 110, float("inf")]
BUCKET_LABELS = ["<85", "85-90", "90-95", "95-98", "98-100", "100-110", ">110 (suspect)"]

# Headline thresholds; the Coverage tab reports the whole sensitivity grid because 54.7% of
# marks cluster at 98-100, so the answer swings on whether "healthy" means 95 or 98.
HEALTHY_START = 95.0
IMPAIRED_END = 90.0
SENSITIVITY = [(95.0, 90.0), (95.0, 85.0), (98.0, 90.0), (90.0, 85.0)]

# A fund whose priced debt sums above this multiple of its tagged portfolio is double-counting.
RECON_MAX = 1.05
# Equity-like instruments carry no meaningful percent-of-par and are excluded from every price
# measure (they remain in the tagged portfolio denominator, which is correct).
EQUITY_TYPES = {"Equity", "Preferred", "Common Equity", "Warrant"}
# Concentration tab: how many issuers account for this share of a fund's drag.
CONCENTRATION_SHARE = 0.80
# A fund's valuation drag is only a statement about the FUND if it rests on most of the fund.
# MEASURED: AB Private Lending Fund's priced-debt-with-par is 2.3% of its portfolio, so its
# -2.84pt "drag" describes a sliver, not the book. Below this share the drag is flagged and
# excluded from the universe benchmark, rather than published as though it were representative.
MIN_DRAG_BASIS = 0.50


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Inputs
# ─────────────────────────────────────────────────────────────────────────────────────────────

def tagged_portfolio_fv() -> pd.DataFrame:
    """Portfolio fair value per (cik, reporting_date) from the XBRL-tagged balance sheet.

    The authoritative portfolio size and the denominator for every share on this workbook.
    Imported behaviour is deliberately duplicated from holdings_compare rather than imported,
    because importing that module pulls in rapidfuzz and the whole clustering path for a
    function that only reads JSON."""
    rows = []
    if not EXTRACTED_DIR.exists():
        return pd.DataFrame(columns=["cik", "reporting_date", "portfolio_fv"])
    for path in EXTRACTED_DIR.glob("*.json"):
        try:
            d = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        fact = (d.get("balance_sheet") or {}).get("investments_at_fair_value") or {}
        if fact.get("value"):
            rows.append({"cik": str(d.get("cik")).zfill(10),
                         "reporting_date": str(d.get("reporting_date"))[:10],
                         "portfolio_fv": float(fact["value"])})
    return pd.DataFrame(rows).drop_duplicates(["cik", "reporting_date"])


def reported_returns() -> pd.DataFrame:
    """Reported financial-highlights total_return per (cik, reporting_date), as a FRACTION,
    with the fiscal year-end needed to chain year-to-date figures into a window return.

    total_return here is fiscal YEAR-TO-DATE and `period_months` does not describe it (MEASURED:
    Golub reports 2.77 / 4.72 / 6.96 across consecutive quarters, then 9.36 on the 10-K, then
    RESETS to 1.65). Whoever consumes this must chain, never compound naively."""
    rows = []
    if not EXTRACTED_DIR.exists():
        return pd.DataFrame()
    for path in EXTRACTED_DIR.glob("*.json"):
        try:
            d = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        fh = d.get("financial_highlights") or {}
        if not isinstance(fh, dict):
            continue
        tr = fh.get("total_return") or {}
        if tr.get("value") is None:
            continue
        rows.append({"cik": str(d.get("cik")).zfill(10),
                     "reporting_date": str(d.get("reporting_date"))[:10],
                     "form_type": d.get("form_type"),
                     "ytd_return": float(tr["value"])})
    return pd.DataFrame(rows).drop_duplicates(["cik", "reporting_date"])


def load_panel() -> pd.DataFrame:
    """Priced debt holdings at the two window dates, with a par amount.

    The par requirement is the whole de-contamination story (see the module docstring). Returns
    holding-level rows; aggregation to (fund, issuer) happens in `pairs()` so the raw rows stay
    available for diagnostics."""
    if not CONSOLIDATED.exists():
        raise SystemExit(f"missing {CONSOLIDATED}; run holdings_compare.py --build first")
    cols = ["cik", "fund_name", "reporting_date", "issuer_cluster", "fair_value", "principal",
            "price", "price_basis", "parse_ok", "instrument_type", "seniority"]
    d = pd.read_csv(CONSOLIDATED, low_memory=False, usecols=cols)
    d["cik"] = d["cik"].astype(str).str.zfill(10)
    d = d[d["reporting_date"].isin([START, END])]
    keep = (d["price"].notna() & (d["parse_ok"] == True)          # noqa: E712 (pandas mask)
            & d["issuer_cluster"].notna()
            & ~d["instrument_type"].isin(EQUITY_TYPES)
            & d["principal"].notna())                             # <- the de-contamination rule
    d = d[keep].copy()
    d["mark"] = d["price"] * 100.0
    return d


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Reconciliation gate (checkpoint 1)
# ─────────────────────────────────────────────────────────────────────────────────────────────

def reconciliation(panel: pd.DataFrame, tag: pd.DataFrame) -> pd.DataFrame:
    """Per (cik, date): priced-debt sum vs tagged portfolio, and whether the fund is usable.

    Under 1.0 is fine and merely means incomplete price coverage, which is reported honestly.
    Above RECON_MAX means rows are double-counted and every dollar share for that fund would be
    overstated, so it is excluded from dollar measures and named on the Coverage tab."""
    g = (panel.groupby(["cik", "fund_name", "reporting_date"])["fair_value"]
         .sum().reset_index(name="priced_debt_fv"))
    j = g.merge(tag, on=["cik", "reporting_date"], how="left")
    j["coverage"] = j["priced_debt_fv"] / j["portfolio_fv"]
    j["reconciles"] = j["coverage"].le(RECON_MAX)
    j.loc[j["portfolio_fv"].isna(), "reconciles"] = False
    j["reason"] = ""
    j.loc[j["portfolio_fv"].isna(), "reason"] = "no tagged portfolio total"
    j.loc[j["coverage"] > RECON_MAX, "reason"] = "priced debt exceeds portfolio (double-counted)"
    return j


def usable_funds(recon: pd.DataFrame) -> tuple[set[str], set[str]]:
    """(usable_at_both, present_at_both) — the two flags every fund-level tab carries.

    Kept separate deliberately. A fund present at both dates but not reconciling is a DATA
    problem; a fund that reconciles but is missing an endpoint is a COVERAGE problem (it launched
    mid-window). Conflating them would let a 2025 launch look like a clean book."""
    ok = recon[recon["reconciles"]]
    at = lambda dt, frame: set(frame.loc[frame["reporting_date"] == dt, "cik"])  # noqa: E731
    usable = at(START, ok) & at(END, ok)
    present = at(START, recon) & at(END, recon)
    return usable, present


def run_gate() -> None:
    """Checkpoint 1. Print the gate and stop — nothing else is trustworthy until this passes."""
    panel, tag = load_panel(), tagged_portfolio_fv()
    recon = reconciliation(panel, tag)
    usable, present = usable_funds(recon)
    tot_end = tag.loc[tag["reporting_date"] == END, "portfolio_fv"].sum()
    share = (tag.loc[(tag["reporting_date"] == END) & (tag["cik"].isin(usable)),
                     "portfolio_fv"].sum() / tot_end) if tot_end else 0.0

    print("=" * 78)
    print(f"RECONCILIATION GATE  window {START} -> {END}")
    print("=" * 78)
    print(f"  priced-debt-with-par holding rows: {len(panel):,}")
    for dt in (START, END):
        s = recon[recon["reporting_date"] == dt]
        print(f"  {dt}: {len(s):3} funds  reconcile {s['reconciles'].sum():3} "
              f"({s['reconciles'].mean():.0%})  median coverage {s['coverage'].median():.1%}")
    print()
    print(f"  usable at BOTH dates:   {len(usable):3}")
    print(f"  present at BOTH dates:  {len(present):3}")
    print(f"  share of END BDC assets covered by usable funds: {share:.1%}")
    print()
    bad = recon[~recon["reconciles"]].sort_values("coverage", ascending=False)
    print(f"  EXCLUDED fund-dates ({len(bad)}) - published, not hidden:")
    for _, r in bad.head(12).iterrows():
        cov = f"{r['coverage']:.2f}x" if pd.notna(r["coverage"]) else "  n/a"
        print(f"    {r['reporting_date']}  {cov:>7}  {str(r['fund_name'])[:42]:42} {r['reason']}")
    if len(bad) > 12:
        print(f"    ... and {len(bad)-12} more")
    print()
    print("  VERDICT: expected ~43 usable funds / ~88.9% of assets (plan section 4.1).")
    print("  If materially lower, dollar answers are unsafe -> fall back to issuer COUNTS only.")


# ─────────────────────────────────────────────────────────────────────────────────────────────
# The (fund, issuer) panel — the object every engine works from
# ─────────────────────────────────────────────────────────────────────────────────────────────

def pairs(panel: pd.DataFrame) -> pd.DataFrame:
    """One row per (fund, issuer) with start and end marks, fair value and par.

    Marks are FAIR-VALUE-WEIGHTED within a (fund, issuer), not median: this feeds dollar
    attribution, so a fund's big tranche must dominate its small one. Grain is issuer rather
    than tranche because tranche keys do not survive across dates — MEASURED in Phase 7, 55% of
    (issuer, seniority, spread) tranches appear on only one as-of date.

    `status` distinguishes held_both / exited / entered. `exited` matters: a fund that SOLD its
    deteriorating credits shows no migration, which is indistinguishable from never having owned
    any unless exits are counted separately."""
    panel = panel.copy()
    panel["_w"] = panel["fair_value"]
    g = (panel.groupby(["cik", "fund_name", "issuer_cluster", "reporting_date"])
         .apply(lambda d: pd.Series({
             "mark": ((d["mark"] * d["_w"]).sum() / d["_w"].sum()) if d["_w"].sum() else float("nan"),
             "fv": d["fair_value"].sum(),
             "par": d["principal"].sum(),
             "lots": len(d),
             "seniority": "; ".join(sorted(set(d["seniority"].dropna()))) or None,
         }), include_groups=False)
         .reset_index())
    wide = g.pivot_table(index=["cik", "fund_name", "issuer_cluster"],
                         columns="reporting_date",
                         values=["mark", "fv", "par", "lots"], aggfunc="first")
    wide.columns = [f"{a}_{'start' if b == START else 'end'}" for a, b in wide.columns]
    wide = wide.reset_index()
    for c in ("mark_start", "mark_end", "fv_start", "fv_end", "par_start", "par_end"):
        if c not in wide.columns:
            wide[c] = float("nan")
    both = wide["mark_start"].notna() & wide["mark_end"].notna()
    wide["status"] = "entered"
    wide.loc[both, "status"] = "held_both"
    wide.loc[wide["mark_start"].notna() & wide["mark_end"].isna(), "status"] = "exited"
    wide["mark_chg"] = (wide["mark_end"] - wide["mark_start"]).round(2)
    wide["bucket_start"] = pd.cut(wide["mark_start"], BUCKET_EDGES, labels=BUCKET_LABELS)
    wide["bucket_end"] = pd.cut(wide["mark_end"], BUCKET_EDGES, labels=BUCKET_LABELS)
    return wide


def _migrated_mask(p: pd.DataFrame, healthy: float = HEALTHY_START,
                   impaired: float = IMPAIRED_END) -> pd.Series:
    """Held at both ends, healthy at the start, impaired at the end — his item 2/4 definition."""
    return ((p["status"] == "held_both") & (p["mark_start"] >= healthy)
            & (p["mark_end"] < impaired))


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Engine A — credit migration (his items 2, 3, 4)
# ─────────────────────────────────────────────────────────────────────────────────────────────

def migration_matrix(p: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Start bucket x end bucket, in start-weighted dollars and in issuer counts.

    Answers items 2, 3 and 4 simultaneously and retires the threshold argument: item 3 is a
    column total, item 2 is the healthy-row/impaired-column block, item 4 is the same block in
    the count matrix. The diagonal (what stayed put) is the baseline all of it must be read
    against, and the Exited column keeps sold positions visible."""
    q = p[p["status"].isin(["held_both", "exited"])].copy()
    q["end_lab"] = q["bucket_end"].astype("object").where(q["status"] == "held_both", "EXITED")
    dollars = (q.pivot_table(index="bucket_start", columns="end_lab", values="fv_start",
                             aggfunc="sum", observed=False).fillna(0.0) / 1e6).round(1)
    counts = q.pivot_table(index="bucket_start", columns="end_lab", values="fv_start",
                           aggfunc="size", observed=False).fillna(0).astype(int)
    order = [c for c in BUCKET_LABELS if c in dollars.columns] + \
            (["EXITED"] if "EXITED" in dollars.columns else [])
    return dollars[order], counts[order]


def fund_migration(p: pd.DataFrame, tag: pd.DataFrame, recon: pd.DataFrame,
                   usable: set[str], present: set[str]) -> pd.DataFrame:
    """Per fund: bucket shares at both ends, what migrated, what exited, and the two flags.

    Every share divides by the fund's XBRL-TAGGED portfolio total, never by the sum of its own
    holdings — so the shares are comparable across funds and cannot be inflated by a fund whose
    schedule double-counts. Shares therefore do NOT sum to 100%: the remainder is equity, cash,
    unpriced debt and holdings without a par amount, and `priced_debt_cov` states it."""
    port = tag.pivot_table(index="cik", columns="reporting_date", values="portfolio_fv",
                           aggfunc="first")
    rows = []
    for (cik, name), d in p.groupby(["cik", "fund_name"]):
        ps = port.at[cik, START] if (cik in port.index and START in port.columns) else float("nan")
        pe = port.at[cik, END] if (cik in port.index and END in port.columns) else float("nan")
        rec = {"cik": cik, "fund_name": name,
               "usable?": "yes" if cik in usable else "no",
               "constant_sample?": "yes" if cik in present else "no",
               "portfolio_start_mm": round(ps / 1e6, 0) if pd.notna(ps) else None,
               "portfolio_end_mm": round(pe / 1e6, 0) if pd.notna(pe) else None}
        held = d[d["status"] == "held_both"]
        for lab in BUCKET_LABELS:
            s = d.loc[d["bucket_start"] == lab, "fv_start"].sum()
            e = d.loc[d["bucket_end"] == lab, "fv_end"].sum()
            rec[f"start % {lab}"] = round(100 * s / ps, 2) if pd.notna(ps) and ps else None
            rec[f"end % {lab}"] = round(100 * e / pe, 2) if pd.notna(pe) and pe else None
        mig = d[_migrated_mask(d)]
        ex = d[d["status"] == "exited"]
        rec["migrated_mm"] = round(mig["fv_start"].sum() / 1e6, 1)
        rec["migrated_pct_of_portfolio"] = (round(100 * mig["fv_start"].sum() / ps, 2)
                                           if pd.notna(ps) and ps else None)
        rec["migrated_issuers"] = int(len(mig))
        rec["exited_mm"] = round(ex["fv_start"].sum() / 1e6, 1)
        rec["exited_issuers"] = int(len(ex))
        rec["issuers_held_both"] = int(len(held))
        cs = recon[(recon["cik"] == cik) & (recon["reporting_date"] == START)]["coverage"]
        ce = recon[(recon["cik"] == cik) & (recon["reporting_date"] == END)]["coverage"]
        rec["priced_debt_cov_start"] = round(float(cs.iloc[0]), 3) if len(cs) else None
        rec["priced_debt_cov_end"] = round(float(ce.iloc[0]), 3) if len(ce) else None
        rows.append(rec)
    out = pd.DataFrame(rows)
    # usable funds first, then by how much migrated - so the flagged block cannot be mistaken
    # for a clean ranking
    return out.sort_values(["usable?", "migrated_pct_of_portfolio"],
                           ascending=[True, False], na_position="last")


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Engine B — issuer importance (his items 1, 5a)
# ─────────────────────────────────────────────────────────────────────────────────────────────

def issuer_impact(p: pd.DataFrame, usable: set[str]) -> tuple[pd.DataFrame, dict]:
    """Per issuer: start-weighted price change, BDC-visible debt, and share of the universe.

    Computed over the RECONCILING funds only, and that is deliberate even though the fund-level
    tabs show all funds. A fund-level row can carry a flag; a cross-fund DOLLAR SUM cannot -
    including a double-counting fund would silently overstate an issuer's exposure and there is
    no per-row warning that would fix it.

    Weights are start-date fair value (headline) with par published beside it, because weighting
    by ENDING value lets a marked-down loan shrink its own contribution."""
    q = p[(p["cik"].isin(usable)) & (p["status"] == "held_both")].copy()
    tot_start = q["fv_start"].sum()
    rows = []
    for iss, d in q.groupby("issuer_cluster"):
        w = d["fv_start"]
        if w.sum() <= 0:
            continue
        wtd = (d["mark_chg"] * w).sum() / w.sum()
        wtd_par = ((d["mark_chg"] * d["par_start"]).sum() / d["par_start"].sum()
                   if d["par_start"].sum() > 0 else float("nan"))
        share = w.sum() / tot_start if tot_start else float("nan")
        rows.append({
            "issuer_cluster": iss,
            "wtd_mark_chg_pts": round(wtd, 2),
            "wtd_mark_chg_pts_par_weighted": round(wtd_par, 2) if pd.notna(wtd_par) else None,
            "bdc_debt_start_mm": round(w.sum() / 1e6, 1),
            "bdc_debt_end_mm": round(d["fv_end"].sum() / 1e6, 1),
            "par_start_mm": round(d["par_start"].sum() / 1e6, 1),
            "universe_share_start_pct": round(100 * share, 4),
            # item 1's "breadth of importance": how many pts of the whole BDC book this one
            # credit moved. Big fall on a small loan scores low; this is the point.
            "contribution_to_universe_pts": round(share * wtd, 4),
            "n_holders": int(d["cik"].nunique()),
            "mark_start": round((d["mark_start"] * w).sum() / w.sum(), 1),
            "mark_end": round((d["mark_end"] * w).sum() / w.sum(), 1),
        })
    out = pd.DataFrame(rows).sort_values("contribution_to_universe_pts")

    # item 5a: the criteria-meeting cohort against the whole universe, at both ends
    mig = q[_migrated_mask(q)]
    summary = {
        "universe priced debt at start ($mm)": round(tot_start / 1e6, 0),
        "universe priced debt at end ($mm)": round(q["fv_end"].sum() / 1e6, 0),
        f"cohort: held both, >={HEALTHY_START:.0f} at start and <{IMPAIRED_END:.0f} at end": "",
        "  issuers": int(mig["issuer_cluster"].nunique()),
        "  (fund, issuer) positions": int(len(mig)),
        "  debt at START ($mm)": round(mig["fv_start"].sum() / 1e6, 0),
        "  debt at END ($mm)": round(mig["fv_end"].sum() / 1e6, 0),
        "  share of universe priced debt at START":
            f"{100*mig['fv_start'].sum()/tot_start:.2f}%" if tot_start else "n/a",
        "  its weighted mark change (pts)":
            round((mig["mark_chg"] * mig["fv_start"]).sum() / mig["fv_start"].sum(), 2)
            if mig["fv_start"].sum() else None,
    }
    return out, summary


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Engine C — fund attribution (his items 5b, 6)
# ─────────────────────────────────────────────────────────────────────────────────────────────

def window_total_return(rr_fund: pd.DataFrame) -> tuple[float | None, str]:
    """Chain fiscal YEAR-TO-DATE total returns into a window return, or say why it cannot.

    total_return resets at each fiscal year end, so it must be CHAINED, never compounded.
    MEASURED: Golub reports 2.77 / 4.72 / 6.96 across consecutive quarters, then 9.36 on the
    10-K, then resets to 1.65 - all tagged period_months=3. Fiscal year ends are identified as
    the reporting dates of 10-K filings.

    Returns (None, reason) rather than an estimate whenever the chain is incomplete."""
    ytd = dict(zip(rr_fund["reporting_date"], rr_fund["ytd_return"]))
    fyes = sorted(rr_fund.loc[rr_fund["form_type"] == "10-K", "reporting_date"])
    if START not in ytd:
        return None, "no YTD at window start"
    if END not in ytd:
        return None, "no YTD at window end"
    # If the window opens exactly on a fiscal year end, the new year starts from zero.
    base = 0.0 if START in fyes else ytd[START]
    mids = [f for f in fyes if START < f < END]
    factor, prev = 1.0, None
    for f in mids:
        if f not in ytd:
            return None, f"no YTD at fiscal year end {f}"
        factor *= (1 + ytd[f]) / (1 + base) if prev is None else (1 + ytd[f])
        prev = f
    if prev is None:                      # no fiscal boundary inside the window
        factor *= (1 + ytd[END]) / (1 + base)
    else:                                 # last stub runs from the final FYE, so YTD as-is
        factor *= (1 + ytd[END])
    return factor - 1.0, ""


def fund_attribution(p: pd.DataFrame, tag: pd.DataFrame, usable: set[str],
                     present: set[str], rr: pd.DataFrame) -> pd.DataFrame:
    """Per fund: the valuation drag on its priced debt, how much came from migrated credits,
    the asset-weighted universe benchmark (item 5b), and the reported return (item 6).

    'Drag' is NOT a return and is never labelled one. It measures only revaluation of priced
    debt; it excludes interest income (which dominates BDC total return), realised gains,
    leverage and fees. The reported-return columns are the actual return, carried separately."""
    port = tag.pivot_table(index="cik", columns="reporting_date", values="portfolio_fv",
                           aggfunc="first")
    rows = []
    for (cik, name), d in p.groupby(["cik", "fund_name"]):
        held = d[d["status"] == "held_both"]
        w = held["fv_start"]
        if w.sum() <= 0:
            continue
        drag = (held["mark_chg"] * w).sum() / w.sum()
        mig = held[_migrated_mask(held)]
        mig_contrib = (mig["mark_chg"] * mig["fv_start"]).sum() / w.sum() if w.sum() else None
        rows.append({
            "cik": cik, "fund_name": name,
            "usable?": "yes" if cik in usable else "no",
            "constant_sample?": "yes" if cik in present else "no",
            "priced_debt_start_mm": round(w.sum() / 1e6, 1),
            "valuation_drag_pts": round(drag, 2),
            "drag_from_migrated_pts": round(mig_contrib, 2) if mig_contrib is not None else None,
            "share_of_drag_from_migrated": (round(mig_contrib / drag, 3)
                                            if drag and mig_contrib is not None
                                            and abs(drag) > 1e-9 else None),
            "migrated_issuers": int(len(mig)),
            "_w": (port.at[cik, START] if (cik in port.index and START in port.columns)
                   else float("nan")),
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out

    # How much of the fund the drag actually describes. A drag measured on 2% of a portfolio is
    # not a fact about the fund, so it is flagged and kept out of the benchmark.
    out["drag_basis_pct"] = (100 * out["priced_debt_start_mm"] * 1e6 / out["_w"]).round(1)
    out["flags"] = ""
    low = out["drag_basis_pct"] < MIN_DRAG_BASIS * 100
    out.loc[low, "flags"] = f"low_basis(<{MIN_DRAG_BASIS:.0%} of portfolio priced)"
    out.loc[out["usable?"] == "no", "flags"] = (
        out["flags"] + "; not_reconciling").str.strip("; ")
    out.loc[out["constant_sample?"] == "no", "flags"] = (
        out["flags"] + "; missing_an_endpoint").str.strip("; ")

    # item 5b: the universe benchmark, asset-weighted over funds that BOTH reconcile and have a
    # representative basis. A double-counting fund would distort the benchmark and no per-row
    # flag can fix that; a sliver-basis fund would contribute a drag that is not its own.
    u = out[(out["usable?"] == "yes") & out["_w"].notna() & ~low]
    bench = ((u["valuation_drag_pts"] * u["_w"]).sum() / u["_w"].sum()) if len(u) else float("nan")
    out["universe_drag_pts_asset_wtd"] = round(bench, 2)
    out["vs_universe_pts"] = (out["valuation_drag_pts"] - bench).round(2)

    # item 6: the actual reported return, chained across fiscal years, flagged when impossible
    wr, why, ytd_end = [], [], []
    for cik in out["cik"]:
        sub = rr[rr["cik"] == cik]
        if sub.empty:
            wr.append(None); why.append("no reported total_return"); ytd_end.append(None); continue
        v, reason = window_total_return(sub)
        wr.append(round(100 * v, 2) if v is not None else None)
        why.append(reason)
        e = sub.loc[sub["reporting_date"] == END, "ytd_return"]
        ytd_end.append(round(100 * float(e.iloc[0]), 2) if len(e) else None)
    out["reported_window_total_return_pct"] = wr
    out["reported_return_gap_reason"] = why
    out["reported_fiscal_ytd_at_end_pct"] = ytd_end
    return out.drop(columns="_w").sort_values(["usable?", "valuation_drag_pts"],
                                              ascending=[True, True])


def concentration(p: pd.DataFrame, usable: set[str]) -> pd.DataFrame:
    """How concentrated each fund's drag is: the fewest issuers accounting for
    CONCENTRATION_SHARE of it. One bad name and thirty bad names produce the same aggregate."""
    rows = []
    for (cik, name), d in p[p["cik"].isin(usable)].groupby(["cik", "fund_name"]):
        held = d[(d["status"] == "held_both") & (d["mark_chg"] < 0)]
        if held.empty:
            continue
        contrib = (held["mark_chg"] * held["fv_start"]).abs().sort_values(ascending=False)
        tot = contrib.sum()
        if tot <= 0:
            continue
        csum = contrib.cumsum() / tot
        n = int((csum < CONCENTRATION_SHARE).sum() + 1)
        top = held.loc[contrib.index[:3], "issuer_cluster"].tolist()
        rows.append({"fund_name": name, "declining_issuers": int(len(held)),
                     f"issuers for {CONCENTRATION_SHARE:.0%} of drag": n,
                     "concentration_ratio": round(n / len(held), 3),
                     "top_3_contributors": "; ".join(top)})
    return (pd.DataFrame(rows)
            .sort_values(f"issuers for {CONCENTRATION_SHARE:.0%} of drag"))


def manager_rollup(attr: pd.DataFrame, fmig: pd.DataFrame) -> pd.DataFrame:
    """Every cut at parent-manager grain. Several managers run five or more vehicles, and
    family-level concentration is usually the finding a per-fund table buries."""
    try:
        from managers import manager_of
    except Exception:
        return pd.DataFrame()
    a = attr[attr["usable?"] == "yes"].copy()
    a["manager"] = a["cik"].map(manager_of)
    m = fmig[fmig["usable?"] == "yes"][["cik", "migrated_mm", "migrated_issuers", "exited_mm"]]
    a = a.merge(m, on="cik", how="left")
    g = (a.groupby("manager")
         .apply(lambda d: pd.Series({
             "funds": d["cik"].nunique(),
             "priced_debt_start_mm": round(d["priced_debt_start_mm"].sum(), 1),
             "valuation_drag_pts_wtd": round(
                 (d["valuation_drag_pts"] * d["priced_debt_start_mm"]).sum()
                 / d["priced_debt_start_mm"].sum(), 2)
             if d["priced_debt_start_mm"].sum() else None,
             "migrated_mm": round(d["migrated_mm"].sum(), 1),
             "migrated_issuers": int(d["migrated_issuers_y"].sum())
             if "migrated_issuers_y" in d else int(d["migrated_issuers"].sum()),
         }), include_groups=False)
         .reset_index())
    return g.sort_values("valuation_drag_pts_wtd")


def sensitivity(p: pd.DataFrame, usable: set[str]) -> pd.DataFrame:
    """The headline thresholds are one arbitrary pair out of several defensible ones, and 54.7%
    of marks sit at 98-100, so the grid ships beside the answer rather than behind it."""
    q = p[(p["cik"].isin(usable)) & (p["status"] == "held_both")]
    tot = q["fv_start"].sum()
    rows = []
    for h, i in SENSITIVITY:
        m = q[_migrated_mask(q, h, i)]
        rows.append({"healthy at start (>=)": h, "impaired at end (<)": i,
                     "issuers": int(m["issuer_cluster"].nunique()),
                     "(fund, issuer) positions": int(len(m)),
                     "debt at start ($mm)": round(m["fv_start"].sum() / 1e6, 0),
                     "share of universe priced debt":
                         round(100 * m["fv_start"].sum() / tot, 2) if tot else None})
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Definitions tab
# ─────────────────────────────────────────────────────────────────────────────────────────────
# One entry per column: (what it means, how it is computed, what to watch out for). Written for
# a colleague who did not build this and will read a number without reading the Overview.
#
# VALIDATED at build time against the columns actually written (see `definitions_table`): a new
# column with no definition, or a definition for a column that no longer exists, is reported
# rather than silently shipped. Documentation that drifts from the data is worse than none.

_DEFS: dict[str, tuple[str, str, str]] = {
    # ---- identifiers, shared across tabs ----
    "cik": ("The fund's unique SEC identifier (Central Index Key), zero-padded to 10 digits.",
            "From the SEC filing.",
            "Join on this, never on fund_name - names are spelled inconsistently across filings."),
    "fund_name": ("The fund's name as reported in its filing.",
                  "From the SEC filing.",
                  "Spelling varies between filings for the same fund; cik is the stable key."),
    "manager": ("Parent asset manager the fund rolls up to (e.g. all Blue Owl vehicles -> "
                "'Blue Owl').",
                "Curated by CIK in src/analysis/managers.py.",
                "'UNMAPPED' means the CIK is not in the curated map - never silently bucketed."),
    "issuer_cluster": ("The borrower (portfolio company).",
                       "Fuzzy-clustered from the issuer names filers tag in their schedule of "
                       "investments.",
                       "One row per BORROWER, not per loan. A borrower's several tranches are "
                       "combined into one weighted mark. A borrower whose name clusters two "
                       "ways will appear as two rows, understating its true size."),
    # ---- flags ----
    "usable?": ("Whether this fund's DOLLAR figures can be trusted.",
                "'yes' = the fund's priced debt reconciles to its own XBRL-tagged portfolio "
                "total at BOTH window dates (ratio <= 1.05).",
                "'no' means the filing double-counts holdings rows, so every $ and % for this "
                "fund is OVERSTATED. Such funds are kept in the file on purpose, in their own "
                "block, and named on Coverage. Do not rank them against the others."),
    "constant_sample?": ("Whether the fund has data at both ends of the window.",
                         "'yes' = priced holdings present at both the start and end date.",
                         "A 'no' fund is usually one that LAUNCHED MID-WINDOW (seven first "
                         "appear at 2025-06-30). It will show little or no migration because it "
                         "has no history - not because it underwrote well. This is the single "
                         "easiest way to misread the file."),
    "flags": ("Machine-readable warnings for the row.",
              "Combination of: low_basis, not_reconciling, missing_an_endpoint.",
              "'low_basis' is the important one - see drag_basis_pct."),
    # ---- portfolio sizes ----
    "portfolio_start_mm": ("The fund's whole investment portfolio at the START date, $ millions.",
                           "XBRL-tagged 'investments at fair value' from the fund's own balance "
                           "sheet.",
                           "This is the denominator for every % on the tab. It is the filer's "
                           "own figure, not a sum of the holdings we parsed."),
    "portfolio_end_mm": ("The fund's whole investment portfolio at the END date, $ millions.",
                         "As portfolio_start_mm, at the end date.",
                         "A fund can grow a lot over the window; a position's % can fall purely "
                         "because the portfolio grew."),
    "priced_debt_cov_start": ("What fraction of the fund's portfolio this analysis actually "
                              "measures, at the START date.",
                              "(priced debt with a par amount) / (tagged portfolio total).",
                              "0.9 means 90% of the book is covered; the rest is equity, cash, "
                              "unpriced debt and holdings with no par. Above 1.05 means "
                              "double-counting, which is what sets usable? = no."),
    "priced_debt_cov_end": ("Same coverage fraction at the END date.",
                            "As priced_debt_cov_start, at the end date.",
                            "Coverage can differ between dates for the same fund."),
    # ---- FundMigration measures ----
    "migrated_mm": ("Start-date value of credits that went from healthy to impaired, $ millions.",
                    f"Sum of START fair value for issuers held at BOTH dates, marked "
                    f">= {HEALTHY_START:.0f} at the start and < {IMPAIRED_END:.0f} at the end.",
                    "Valued at the START date deliberately, so a markdown does not shrink its "
                    "own measured size."),
    "migrated_pct_of_portfolio": ("The same figure as a share of the whole fund.",
                                  "migrated_mm / portfolio_start_mm x 100.",
                                  "The headline 'how much of this fund went bad' number. Read "
                                  "usable? and constant_sample? before comparing funds."),
    "migrated_issuers": ("How many borrowers made that healthy-to-impaired move.",
                         "Count of issuers meeting the migration test.",
                         "A count, not a size - one big name and ten small ones both count as "
                         "their number of borrowers."),
    "exited_mm": ("Start-date value of credits the fund held at the start and no longer held at "
                  "the end, $ millions.",
                  "Sum of START fair value where status = exited.",
                  "Sold, repaid at maturity, or restructured - THE FILINGS DO NOT SAY WHICH. A "
                  "fund that sold its deteriorating loans shows LOW migration and HIGH exits, "
                  "which can look like good underwriting but is not the same thing."),
    "exited_issuers": ("How many borrowers left the portfolio during the window.",
                       "Count of issuers with status = exited.",
                       "Turnover is large across this universe - 4,443 of 15,327 positions "
                       "exited - so a low migration figure often means churn, not resilience."),
    "issuers_held_both": ("How many borrowers the fund held, with a usable mark, at BOTH dates.",
                          "Count of issuers with status = held_both.",
                          "The population every migration measure is drawn from. A small number "
                          "makes the fund's percentages noisy."),
    # ---- IssuerImpact ----
    "wtd_mark_chg_pts": ("How far this borrower's price moved over the window, in points of par "
                         "(100 = par).",
                         "Change in mark from start to end, weighted across holders by each "
                         "holder's START fair value.",
                         "Negative = marked down. Weighted at the start on purpose: weighting "
                         "by ending value would let a marked-down loan shrink its own weight."),
    "wtd_mark_chg_pts_par_weighted": ("The same price move, weighted by par instead of by fair "
                                      "value.",
                                      "As wtd_mark_chg_pts, weighted by START par.",
                                      "A robustness check. Par does not move with the mark, so "
                                      "if this differs materially from the fair-value-weighted "
                                      "figure, the weighting is being influenced by the "
                                      "revaluation itself."),
    "bdc_debt_start_mm": ("How much of this borrower's debt the BDCs held at the START, "
                          "$ millions of fair value.",
                          "Sum of START fair value across the reconciling BDCs.",
                          "BDC-VISIBLE ONLY. CLOs, insurance accounts and other private funds "
                          "hold the rest of these loans and are invisible here, so this is not "
                          "the loan's market size."),
    "bdc_debt_end_mm": ("The same exposure at the END, $ millions of fair value.",
                        "Sum of END fair value across the reconciling BDCs.",
                        "A fall here mixes THREE things: markdowns, repayments, and holders "
                        "dropping out of the sample. It is not a measure of price alone."),
    "par_start_mm": ("Face (par) amount of the borrower's debt held at the START, $ millions.",
                     "Sum of START principal across the reconciling BDCs.",
                     "Par does not move when a loan is marked down, so it is the more stable "
                     "measure of how big a position really is."),
    "universe_share_start_pct": ("How much of all BDC priced debt this one borrower represents, "
                                 "at the START.",
                                 "bdc_debt_start_mm / (all reconciling BDCs' priced debt) x 100.",
                                 "Denominator is the reconciling funds only, so it is a share of "
                                 "the measured universe (88.9% of BDC assets), not of everything."),
    "contribution_to_universe_pts": ("THE IMPORTANCE RANKING: how many points this single "
                                     "borrower moved the entire BDC book.",
                                     "universe_share_start_pct/100 x wtd_mark_chg_pts.",
                                     "This is what makes a mild fall on a huge position outrank "
                                     "a severe fall on a tiny one - by design. Sort ascending "
                                     "for the most damaging credits."),
    "n_holders": ("How many BDCs held this borrower at both dates.",
                  "Distinct CIK count.",
                  "1 means NO cross-holder corroboration - the mark is one manager's opinion "
                  "with nothing to check it against. Treat single-holder rows cautiously."),
    "mark_start": ("The borrower's average price at the START, points of par.",
                   "Fair-value-weighted average across holders.",
                   "Above 100 usually means accrued interest landed in fair value, not a "
                   "genuine premium."),
    "mark_end": ("The borrower's average price at the END, points of par.",
                 "Fair-value-weighted average across holders.",
                 "Compare with mark_start; the difference is wtd_mark_chg_pts."),
    # ---- FundAttribution ----
    "priced_debt_start_mm": ("Start-date value of the fund's debt that this analysis measures, "
                             "$ millions.",
                             "Sum of START fair value of priced debt with a par amount.",
                             "Compare with portfolio_start_mm: the gap is what we cannot "
                             "measure. See drag_basis_pct."),
    "valuation_drag_pts": ("How far the fund's measured credits were marked down over the "
                           "window, in points of par.",
                           "Fair-value-weighted average mark change across credits held at both "
                           "dates, weighted at the START.",
                           "*** THIS IS NOT A RETURN. *** It captures revaluation only. It "
                           "excludes interest income - which dominates BDC total return - plus "
                           "realised gains, leverage and fees. Funds here carry -2 to -4 points "
                           "of drag and still reported returns above +10%."),
    "drag_from_migrated_pts": ("How much of that drag came from the healthy-to-impaired credits.",
                               "Sum(mark change x START fair value) for migrated issuers, "
                               "divided by the START fair value of ALL credits held at both "
                               "dates.",
                               "Same units as valuation_drag_pts, so the two are directly "
                               "comparable."),
    "share_of_drag_from_migrated": ("What proportion of the fund's drag the migrated credits "
                                    "explain.",
                                    "drag_from_migrated_pts / valuation_drag_pts.",
                                    "CAN EXCEED 1.0. That happens when the fund's other credits "
                                    "moved UP and offset part of the decline - it is meaningful, "
                                    "not an error."),
    "drag_basis_pct": ("How much of the fund the drag figure actually describes.",
                       "priced_debt_start_mm / portfolio_start_mm x 100.",
                       f"Below {MIN_DRAG_BASIS:.0%} the drag is a statement about a SLICE, not "
                       f"the fund, and the row is flagged low_basis and left out of the "
                       f"benchmark. One fund's basis is 2.3% - its drag is arithmetically "
                       f"correct and analytically meaningless."),
    "universe_drag_pts_asset_wtd": ("The benchmark: the average valuation drag across the BDC "
                                    "universe.",
                                    "Asset-weighted mean valuation_drag_pts over funds that both "
                                    "reconcile and have an adequate basis, weighted by START "
                                    "portfolio size.",
                                    "The SAME value on every row - it is a single universe "
                                    "figure, not a per-fund one."),
    "vs_universe_pts": ("How this fund compares with that benchmark.",
                        "valuation_drag_pts - universe_drag_pts_asset_wtd.",
                        "Negative = marked down MORE than the universe. This measures relative "
                        "valuation change, not skill - a fund can be worse here because it held "
                        "riskier credits or because it marks more conservatively. Cross-check "
                        "marking_bias.xlsx to tell those apart."),
    "reported_window_total_return_pct": ("The fund's ACTUAL total return over the window, "
                                         "percent.",
                                         "Chained from the fund's reported fiscal YEAR-TO-DATE "
                                         "total return, across fiscal year ends (identified by "
                                         "10-K dates).",
                                         "THIS is a return; valuation_drag_pts is not. Blank "
                                         "where the year-to-date chain is incomplete - the "
                                         "reason is in the next column. Never estimated."),
    "reported_return_gap_reason": ("Why the window return above is blank.",
                                   "Set by the chaining routine.",
                                   "Blank here means the return was computed successfully."),
    "reported_fiscal_ytd_at_end_pct": ("The fund's reported fiscal year-to-date total return as "
                                       "of the END date, percent.",
                                       "Read straight from the filing's financial highlights.",
                                       "NOT WINDOW-ALIGNED and NOT COMPARABLE ACROSS FUNDS: each "
                                       "fund's fiscal year starts in a different month, so this "
                                       "covers a different number of months for each one."),
    # ---- Concentration ----
    "declining_issuers": ("How many of the fund's borrowers were marked DOWN over the window.",
                          "Count of held-both issuers with a negative mark change.",
                          "Ignores size - a 1-point and a 40-point fall both count once."),
    "issuers for 80% of drag": ("How few borrowers explain most of the fund's decline.",
                                "Borrowers ranked by |mark change x START fair value|; the count "
                                "needed to reach 80% of the total.",
                                "A low number means the damage is one or two names; a high "
                                "number means it is broad. Both can produce the same "
                                "valuation_drag_pts, which is why this tab exists."),
    "concentration_ratio": ("The same idea as a fraction.",
                            "'issuers for 80% of drag' / declining_issuers.",
                            "Low = concentrated. Compare funds of similar size; a fund with "
                            "only two declining credits will always look concentrated."),
    "top_3_contributors": ("The three borrowers doing the most damage to this fund.",
                           "Largest three by |mark change x START fair value|.",
                           "Contribution combines size and price move, so a large stable-ish "
                           "position can outrank a small collapsing one."),
    # ---- ManagerRollup ----
    "funds": ("How many of this manager's BDCs are included.",
              "Distinct CIK count among reconciling funds.",
              "Only reconciling funds are rolled up, so a manager's total may exclude some of "
              "its vehicles."),
    "valuation_drag_pts_wtd": ("The manager's overall valuation drag, in points of par.",
                               "Mean valuation_drag_pts across its funds, weighted by each "
                               "fund's priced debt.",
                               "Same caveat as valuation_drag_pts: NOT a return."),
}

# Extra guidance for tabs that are not simple column tables.
_TAB_NOTES: dict[str, list[tuple[str, str, str, str]]] = {
    "MigrationMatrix": [
        ("rows (start bucket)", "The price band a credit sat in at the START of the window.",
         "Bands are points of par: <85, 85-90, 90-95, 95-98, 98-100, 100-110, >110.",
         "'>110 (suspect)' is held separately because a mark that high usually means accrued "
         "interest landed in fair value, not a loan trading above par."),
        ("columns (end bucket)", "The band the same credit sat in at the END.",
         "Same bands, plus EXITED.",
         "The DIAGONAL is what did not move - read every other cell against it."),
        ("EXITED column", "Positions held at the start that were gone by the end.",
         "Sold, repaid or restructured; the filings do not distinguish.",
         "This is the biggest block in the matrix ($38.3bn left the 98-100 band alone), so "
         "turnover - not migration - is the dominant story in this data."),
        ("upper block: DOLLARS", "Start-date fair value in each cell, $ millions.",
         "Sum of START fair value.",
         "Valued at the start so a markdown does not shrink its own measured size."),
        ("lower block: COUNTS", "Number of (fund, borrower) positions in each cell.",
         "Row count.",
         "Use with the dollar block: many small positions and one huge one look identical in "
         "counts alone."),
    ],
    "Coverage": [
        ("reconciliation gate", "Which funds' dollar figures are trustworthy, and which are not.",
         "Priced debt summed against each fund's XBRL-tagged portfolio total.",
         "Excluded fund-dates are listed BY NAME with their ratio - nothing is hidden."),
        ("threshold sensitivity", "How much the answer depends on where 'healthy' and 'impaired' "
         "are drawn.",
         "The migration test re-run at 95/90, 95/85, 98/90 and 90/85.",
         "54.7% of all marks sit between 98 and 100, so the result is genuinely sensitive to "
         "whether healthy means 95 or 98. Check this before quoting a single figure."),
    ],
}


def definitions_table(sheet_cols: dict[str, list[str]]) -> tuple[pd.DataFrame, list[str]]:
    """Build the Definitions rows, organised by tab, and report any drift.

    Returns (frame, problems). `problems` lists columns written to the workbook with no
    definition, and definitions for columns no longer written - so documentation cannot quietly
    fall out of step with the data."""
    rows, problems = [], []
    for tab, cols in sheet_cols.items():
        for label, meaning, how, watch in _TAB_NOTES.get(tab, []):
            rows.append({"Tab": tab, "Column / element": label, "What it means": meaning,
                         "How it is computed": how, "Watch out for": watch})
        for c in cols:
            key = c
            if c.startswith(("start % ", "end % ")):       # generated bucket columns
                when, band = ("START", c[8:]) if c.startswith("start % ") else ("END", c[6:])
                rows.append({
                    "Tab": tab, "Column / element": c,
                    "What it means": f"Share of the fund's WHOLE portfolio held in credits "
                                     f"marked {band} points of par, at the {when} date.",
                    "How it is computed": f"Sum of {when.lower()}-date fair value of priced debt "
                                          f"in the {band} band, divided by the fund's tagged "
                                          f"portfolio total at that date, x 100.",
                    "Watch out for": "These bands do NOT sum to 100%. The remainder is equity, "
                                     "cash, unpriced debt and holdings with no par amount - see "
                                     "priced_debt_cov. Bands are shares of the fund, not of its "
                                     "debt.",
                })
                continue
            if key not in _DEFS:
                problems.append(f"NO DEFINITION: {tab}.{c}")
                continue
            meaning, how, watch = _DEFS[key]
            rows.append({"Tab": tab, "Column / element": c, "What it means": meaning,
                         "How it is computed": how, "Watch out for": watch})
    written = {c for cols in sheet_cols.values() for c in cols}
    for k in _DEFS:
        if k not in written:
            problems.append(f"ORPHAN DEFINITION (column not written): {k}")
    return pd.DataFrame(rows), problems


# ─────────────────────────────────────────────────────────────────────────────────────────────
# Workbook
# ─────────────────────────────────────────────────────────────────────────────────────────────

CAVEATS = [
    f"CREDIT MIGRATION & FUND ATTRIBUTION   window {START} -> {END}",
    "",
    "Marks are POINTS OF PAR (100 = par). Plan: docs/CREDIT_MIGRATION_PLAN.md.",
    "",
    "READ THIS FIRST - four things that decide what these numbers can claim:",
    "",
    "1. A DEBT HOLDING MUST CARRY A PAR AMOUNT to be counted. Filers tag industry-level",
    "   AGGREGATE rows on the same XBRL axis, so summing holdings does not reconcile to the",
    "   tagged portfolio (median 1.39x at the start date). Requiring par fixes it, because",
    "   aggregates are fair-value-only sector totals - and percent of par is undefined without",
    "   par. Measured: this lifts usable asset coverage from 51.5% to 88.9%.",
    "",
    "2. FUNDS THAT STILL DOUBLE-COUNT ARE FLAGGED, NOT REPAIRED. 'usable? = no' means the",
    "   fund's priced debt exceeds its own tagged portfolio, so its DOLLAR shares would be",
    "   overstated. They are kept in the file, sorted into their own block, and named on",
    "   Coverage. Cross-fund dollar sums (IssuerImpact, the universe benchmark) use reconciling",
    "   funds ONLY - a per-row flag cannot fix a contaminated sum.",
    "",
    "3. 'constant_sample? = no' means the fund is missing an endpoint - usually it launched",
    "   mid-window. Seven funds first appear at 2025-06-30. SUCH A FUND SHOWS NO MIGRATION AND",
    "   WILL LOOK LIKE THE CLEANEST BOOK IN THE FILE. It is not; it simply has no history.",
    "",
    "4. VALUATION DRAG IS NOT A RETURN. It measures revaluation of priced debt only. It",
    "   excludes interest income - which dominates BDC total return - plus realised gains,",
    "   leverage and fees. A fund at -2 pts of drag can post a +8% total return. The reported",
    "   return is carried separately, chained from fiscal YEAR-TO-DATE figures (total_return",
    "   resets each fiscal year and period_months does NOT describe it), and is left BLANK with",
    "   a reason wherever the chain breaks rather than estimated.",
    "",
    "5. A DRAG IS ONLY A FACT ABOUT A FUND IF IT RESTS ON MOST OF THE FUND. drag_basis_pct is",
    "   the share of the portfolio the drag is measured over. Measured: one fund's priced-debt",
    "   basis is 2.3% of its book, so its -2.84pt 'drag' describes a sliver. Below 50% the row",
    "   is flagged low_basis and left out of the universe benchmark.",
    "",
    "Also: weights are fixed at the START of the window, because weighting by ending fair value",
    "lets a marked-down loan shrink its own contribution. Bucket shares divide by the TAGGED",
    "portfolio, so they do NOT sum to 100% - the remainder is equity, cash, unpriced debt and",
    "holdings without par; see priced_debt_cov. Grain is ISSUER, not tranche (55% of tranche",
    "keys survive only one date), so a borrower's tranches collapse to a weighted mark.",
    "'Importance to the market' is really importance to the BDCs we track - CLOs, insurance and",
    "other private funds hold the rest of these loans and are invisible here.",
    "",
    "TABS:",
    "  MigrationMatrix - start bucket x end bucket, in $mm and in issuer counts. The diagonal is",
    "                    what stayed put; EXITED is what the fund sold or had repaid.",
    "  FundMigration   - per fund: bucket shares at both ends, what migrated, what exited, flags.",
    "  IssuerImpact    - per issuer: weighted price change, BDC-visible debt, share of universe,",
    "                    and contribution_to_universe_pts (the breadth measure).",
    "  FundAttribution - valuation drag, how much came from migrated credits, vs the",
    "                    asset-weighted universe, and vs the reported return.",
    "  Concentration   - how few issuers account for 80% of each fund's drag.",
    "  ManagerRollup   - the same cuts at parent-manager grain.",
    "  Coverage        - the reconciliation gate, excluded funds by name, threshold sensitivity.",
]


def _style(ws, n_cols: int, one_dp: list[int] | None = None,
           widths: dict[int, int] | None = None, freeze: str = "A2") -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    fill, font = PatternFill("solid", fgColor="1F4E78"), Font(bold=True, color="FFFFFF")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    for col, w in (widths or {}).items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for col in (one_dp or []):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "0.00"


def build() -> None:
    panel, tag = load_panel(), tagged_portfolio_fv()
    recon = reconciliation(panel, tag)
    usable, present = usable_funds(recon)
    p = pairs(panel)
    rr = reported_returns()

    dollars, counts = migration_matrix(p[p["cik"].isin(usable)])
    fmig = fund_migration(p, tag, recon, usable, present)
    iimp, isum = issuer_impact(p, usable)
    attr = fund_attribution(p, tag, usable, present, rr)
    conc = concentration(p, usable)
    mgr = manager_rollup(attr, fmig)
    sens = sensitivity(p, usable)

    tot_end = tag.loc[tag["reporting_date"] == END, "portfolio_fv"].sum()
    share = (tag.loc[(tag["reporting_date"] == END) & (tag["cik"].isin(usable)),
                     "portfolio_fv"].sum() / tot_end) if tot_end else 0.0
    cov = {
        "RECONCILIATION GATE": "",
        "window": f"{START} -> {END}",
        "priced-debt-with-par holding rows": int(len(panel)),
        "(fund, issuer) pairs": int(len(p)),
        "  held at both ends": int((p["status"] == "held_both").sum()),
        "  exited during the window": int((p["status"] == "exited").sum()),
        "  entered during the window": int((p["status"] == "entered").sum()),
        "funds usable at BOTH dates (reconcile)": len(usable),
        "funds present at BOTH dates": len(present),
        "funds in the file overall": int(p["cik"].nunique()),
        "share of END BDC assets covered by usable funds": f"{share:.1%}",
        "EXCLUDED FUND-DATES (named, not hidden)": "",
    }
    bad = recon[~recon["reconciles"]].sort_values("coverage", ascending=False)
    for _, r in bad.iterrows():
        c = f"{r['coverage']:.2f}x" if pd.notna(r["coverage"]) else "n/a"
        cov[f"  {r['reporting_date']} {str(r['fund_name'])[:40]}"] = f"{c} - {r['reason']}"
    cov["ITEM 5a - COHORT vs UNIVERSE"] = ""
    cov.update(isum)
    n_ret = int(attr["reported_window_total_return_pct"].notna().sum()) if not attr.empty else 0
    cov["ITEM 6 - REPORTED RETURN COVERAGE"] = ""
    cov["  funds with a complete YTD chain over the window"] = n_ret
    cov["  funds without one (blank, reason given per fund)"] = (
        int(len(attr) - n_ret) if not attr.empty else 0)
    if not attr.empty:
        nlow = int((attr["drag_basis_pct"] < MIN_DRAG_BASIS * 100).sum())
        cov["DRAG BASIS GUARD"] = ""
        cov[f"  funds whose drag rests on <{MIN_DRAG_BASIS:.0%} of the portfolio (flagged, "
            f"and excluded from the benchmark)"] = nlow
        cov["  median drag basis across funds"] = f"{attr['drag_basis_pct'].median():.1f}%"

    # Definitions, built from the columns actually being written so it cannot drift
    sheet_cols = {"MigrationMatrix": [], "FundMigration": list(fmig.columns),
                  "IssuerImpact": list(iimp.columns)}
    if not attr.empty:
        sheet_cols["FundAttribution"] = list(attr.columns)
    if not conc.empty:
        sheet_cols["Concentration"] = list(conc.columns)
    if not mgr.empty:
        sheet_cols["ManagerRollup"] = list(mgr.columns)
    sheet_cols["Coverage"] = []
    defs, def_problems = definitions_table(sheet_cols)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(WORKBOOK, engine="openpyxl") as xl:
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Overview", index=False)
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="MigrationMatrix", index=False)
        fmig.to_excel(xl, sheet_name="FundMigration", index=False)
        iimp.to_excel(xl, sheet_name="IssuerImpact", index=False)
        if not attr.empty:
            attr.to_excel(xl, sheet_name="FundAttribution", index=False)
        if not conc.empty:
            conc.to_excel(xl, sheet_name="Concentration", index=False)
        if not mgr.empty:
            mgr.to_excel(xl, sheet_name="ManagerRollup", index=False)
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Coverage", index=False)
        defs.to_excel(xl, sheet_name="Definitions", index=False)

        wb = xl.book
        from openpyxl.styles import Font
        ov = wb["Overview"]
        ov["A1"] = CAVEATS[0]
        ov["A1"].font = Font(bold=True, size=14)
        for i, line in enumerate(CAVEATS[1:], start=3):
            ov.cell(row=i, column=1, value=line)
        ov.column_dimensions["A"].width = 100

        # MigrationMatrix: the two matrices stacked, labelled
        ms = wb["MigrationMatrix"]
        ms["A1"] = f"START bucket (rows) x END bucket (columns).  Window {START} -> {END}"
        ms["A1"].font = Font(bold=True, size=12)
        r = 3
        for title, mat, fmt in (("START-WEIGHTED DOLLARS ($mm)", dollars, "0.0"),
                                ("ISSUER COUNTS", counts, "0")):
            ms.cell(row=r, column=1, value=title).font = Font(bold=True)
            r += 1
            ms.cell(row=r, column=1, value="start \\ end").font = Font(bold=True)
            for j, c in enumerate(mat.columns, start=2):
                ms.cell(row=r, column=j, value=str(c)).font = Font(bold=True)
            r += 1
            for idx, row in mat.iterrows():
                ms.cell(row=r, column=1, value=str(idx)).font = Font(bold=True)
                for j, c in enumerate(mat.columns, start=2):
                    cell = ms.cell(row=r, column=j, value=float(row[c]))
                    cell.number_format = fmt
                r += 1
            r += 2
        ms.column_dimensions["A"].width = 18

        cv = wb["Coverage"]
        cv["A1"] = "Coverage, exclusions and threshold sensitivity"
        cv["A1"].font = Font(bold=True, size=13)
        rr_ = 3
        for k, v in cov.items():
            cell = cv.cell(row=rr_, column=1, value=k)
            if v == "":
                cell.font = Font(bold=True)
            else:
                cv.cell(row=rr_, column=2, value=v)
            rr_ += 1
        rr_ += 1
        cv.cell(row=rr_, column=1, value="THRESHOLD SENSITIVITY").font = Font(bold=True)
        rr_ += 1
        for j, c in enumerate(sens.columns, start=1):
            cv.cell(row=rr_, column=j, value=str(c)).font = Font(bold=True)
        rr_ += 1
        for _, srow in sens.iterrows():
            for j, c in enumerate(sens.columns, start=1):
                cv.cell(row=rr_, column=j, value=srow[c])
            rr_ += 1
        cv.column_dimensions["A"].width = 56
        cv.column_dimensions["B"].width = 46

        _style(wb["FundMigration"], fmig.shape[1], widths={2: 34}, freeze="C2")
        _style(wb["IssuerImpact"], iimp.shape[1], widths={1: 30})
        if not attr.empty:
            _style(wb["FundAttribution"], attr.shape[1], widths={2: 34, 15: 30}, freeze="C2")
        if not conc.empty:
            _style(wb["Concentration"], conc.shape[1], widths={1: 34, 5: 52})
        if not mgr.empty:
            _style(wb["ManagerRollup"], mgr.shape[1], widths={1: 26})

        # Definitions: wrapped prose, grouped by tab with a visible band on each tab's first row
        ds = wb["Definitions"]
        _style(ds, defs.shape[1], widths={1: 17, 2: 30, 3: 62, 4: 62, 5: 62}, freeze="C2")
        from openpyxl.styles import Alignment as _Al, Font as _Fn, PatternFill as _Pf
        band = _Pf("solid", fgColor="DCE6F1")
        prev = None
        for i, tab in enumerate(defs["Tab"], start=2):
            for c in range(1, defs.shape[1] + 1):
                cell = ds.cell(row=i, column=c)
                cell.alignment = _Al(vertical="top", wrap_text=True)
                if tab != prev:
                    cell.fill = band
                    if c <= 2:
                        cell.font = _Fn(bold=True)
            prev = tab

    print(f"wrote {WORKBOOK}")
    print(f"  Definitions: {len(defs)} entries across {defs['Tab'].nunique()} tabs")
    if def_problems:
        print("  DEFINITIONS DRIFT - fix before sharing:")
        for pb in def_problems:
            print(f"    {pb}")
    else:
        print("  every written column has a definition; no orphans")
    print(f"  usable funds {len(usable)} / {p['cik'].nunique()} in file, {share:.1%} of assets")
    print(f"  pairs {len(p):,}  (held_both {(p['status']=='held_both').sum():,}, "
          f"exited {(p['status']=='exited').sum():,}, entered {(p['status']=='entered').sum():,})")
    print(f"  issuers on IssuerImpact {len(iimp):,}   reported-return chains {n_ret}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--gate", action="store_true",
                    help="checkpoint 1: print the reconciliation gate and stop")
    ap.add_argument("--build", action="store_true", help="write the full workbook")
    args = ap.parse_args()
    if args.gate:
        run_gate()
    elif args.build:
        build()
    else:
        print("nothing to do; pass --gate (checkpoint 1) or --build")
