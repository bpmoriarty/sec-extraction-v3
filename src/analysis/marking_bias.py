"""
marking_bias.py — Do some managers value the same loans richer / cheaper than peers?

Plan: docs/MARKING_BIAS_PLAN.md. Reuses the cleaned/clustered/matched holdings pipeline
(holdings_compare.py) but writes its OWN workbook (data/dataset/marking_bias.xlsx).

Unit of analysis = a manager's deviation from the LEAVE-ONE-OUT consensus on the same loan at the
same date (points of par). Comparing within the identical issue controls for which loans a manager
holds, so this measures relative valuation tendency, not portfolio mix. Positive = marks ABOVE
peers (richer / more optimistic); negative = below (cheaper / more conservative). Descriptive, not
a judgment of who is correct.

Layers (per the plan):
  - Descriptive: median/mean deviation, % rich/in-line/cheap, ex-distressed cut.
  - Robust uncertainty: bootstrap CI on the median (resampling the manager's own per-issue
    deviations — an issue-level resample for that manager, NOT a cross-manager cluster bootstrap)
    + a sign test.
  - Formal model: OLS of the leave-one-out deviation on manager dummies with issue-clustered SEs.
    The leave-one-out deviation is preferred over include-self issue FE, which attenuates manager
    coefficients by k/(k-1) on thin loans. BH-FDR computed over >=20-loan managers only.
  - Drift: median deviation by reporting date.

Run:  uv run python src/analysis/marking_bias.py [--threshold 92]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "src" / "analysis"))
from holdings_compare import add_clusters, fund_marks, load_consolidated, match_issues  # noqa: E402
from managers import manager_of  # noqa: E402

OUT_DIR = PROJECT_ROOT / "data" / "dataset"
WORKBOOK = OUT_DIR / "marking_bias.xlsx"
CHART_DIR = OUT_DIR / "_bias_charts"

MIN_MANAGERS_PER_ISSUE = 3     # need >=3 managers so the leave-one-out median is over >=2 others
MIN_ISSUES_RANK = 20           # below this a manager is shown but flagged low-confidence
DISTRESSED_BELOW = 0.90        # issue consensus < this = distressed (for the ex-distressed cut)
RICH_CHEAP_PTS = 1.0           # |dev| <= this = "in line"
N_BOOT = 800


# ---------------------------------------------------------------------------
# Build the (manager, issue) deviation table
# ---------------------------------------------------------------------------

def build_deviations(threshold: int = 92) -> pd.DataFrame:
    """One row per (issue, manager): the manager's mark and its leave-one-out deviation (points)."""
    df = add_clusters(load_consolidated(), threshold=threshold)
    holdings, issues = match_issues(df)
    fm = fund_marks(holdings, issues)
    fm = fm[fm["confidence"].isin(["High", "Medium"])].copy()   # reliable tranche matches only
    fm["manager"] = fm["cik"].map(manager_of)

    # one mark per (issue, manager): collapse a manager's vehicles to their median
    mi = (fm.groupby(["issue_id", "manager"])
          .agg(mark=("cmp_price", "median"),
               issuer_cluster=("issuer_cluster", "first"),
               reporting_date=("reporting_date", "first"),
               issue_med=("price_median", "first")).reset_index())

    # keep issues with >=3 managers
    n_mgr = mi.groupby("issue_id")["manager"].transform("size")
    mi = mi[n_mgr >= MIN_MANAGERS_PER_ISSUE].copy()

    # leave-one-out consensus per issue
    devs = []
    for iid, g in mi.groupby("issue_id"):
        marks = g["mark"].to_numpy()
        for pos, (_, row) in enumerate(g.iterrows()):
            loo = np.median(np.delete(marks, pos))
            devs.append(((row["mark"] - loo) * 100.0))
    mi["dev_pts"] = devs
    mi["distressed"] = mi["issue_med"] < DISTRESSED_BELOW
    return mi


# ---------------------------------------------------------------------------
# Per-manager statistics
# ---------------------------------------------------------------------------

def _boot_median_ci(devs: np.ndarray, rng) -> tuple[float, float]:
    """Bootstrap 95% CI on a manager's median deviation. Resamples the manager's own per-issue
    deviations (an issue-level resample for THAT manager) — not a cross-manager cluster bootstrap."""
    if len(devs) < 3:
        return (float("nan"), float("nan"))
    samp = rng.choice(devs, size=(N_BOOT, len(devs)), replace=True)
    meds = np.median(samp, axis=1)
    return (round(float(np.percentile(meds, 2.5)), 2), round(float(np.percentile(meds, 97.5)), 2))


def _sign_test_p(devs: np.ndarray) -> float:
    from scipy.stats import binomtest
    pos = int((devs > 0.01).sum())
    neg = int((devs < -0.01).sum())
    if pos + neg == 0:
        return 1.0
    return float(binomtest(pos, pos + neg, 0.5).pvalue)


def manager_table(mi: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(7)
    rows = []
    for mgr, g in mi.groupby("manager"):
        d = g["dev_pts"].to_numpy()
        d_ex = g.loc[~g["distressed"], "dev_pts"].to_numpy()
        lo, hi = _boot_median_ci(d, rng)
        rows.append({
            "Manager": mgr,
            "n_issues": len(d),
            "n_dates": g["reporting_date"].nunique(),
            "median_dev": round(float(np.median(d)), 2),
            "mean_dev": round(float(np.mean(d)), 2),
            "boot_lo": lo, "boot_hi": hi,
            "sign_p": round(_sign_test_p(d), 4),
            "pct_rich": round(float((d > RICH_CHEAP_PTS).mean() * 100), 1),
            "pct_inline": round(float((np.abs(d) <= RICH_CHEAP_PTS).mean() * 100), 1),
            "pct_cheap": round(float((d < -RICH_CHEAP_PTS).mean() * 100), 1),
            "median_dev_exdistressed": (round(float(np.median(d_ex)), 2) if len(d_ex) else None),
        })
    tab = pd.DataFrame(rows)
    tab["low_confidence"] = tab["n_issues"] < MIN_ISSUES_RANK
    return tab


def fixed_effects(mi: pd.DataFrame) -> pd.DataFrame:
    """OLS of the leave-one-out deviation on manager dummies with issue-clustered SEs.

    The DV is the leave-one-out deviation: each manager's mark minus the median of the OTHER
    managers on that issue. The issue-level consensus (excluding self) is already differenced out,
    so regressing on manager dummies recovers each manager's issue-controlled average bias.

    This estimator is preferred over an include-self issue fixed effect (mark ~ C(manager) +
    C(issue_id)) because the include-self form reintroduces the self-inclusion it tries to remove:
    the issue FE absorbs part of each manager's own mark, attenuating coefficients by k/(k-1)
    (worst on thin loans: -33% at k=3). Synthetic calibration confirmed the leave-one-out form
    recovers a known +2-pt bias accurately across holder counts; include-self understates it.

    Returns raw fe_p only — BH-FDR is computed in build() over the >=20-loan subset exclusively."""
    import statsmodels.api as sm

    d = mi.dropna(subset=["dev_pts"]).copy()
    D = pd.get_dummies(d["manager"]).astype(float)
    groups = d["issue_id"].astype("category").cat.codes.to_numpy()
    res = sm.OLS(d["dev_pts"].to_numpy(), D.to_numpy()).fit(
        cov_type="cluster", cov_kwds={"groups": groups})
    out = pd.DataFrame({
        "Manager": D.columns,
        "fe_coef": np.round(res.params, 2),
        "fe_se": np.round(res.bse, 2),
        "fe_p": np.round(res.pvalues, 4),
    })
    out["fe_ci_lo"] = np.round(out["fe_coef"] - 1.96 * out["fe_se"], 2)
    out["fe_ci_hi"] = np.round(out["fe_coef"] + 1.96 * out["fe_se"], 2)
    # fe_p_fdr is NOT computed here; it is computed in build() over eligible managers only.
    return out


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------

def _caterpillar(tab: pd.DataFrame, path: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    t = tab[tab["n_issues"] >= MIN_ISSUES_RANK].sort_values("fe_coef")
    if t.empty:
        return
    y = range(len(t))
    colors = ["#c0392b" if c > 0 else "#2471a3" for c in t["fe_coef"]]
    fig, ax = plt.subplots(figsize=(9, max(4, 0.32 * len(t))))
    ax.errorbar(t["fe_coef"], y,
                xerr=[t["fe_coef"] - t["fe_ci_lo"], t["fe_ci_hi"] - t["fe_coef"]],
                fmt="o", ecolor="gray", elinewidth=1, capsize=3, mfc="white", mec="gray")
    ax.scatter(t["fe_coef"], y, c=colors, zorder=3, s=36)
    ax.axvline(0, color="black", lw=1)
    ax.set_yticks(list(y)); ax.set_yticklabels(t["Manager"], fontsize=8)
    ax.set_xlabel("Issue-controlled marking bias (points of par)  —  +rich / −cheap")
    ax.set_title("Manager marking bias (OLS coefficient ± 95% issue-clustered CI)", fontsize=11)
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)


def _boxplot(mi: pd.DataFrame, tab: pd.DataFrame, path: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    keep = tab[tab["n_issues"] >= MIN_ISSUES_RANK].sort_values("median_dev")["Manager"].tolist()
    if not keep:
        return
    data = [mi.loc[mi["manager"] == m, "dev_pts"].clip(-30, 30).to_numpy() for m in keep]
    fig, ax = plt.subplots(figsize=(10, max(4, 0.34 * len(keep))))
    ax.boxplot(data, orientation="horizontal", showfliers=False, whis=(10, 90))
    ax.set_yticklabels(keep, fontsize=8)
    ax.axvline(0, color="red", lw=1)
    ax.set_xlabel("Per-loan deviation from peers (points, clipped ±30; whiskers 10–90 pct)")
    ax.set_title("Distribution of per-loan marking deviations by manager", fontsize=11)
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)


# ---------------------------------------------------------------------------
# Assemble
# ---------------------------------------------------------------------------

def build(threshold: int = 92) -> None:
    mi = build_deviations(threshold=threshold)
    if mi.empty:
        raise SystemExit("no (manager, issue) deviations — nothing to analyze")
    tab = manager_table(mi)
    fe = fixed_effects(mi)
    table = tab.merge(fe, on="Manager", how="left").sort_values("fe_coef", ascending=False)
    # ── Thin-manager significance guard ──────────────────────────────────────────────────────────
    # The cluster-robust SE can collapse to a false p≈0 when a manager has very few comparable loans
    # (e.g. Fidus n=2: fe_se=0.19, raw p=0.000 — an artefact, not a finding). Calibration simulations
    # show CI coverage is only 74% at n=2-4 and 84% at 4-6; it reaches ~92% at n≥20 (MIN_ISSUES_RANK).
    # Resolution: blank the FE inference below the floor and restrict FDR to the eligible family only.
    from statsmodels.stats.multitest import multipletests
    eligible = table["n_issues"] >= MIN_ISSUES_RANK

    # Option 1: blank FE regression inference below the floor. KEEP the point estimate (Bias, pts)
    # as an indicative direction; KEEP bootstrap CI + sign test (those degrade honestly by WIDENING,
    # not collapsing, so they remain a useful visible uncertainty signal for thin managers).
    table.loc[~eligible, ["fe_p", "fe_p_fdr", "fe_se", "fe_ci_lo", "fe_ci_hi"]] = np.nan

    # Option 2: recompute BH-FDR over the eligible family only so thin managers don't contaminate it.
    table["fe_p_fdr"] = np.nan
    table.loc[eligible, "fe_p_fdr"] = np.round(
        multipletests(table.loc[eligible, "fe_p"], method="fdr_bh")[1], 4)

    # Significant (robust): requires >=20 loans AND FE model significant after FDR AND bootstrap
    # median CI excludes 0 in the same direction as the FE coefficient. A small p alone is never
    # sufficient — both tests must agree, in the same direction, on sufficient data.
    table["significant_robust"] = (
        eligible
        & (table["fe_p_fdr"] < 0.05)
        & table["boot_lo"].notna() & table["boot_hi"].notna()
        & (np.sign(table["boot_lo"]) == np.sign(table["boot_hi"]))
        & (np.sign(table["boot_lo"]) == np.sign(table["fe_coef"]))
    )
    # ─────────────────────────────────────────────────────────────────────────────────────────────

    by_date = (mi.groupby(["manager", "reporting_date"])["dev_pts"].median().round(2)
               .unstack("reporting_date"))
    by_date = by_date.reindex(sorted(by_date.columns), axis=1).reset_index()

    # issue sample (>=4 managers), each manager's mark vs the issue consensus, for spot-checking
    big = mi.groupby("issue_id")["manager"].transform("size") >= 4
    samp_ids = (pd.Series(mi.loc[big, "issue_id"].unique())
                .sample(min(40, mi.loc[big, "issue_id"].nunique()), random_state=11))
    srows = []
    for iid in samp_ids:
        g = mi[mi["issue_id"] == iid].sort_values("mark")
        marks = "  ".join(f"{m}:{v*100:.0f}" for m, v in zip(g["manager"], g["mark"]))
        srows.append({"issue_id": iid, "Issuer": g["issuer_cluster"].iat[0],
                      "Date": g["reporting_date"].iat[0],
                      "Consensus (pts)": round(g["issue_med"].iat[0] * 100, 1),
                      "Managers (mgr:pts, low->high)": marks, "Verdict (Y/N)": "", "Notes": ""})
    issue_sample = pd.DataFrame(srows)

    fmap = (pd.read_csv(OUT_DIR / "fund_manager_map.csv", dtype=str)
            if (OUT_DIR / "fund_manager_map.csv").exists() else pd.DataFrame())

    CHART_DIR.mkdir(parents=True, exist_ok=True)
    cat, box = CHART_DIR / "caterpillar.png", CHART_DIR / "boxplot.png"
    _caterpillar(table, cat)
    _boxplot(mi, table, box)

    _write(table, by_date, issue_sample, fmap, mi, cat, box)
    print(f"wrote {WORKBOOK}  ({mi['manager'].nunique()} managers, {len(mi):,} manager-issue obs)")


def _write(table, by_date, issue_sample, fmap, mi, cat, box) -> None:
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    disp = {
        "Manager": "Manager", "n_issues": "n issues", "n_dates": "n dates",
        "fe_coef": "Bias (pts)", "fe_ci_lo": "CI low", "fe_ci_hi": "CI high",
        "fe_p_fdr": "p (FDR)", "significant_robust": "Significant (robust)",
        "median_dev": "Median dev", "boot_lo": "Median CI low",
        "boot_hi": "Median CI high", "sign_p": "Sign-test p", "pct_rich": "% rich",
        "pct_inline": "% in-line", "pct_cheap": "% cheap",
        "median_dev_exdistressed": "Median dev (ex-distressed)", "low_confidence": "Low-confidence",
    }
    overview = [
        "BDC Manager Marking Bias — who values the same loans richer/cheaper than peers.",
        "",
        "Unit: a manager's deviation from the LEAVE-ONE-OUT consensus on the SAME loan at the SAME",
        "date (points of par). Within-issue, so it controls for which loans a manager holds.",
        "POSITIVE = marks ABOVE peers (richer / more optimistic); NEGATIVE = below (more conservative).",
        "It is DESCRIPTIVE — a relative tendency among overlapping loans, not a judgment of who is right.",
        "",
        "Reading the ManagerBias tab:",
        "  'n issues' is the manager's effective cluster count for the regression.",
        "  Significance (p (FDR), CI, Significant (robust)) is intentionally BLANK for managers with",
        "  fewer than 20 comparable loans — the cluster-robust SE is unreliable below that floor and",
        "  can produce false p~0 (e.g. a manager with 2 loans may show a precise-looking p=0.000 that",
        "  is a statistical artefact). 'Bias (pts)' is still shown as an indicative direction.",
        "  'Significant (robust)' = True only when ALL of: (a) >=20 comparable loans, (b) FE model",
        "  significant after BH-FDR (computed over the >=20-loan managers only), (c) bootstrap median",
        "  CI excludes 0 in the same direction as the FE coefficient. A small p (FDR) alone is not",
        "  sufficient — both tests must agree. See docs/MARKING_BIAS_PLAN.md §7 for full methodology.",
        "",
        "Tabs:",
        "  ManagerBias   — per manager: Bias (pts) + CI + p (FDR) + Significant (robust) +",
        "                  median deviation, bootstrap CI, sign test, % rich/cheap, ex-distressed cut.",
        "  ManagerBiasByDate — median deviation by reporting date (drift over time).",
        "  IssueSample   — sampled loans with each manager's mark vs the consensus (spot-check; Verdict).",
        "  FundManagerMap — the fund->manager rollup used.",
        "  Charts        — caterpillar (bias ± CI) and per-loan deviation box plots.",
        "",
        f"Built on issues held by >= {MIN_MANAGERS_PER_ISSUE} managers, High/Medium-confidence matches.",
        "Caveats: marks are quarterly manager estimates of illiquid loans; only loans matched on shared",
        "dates count; managers with thin overlap get wide CIs (flagged). A non-zero grand mean would",
        "signal a bug — by construction deviations roughly center at zero across managers.",
        "",
        "Universe & coverage caveats (see _verification/DATA_INTEGRITY_REPORT.md for detail):",
        "1. SURVIVOR-ONLY UNIVERSE. The fund universe was seeded from current Morningstar lists and",
        "   contains only BDCs that still exist. An estimated 77 BDCs that deregistered during the",
        "   XBRL window (2016+) are absent — many are the distressed failures (American Capital,",
        "   Medallion, Sierra Income/Medley, Garrison, Logan Ridge, Alcentra, OHA, Newtek). The",
        "   bias rankings are therefore built on the surviving, performing subset and may understate",
        "   cross-manager dispersion. Do NOT present these rankings as universe-complete.",
        "2. OVERLAP-ONLY COMPARISON. Only managers whose books overlap with others can be compared.",
        "   Approximately 18 of ~57 managers contribute zero comparable observations and are absent",
        "   from the rankings — every venture/specialty lender (Hercules, Horizon Technology, Runway",
        "   Growth, Trinity Capital, TriplePoint, Firsthand, Advanced Flower Capital, Chicago",
        "   Atlantic, Gladstone, Princeton, Rand, Saratoga, PGIM, Fidelity, PennantPark, SLR) plus",
        "   2 lost to SOI parse format (BlackRock, AllianceBernstein). Findings describe the",
        "   club-deal / broadly-syndicated direct-lending core, not venture-debt or specialty lenders.",
        "3. EXACT-DATE MATCHING. Only managers reporting on the same fiscal period-end are ever",
        "   compared. Funds with different quarter-ends do not co-appear on shared dates.",
    ]
    with pd.ExcelWriter(WORKBOOK, engine="openpyxl") as xl:
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Overview", index=False)
        cols = ["Manager", "n_issues", "n_dates", "fe_coef", "fe_ci_lo", "fe_ci_hi", "fe_p_fdr",
                "significant_robust", "median_dev", "boot_lo", "boot_hi", "sign_p",
                "pct_rich", "pct_inline", "pct_cheap",
                "median_dev_exdistressed", "low_confidence"]
        table[cols].rename(columns=disp).to_excel(xl, sheet_name="ManagerBias", index=False)
        by_date.to_excel(xl, sheet_name="ManagerBiasByDate", index=False)
        issue_sample.to_excel(xl, sheet_name="IssueSample", index=False)
        if not fmap.empty:
            fmap.to_excel(xl, sheet_name="FundManagerMap", index=False)
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Charts", index=False)

        wb = xl.book
        ov = wb["Overview"]
        ov["A1"].font = Font(bold=True, size=14)
        for i, line in enumerate(overview, start=1):
            ov.cell(row=i, column=1, value=line)
        ov.column_dimensions["A"].width = 100

        hdr_fill, hdr_font = PatternFill("solid", fgColor="1F4E78"), Font(bold=True, color="FFFFFF")
        for name in ("ManagerBias", "ManagerBiasByDate", "IssueSample", "FundManagerMap"):
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill, cell.font = hdr_fill, hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        wb["ManagerBias"].column_dimensions["A"].width = 24
        wb["IssueSample"].column_dimensions[get_column_letter(5)].width = 70

        ch = wb["Charts"]
        ch["A1"] = "Manager marking bias — coefficient (±95% CI) and per-loan deviation spread"
        ch["A1"].font = Font(bold=True, size=12)
        row = 3
        for img in (cat, box):
            if Path(img).exists():
                ch.add_image(XLImage(str(img)), f"A{row}")
                row += 48


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--threshold", type=int, default=92, help="fuzzy merge threshold (Phase 2)")
    args = ap.parse_args()
    build(threshold=args.threshold)
