"""
portfolio_overlap.py — Pairwise portfolio-overlap analysis across BDC funds.

Plan: docs/PORTFOLIO_OVERLAP_PLAN.md. Reuses the cleaned/clustered/matched holdings pipeline
(holdings_compare.py) but writes its OWN workbook (data/dataset/portfolio_overlap.xlsx) — nothing
mixes into the marks-comparison output.

For each reporting date, for each pair of funds, at two grains (issuer = same borrower; issue =
same specific loan), we compute how much of their books is the same:
  - common count, directional coverage (|A∩B|/|A| and |A∩B|/|B| — the "vice versa"),
  - Jaccard (size-fair similarity), overlap coefficient (handles big-vs-small),
  - $-weighted overlap (Σ min(weightA, weightB)), and
  - expected overlap + lift (vs a random draw of the same sizes from that date's universe).
Pairs carry a same-manager flag so mechanical intra-manager co-investment can be separated from
genuine cross-manager club deals.

Run:  uv run python src/analysis/portfolio_overlap.py [--threshold 92]
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from itertools import combinations
from pathlib import Path

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "src" / "analysis"))
from holdings_compare import add_clusters, load_consolidated, match_issues  # noqa: E402
from managers import manager_of  # noqa: E402

OUT_DIR = PROJECT_ROOT / "data" / "dataset"
WORKBOOK = OUT_DIR / "portfolio_overlap.xlsx"
CHART_DIR = OUT_DIR / "_overlap_charts"

MIN_HOLDINGS = 5      # ignore a fund-date with fewer than this many priced names (too thin)
NET_EDGE_JACCARD = 0.12   # draw a network edge when issuer-Jaccard >= this


# ---------------------------------------------------------------------------
# Core: per-date pairwise overlap at one grain
# ---------------------------------------------------------------------------

def _pairs_for_date(sub: pd.DataFrame, key: str) -> list[dict]:
    """All overlapping fund pairs for one reporting date at one grain (`key` = issuer_cluster or
    issue_id). Returns a list of metric dicts (only pairs sharing >=1 element)."""
    sub = sub[sub[key].notna()]
    if sub.empty:
        return []
    # weight = a name's fair value as a share of the fund's total (that grain, that date)
    fv = sub.groupby(["cik", key])["fair_value"].sum().abs()
    tot = fv.groupby(level=0).sum()
    fund_sets: dict[str, set] = {}
    fund_w: dict[str, dict] = {}
    for cik, s in fv.groupby(level=0):
        if tot[cik] <= 0 or s.shape[0] < MIN_HOLDINGS:
            continue
        elems = s.droplevel(0)
        fund_sets[cik] = set(elems.index)
        fund_w[cik] = (elems / tot[cik]).to_dict()
    if len(fund_sets) < 2:
        return []

    universe = len({e for s in fund_sets.values() for e in s})
    # candidate pairs via an inverted index (only pairs that share >=1 element)
    inv: dict = defaultdict(list)
    for f, s in fund_sets.items():
        for e in s:
            inv[e].append(f)
    cand: set = set()
    for fs in inv.values():
        if len(fs) >= 2:
            cand.update(combinations(sorted(fs), 2))

    rows = []
    for a, b in cand:
        sa, sb = fund_sets[a], fund_sets[b]
        common = sa & sb
        inter = len(common)
        if not inter:
            continue
        na, nb = len(sa), len(sb)
        union = na + nb - inter
        wa, wb = fund_w[a], fund_w[b]
        wov = sum(min(wa[e], wb[e]) for e in common)          # Σ min(weight) — $-weighted overlap
        expected = na * nb / universe if universe else 0.0     # E[|A∩B|] for random draws
        rows.append({
            "cik_a": a, "cik_b": b,
            "n_a": na, "n_b": nb, "common": inter,
            "cov_a_in_b": round(inter / na, 4), "cov_b_in_a": round(inter / nb, 4),
            "jaccard": round(inter / union, 4),
            "overlap_coef": round(inter / min(na, nb), 4),
            "wt_overlap": round(wov, 4),
            "lift": round(inter / expected, 2) if expected > 0 else None,
        })
    return rows


def compute_overlap(threshold: int = 92):
    """Returns (pairs_panel, fund_names) — pairs_panel has one row per (date, grain, fund-pair).

    Issuer grain uses all parsed holdings (issuer_cluster).
    Issue grain uses ONLY High/Medium-confidence issues — seniority-only (Low/Single) matches are
    excluded so the issue grain reflects genuinely-matched tranches, not coincidentally shared IDs."""
    df = add_clusters(load_consolidated(), threshold=threshold)
    holdings, issues = match_issues(df)
    h = holdings[holdings["parse_ok"] & holdings["issuer_cluster"].notna()].copy()
    h["fair_value"] = pd.to_numeric(h["fair_value"], errors="coerce")
    # Blank issue_id for Low/Single confidence so the issue-grain pass only sees genuine matches.
    # Issuer grain keys on issuer_cluster and is unaffected.
    himed = set(issues.loc[issues["confidence"].isin(["High", "Medium"]), "issue_id"])
    h["issue_id"] = h["issue_id"].where(h["issue_id"].isin(himed))
    fund_names = h.drop_duplicates("cik").set_index("cik")["fund_name"].to_dict()

    panel = []
    for dt, day in h.groupby("reporting_date"):
        for grain, key in (("issuer", "issuer_cluster"), ("issue", "issue_id")):
            for r in _pairs_for_date(day, key):
                r["reporting_date"] = dt
                r["grain"] = grain
                panel.append(r)
    pairs = pd.DataFrame(panel)
    if not pairs.empty:
        pairs["fund_a"] = pairs["cik_a"].map(fund_names)
        pairs["fund_b"] = pairs["cik_b"].map(fund_names)
        pairs["mgr_a"] = pairs["cik_a"].map(manager_of)
        pairs["mgr_b"] = pairs["cik_b"].map(manager_of)
        pairs["same_manager"] = (pairs["mgr_a"] == pairs["mgr_b"])
    return pairs, fund_names


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------

def _heatmap(jac: pd.DataFrame, path: Path, title: str) -> None:
    """Clustered heatmap of an issuer-Jaccard matrix (funds ordered so co-lending groups sit
    together)."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from scipy.cluster.hierarchy import leaves_list, linkage
    from scipy.spatial.distance import squareform

    order = list(jac.index)
    if len(order) >= 3:
        dist = (1 - jac).clip(lower=0).values
        dist = (dist + dist.T) / 2
        for i in range(len(dist)):
            dist[i, i] = 0.0
        try:
            order = [jac.index[i] for i in leaves_list(linkage(squareform(dist, checks=False),
                                                               method="average"))]
        except Exception:
            pass
    m = jac.loc[order, order]
    fig, ax = plt.subplots(figsize=(min(22, 0.32 * len(order) + 4),) * 2)
    im = ax.imshow(m.values, cmap="YlOrRd", vmin=0, vmax=min(1.0, float(m.values.max() or 1)))
    ax.set_xticks(range(len(order))); ax.set_yticks(range(len(order)))
    ax.set_xticklabels([s[:22] for s in m.columns], rotation=90, fontsize=6)
    ax.set_yticklabels([s[:22] for s in m.index], fontsize=6)
    ax.set_title(title, fontsize=11)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="Jaccard overlap")
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


def _network(jac: pd.DataFrame, names: dict, path: Path, title: str) -> None:
    """Co-lending network: funds = nodes, an edge when issuer-Jaccard >= NET_EDGE_JACCARD, node
    color = detected co-lending community."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import networkx as nx

    g = nx.Graph()
    for f in jac.index:
        g.add_node(f)
    for i, a in enumerate(jac.index):
        for b in jac.index[i + 1:]:
            w = float(jac.loc[a, b])
            if w >= NET_EDGE_JACCARD:
                g.add_edge(a, b, weight=w)
    g.remove_nodes_from([n for n in list(g.nodes) if g.degree(n) == 0])
    if g.number_of_nodes() < 3:
        return
    try:
        comms = list(nx.community.greedy_modularity_communities(g, weight="weight"))
    except Exception:
        comms = [set(g.nodes)]
    color_of = {n: i for i, c in enumerate(comms) for n in c}
    pos = nx.spring_layout(g, weight="weight", seed=7, k=0.6)
    fig, ax = plt.subplots(figsize=(16, 12))
    nx.draw_networkx_edges(g, pos, ax=ax, alpha=0.25,
                           width=[2.5 * g[u][v]["weight"] for u, v in g.edges])
    nx.draw_networkx_nodes(g, pos, ax=ax, node_size=320,
                           node_color=[color_of.get(n, 0) for n in g.nodes], cmap="tab20")
    nx.draw_networkx_labels(g, pos, ax=ax,
                            labels={n: str(names.get(n, n))[:18] for n in g.nodes}, font_size=7)
    ax.set_title(title, fontsize=12); ax.axis("off")
    fig.tight_layout(); fig.savefig(path, dpi=130); plt.close(fig)


def _trend_chart(trend_long: pd.DataFrame, path: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(12, 7))
    for lbl, grp in trend_long.groupby("pair"):
        grp = grp.sort_values("reporting_date")
        ax.plot(grp["reporting_date"], grp["jaccard"], marker="o", label=lbl[:40])
    ax.set_ylabel("Issuer Jaccard overlap"); ax.set_xlabel("Reporting date")
    ax.set_title("Top cross-manager fund pairs — overlap over time")
    ax.legend(fontsize=7, ncol=2); ax.grid(alpha=0.3)
    plt.xticks(rotation=45, ha="right"); fig.tight_layout()
    fig.savefig(path, dpi=130); plt.close(fig)


# ---------------------------------------------------------------------------
# Assemble workbook
# ---------------------------------------------------------------------------

def build(threshold: int = 92) -> None:
    pairs, names = compute_overlap(threshold=threshold)
    if pairs.empty:
        raise SystemExit("no overlapping pairs found")

    # best snapshot date = the date with the most funds present (issuer grain)
    cov = (pairs[pairs["grain"] == "issuer"]
           .assign(f=lambda d: d[["cik_a", "cik_b"]].values.tolist()))
    funds_per_date = (pairs[pairs["grain"] == "issuer"]
                      .groupby("reporting_date")
                      .apply(lambda d: len(set(d["cik_a"]) | set(d["cik_b"])), include_groups=False))
    best_date = funds_per_date.idxmax()

    iss = pairs[pairs["grain"] == "issuer"]
    isu = pairs[pairs["grain"] == "issue"]

    # PairsLatest: merge issuer + issue metrics for the best date, one row per pair
    li = iss[iss["reporting_date"] == best_date].copy()
    lu = isu[isu["reporting_date"] == best_date][
        ["cik_a", "cik_b", "common", "jaccard", "cov_a_in_b", "cov_b_in_a", "wt_overlap", "lift"]
    ].rename(columns={c: c + "_issue" for c in
                      ["common", "jaccard", "cov_a_in_b", "cov_b_in_a", "wt_overlap", "lift"]})
    latest = li.merge(lu, on=["cik_a", "cik_b"], how="left").sort_values("jaccard", ascending=False)
    latest_cols = {
        "fund_a": "Fund A", "fund_b": "Fund B", "mgr_a": "Mgr A", "mgr_b": "Mgr B",
        "same_manager": "Same mgr", "n_a": "A names", "n_b": "B names",
        "common": "Common issuers", "cov_a_in_b": "A in B", "cov_b_in_a": "B in A",
        "jaccard": "Issuer Jaccard", "overlap_coef": "Issuer overlap-coef",
        "wt_overlap": "Issuer $-overlap", "lift": "Issuer lift",
        "common_issue": "Common issues", "jaccard_issue": "Issue Jaccard",
        "wt_overlap_issue": "Issue $-overlap", "lift_issue": "Issue lift",
    }

    # FundSummary (best date, issuer grain)
    fs_rows = []
    for cik, nm in names.items():
        mine = li[(li["cik_a"] == cik) | (li["cik_b"] == cik)]
        if mine.empty:
            continue
        partner = mine.apply(lambda r: r["cik_b"] if r["cik_a"] == cik else r["cik_a"], axis=1)
        jac = mine["jaccard"]
        top_i = jac.idxmax()
        nn = (li.loc[top_i, "cik_b"] if li.loc[top_i, "cik_a"] == cik else li.loc[top_i, "cik_a"])
        fs_rows.append({
            "Fund": nm, "Manager": manager_of(cik),
            "Overlapping partners (J>=0.1)": int((jac >= 0.1).sum()),
            "Nearest fund": names.get(nn, nn), "Max Jaccard": round(float(jac.max()), 3),
            "Mean Jaccard": round(float(jac.mean()), 3),
            "Differentiation (1-maxJ)": round(1 - float(jac.max()), 3),
        })
    fund_summary = pd.DataFrame(fs_rows).sort_values("Max Jaccard", ascending=False)

    # Trend: top cross-manager pairs (by best-date issuer Jaccard) tracked over all dates
    xmgr = latest[(~latest["same_manager"])].head(12)
    key_pairs = set(zip(xmgr["cik_a"], xmgr["cik_b"]))
    tr = iss[iss.apply(lambda r: (r["cik_a"], r["cik_b"]) in key_pairs, axis=1)].copy()
    tr["pair"] = tr["fund_a"].str[:18] + " / " + tr["fund_b"].str[:18]
    trend_long = tr[["pair", "reporting_date", "jaccard"]].copy()
    trend_wide = (trend_long.pivot_table(index="pair", columns="reporting_date", values="jaccard")
                  .reset_index())

    # issuer-Jaccard matrix for the best date (heatmap + network)
    present = sorted(set(li["cik_a"]) | set(li["cik_b"]))
    jac_m = pd.DataFrame(1.0, index=present, columns=present)
    for _, r in li.iterrows():
        jac_m.loc[r["cik_a"], r["cik_b"]] = r["jaccard"]
        jac_m.loc[r["cik_b"], r["cik_a"]] = r["jaccard"]
    name_idx = [names.get(c, c) for c in present]
    jac_named = jac_m.copy(); jac_named.index = name_idx; jac_named.columns = name_idx

    CHART_DIR.mkdir(parents=True, exist_ok=True)
    hp = CHART_DIR / "heatmap.png"
    npth = CHART_DIR / "network.png"
    tp = CHART_DIR / "trend.png"
    _heatmap(jac_named, hp, f"Issuer-overlap (Jaccard) — {best_date}")
    _network(jac_m, names, npth, f"Co-lending network — {best_date} (edge: issuer-Jaccard >= {NET_EDGE_JACCARD})")
    if not trend_long.empty:
        _trend_chart(trend_long, tp)

    _write(latest, latest_cols, fund_summary, trend_wide, pairs, best_date,
           int(funds_per_date.max()), hp, npth, tp if not trend_long.empty else None)
    print(f"wrote {WORKBOOK}  (best snapshot date {best_date})")


def _write(latest, latest_cols, fund_summary, trend_wide, pairs, best_date, n_funds,
           hp, npth, tp) -> None:
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    overview = [
        "BDC Portfolio Overlap — how much two funds own in common.",
        "",
        f"Best snapshot date (most funds present): {best_date}  ({n_funds} funds).",
        "Two grains: ISSUER = same borrower; ISSUE = same specific loan/tranche.",
        "",
        "Metrics (per fund pair):",
        "  Common      — # of shared borrowers / loans.",
        "  A in B / B in A — directional: share of A's book also in B, and vice versa (asymmetric).",
        "  Jaccard     — shared / combined (0-1), size-fair similarity.",
        "  overlap-coef— shared / smaller book (a small fund inside a big one -> ~1).",
        "  $-overlap   — Σ min(weight in A, weight in B): overlap weighted by position size.",
        "  lift        — actual common vs. expected if both drew names at random (>1 = more than chance).",
        "  Same mgr    — both funds share a parent manager (intra-manager co-investment is mechanical).",
        "",
        "Tabs: PairsLatest (best date), FundSummary (each fund's nearest neighbour + differentiation),",
        "Trend (top cross-manager pairs over time), PairsByDate (full panel), Charts (heatmap, network,",
        "trend). Heatmap/network are issuer-grain at the best snapshot date.",
        "",
        "Grains: issuer-grain uses all parsed holdings (any matched issuer_cluster). Issue-grain",
        "uses ONLY High/Medium-confidence matched issues — Low/Single (seniority-only) matches are",
        "excluded so the issue grain counts only genuinely-matched tranches.",
        "",
        "Caveats: overlap is on matched, priced holdings (equity/unmatched excluded); funds with",
        "different fiscal quarter-ends only co-appear on shared dates; issue-grain overlap inherits",
        "the matcher's confidence.",
    ]
    with pd.ExcelWriter(WORKBOOK, engine="openpyxl") as xl:
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Overview", index=False)
        latest[list(latest_cols)].rename(columns=latest_cols).to_excel(
            xl, sheet_name="PairsLatest", index=False)
        fund_summary.to_excel(xl, sheet_name="FundSummary", index=False)
        if not trend_wide.empty:
            trend_wide.to_excel(xl, sheet_name="Trend", index=False)
        # full panel (rename for readability, keep grain/date)
        panel_cols = ["reporting_date", "grain", "fund_a", "fund_b", "mgr_a", "mgr_b",
                      "same_manager", "n_a", "n_b", "common", "cov_a_in_b", "cov_b_in_a",
                      "jaccard", "overlap_coef", "wt_overlap", "lift"]
        pairs[panel_cols].to_excel(xl, sheet_name="PairsByDate", index=False)
        pd.DataFrame({"": []}).to_excel(xl, sheet_name="Charts", index=False)

        wb = xl.book
        ov = wb["Overview"]
        ov["A1"].font = Font(bold=True, size=14)
        for i, line in enumerate(overview, start=1):
            ov.cell(row=i, column=1, value=line)
        ov.column_dimensions["A"].width = 100

        hdr_fill, hdr_font = PatternFill("solid", fgColor="1F4E78"), Font(bold=True, color="FFFFFF")
        for name in ("PairsLatest", "FundSummary", "Trend", "PairsByDate"):
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill, cell.font = hdr_fill, hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for col, w in {1: 30, 2: 30, 3: 16, 4: 16}.items():
            wb["PairsLatest"].column_dimensions[get_column_letter(col)].width = w

        ch = wb["Charts"]
        ch["A1"] = "Issuer-overlap heatmap, co-lending network, and top-pair trend (best snapshot)"
        ch["A1"].font = Font(bold=True, size=12)
        row = 3
        for img in (hp, npth, tp):
            if img and Path(img).exists():
                ch.add_image(XLImage(str(img)), f"A{row}")
                row += 60


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--threshold", type=int, default=92, help="fuzzy merge threshold (Phase 2)")
    args = ap.parse_args()
    build(threshold=args.threshold)
